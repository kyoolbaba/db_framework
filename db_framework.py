import warnings
warnings.filterwarnings("ignore", category=SyntaxWarning)

import pandas as pd
import numpy as np
import json
import os
import webbrowser
from datetime import datetime

# ─────────────────────────────────────────────
#  HOW TO USE
# ─────────────────────────────────────────────
#
#  generate_dashboard(
#      dicts         = [dict1, dict2, dict3, dict4],
#      names         = ["Payment Summary", "Bank Wise", "PG Wise", "Merchant Wise"],
#      output        = "dashboard.html",
#      default_theme = "Dark Blue",       # Dark Blue | Carbon | Midnight Green | Slate Light | Crimson | Night
#      rows_per_page = 25,                # pagination default: 25 | 50 | 100
#      sparkline_col = None,              # column name to draw sparklines for (optional)
#      cf_cols       = None,              # list of column names for conditional formatting (optional)
#  )
#
#  load_and_generate(pkl_path="data.pkl", output="dashboard.html")
# ─────────────────────────────────────────────


# ── PYTHON HELPERS ────────────────────────────

def _round_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.select_dtypes(include=[np.number]).columns:
        df[col] = df[col].round(2)
    return df


def _build_thead_html(df: pd.DataFrame) -> str:
    cols = df.columns
    if not isinstance(cols, pd.MultiIndex):
        idx_label = df.index.name if df.index.name is not None else "Index"
        all_cols  = [idx_label] + list(cols)
        ths = "".join(f'<th class="sortable" data-col="{i}">{c}<span class="sort-icon">⇅</span></th>' for i, c in enumerate(all_cols))
        return f"<thead><tr>{ths}</tr></thead>"

    n_levels   = cols.nlevels
    col_tuples = list(cols)
    n_cols     = len(col_tuples)
    idx_label  = df.index.name if df.index.name is not None else "Index"

    thead = "<thead>"
    for level in range(n_levels):
        thead += "<tr>"
        if level == 0:
            thead += f'<th rowspan="{n_levels}" class="idx-th sortable" data-col="0">{idx_label}<span class="sort-icon">⇅</span></th>'
        i = 0
        labels = [t[level] if isinstance(t, tuple) else t for t in col_tuples]
        while i < n_cols:
            label = labels[i]
            span  = 1
            while i + span < n_cols and labels[i + span] == label:
                span += 1
            rowspan = 1
            if level < n_levels - 1:
                all_sub_empty = True
                for sub_level in range(level + 1, n_levels):
                    sub_labels = [col_tuples[i+k][sub_level] if isinstance(col_tuples[i+k], tuple) else "" for k in range(span)]
                    if any(str(s).strip() not in ("","nan") for s in sub_labels):
                        all_sub_empty = False; break
                if all_sub_empty:
                    rowspan = n_levels - level
            col_idx = i + 1
            cs = f' colspan="{span}"'   if span    > 1 else ""
            rs = f' rowspan="{rowspan}"' if rowspan > 1 else ""
            sort_cls = ' class="sortable"' if rowspan > 1 or level == n_levels - 1 else ""
            data_col = f' data-col="{col_idx}"' if (rowspan > 1 or level == n_levels - 1) else ""
            sort_icon = '<span class="sort-icon">⇅</span>' if (rowspan > 1 or level == n_levels - 1) else ""
            thead += f"<th{cs}{rs}{sort_cls}{data_col}>{label}{sort_icon}</th>"
            i += span
        thead += "</tr>"
    thead += "</thead>"
    return thead


def _build_summary_row(df: pd.DataFrame) -> str:
    """Auto summary row: sum for numeric, blank for non-numeric."""
    cells = ["<td class='summary-cell summary-idx'>∑ Summary</td>"]
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            total = df[col].sum()
            cells.append(f"<td class='summary-cell'>{round(total, 2)}</td>")
        else:
            cells.append("<td class='summary-cell'>—</td>")
    return f"<tr class='summary-row'>{''.join(cells)}</tr>"


def export_to_excel(dicts: list, names: list, output: str = "dashboard.xlsx"):
    """Export all DataFrames to a single Excel file with one sheet per table."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill  = PatternFill("solid", fgColor="1E2236")
    header_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin', color="2A2D3A"),
        right=Side(style='thin', color="2A2D3A"),
        bottom=Side(style='thin', color="2A2D3A")
    )

    def _add_sheet(wb, sheet_name, df):
        safe_name = sheet_name[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '(').replace(']', ')')
        ws = wb.create_sheet(title=safe_name)
        # Header row
        for ci, col in enumerate([df.index.name or 'Index'] + list(df.columns), 1):
            cell = ws.cell(row=1, column=ci, value=str(col))
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = thin_border
        # Data rows
        for ri, (idx, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=ri, column=1, value=idx)
            for ci, val in enumerate(row, 2):
                try:    ws.cell(row=ri, column=ci, value=float(val) if pd.api.types.is_numeric_dtype(type(val)) else str(val))
                except: ws.cell(row=ri, column=ci, value=str(val))
        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col_cells if cell.value), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

    def _process(d, prefix=''):
        for key, val in d.items():
            name = (prefix + ' › ' + key if prefix else key)
            if isinstance(val, pd.DataFrame):
                _add_sheet(wb, name, val)
            elif isinstance(val, dict):
                _process(val, name)

    for name, d in zip(names, dicts):
        _process(d, name)

    wb.save(output)
    abs_path = os.path.abspath(output)
    print(f"✅ Excel saved → {abs_path}")
    return abs_path


def _df_to_json_rows(df: pd.DataFrame, sparkline_col=None, cf_cols=None, description=None) -> dict:
    """Convert df to row data for JS rendering with metadata."""
    df = _round_df(df)

    # Flatten MultiIndex columns to readable strings
    def _flatten_col(c):
        if isinstance(c, tuple):
            return ' | '.join(str(x) for x in c if str(x) not in ('', 'nan'))
        return str(c)

    idx_label = df.index.name or "Index"
    flat_cols = [idx_label] + [_flatten_col(c) for c in df.columns]

    cf_cols  = cf_cols or []
    sp_col   = sparkline_col

    # Compute cf min/max per column
    cf_ranges = {}
    for col in df.columns:
        if col in cf_cols and pd.api.types.is_numeric_dtype(df[col]):
            cf_ranges[col] = {"min": float(df[col].min()), "max": float(df[col].max())}

    # Compute data bar max per column (all numeric)
    bar_ranges = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            bar_ranges[col] = float(df[col].abs().max()) or 1

    # Summary row
    summary = []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            summary.append(round(float(df[col].sum()), 2))
        else:
            summary.append(None)

    # Rows — sanitize values to prevent JS/JSON injection
    def _safe(val):
        import re as _re
        # Preserve period/date strings like 2021-01, 2021-01-01 BEFORE any conversion
        _s_raw = str(val)
        if _re.match(r'^\d{4}-\d{2}(-\d{2})?$', _s_raw.strip()):
            # Looks like YYYY-MM or YYYY-MM-DD — preserve exactly
            s = _s_raw.strip()
            s = s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            return s
        # Also preserve Pandas Period strings like 2021-01 (Period dtype)
        if hasattr(val, 'strftime'):
            try:
                # Period or Timestamp — format as YYYY-MM or YYYY-MM-DD
                fmt = val.strftime('%Y-%m') if hasattr(val, 'freq') else val.strftime('%Y-%m-%d')
                return fmt
            except: pass
        # Whole float → int (fixes 2024.0 → 2024)
        if isinstance(val, float) and val == int(val) and not (val != val):
            val = int(val)
        s = str(val)
        # Catch Timestamp/Period string representations like "2021-01-01 00:00:00"
        ts_match = _re.match(r'^(\d{4}-\d{2}-\d{2}) 00:00:00$', s)
        if ts_match:
            s = ts_match.group(1)  # strip time component if midnight
        # Replace NaN/None representations
        if s in ('nan', 'NaN', 'None', 'NaT'):
            return ''
        s = s.replace('&',  '&amp;')
        s = s.replace('<',  '&lt;')
        s = s.replace('>',  '&gt;')
        # Remove control characters that break JSON
        s = ''.join(c for c in s if ord(c) >= 32 or c in '\t\n\r')
        return s

    rows = []
    for idx, row in df.iterrows():
        cells = [_safe(idx)]
        for col in df.columns:
            cells.append(_safe(row[col]))
        rows.append(cells)

    # Compute column totals for % of Total
    col_totals = {}
    for ci, col in enumerate(df.columns):
        if pd.api.types.is_numeric_dtype(df.iloc[:, ci]):
            tot = float(df.iloc[:, ci].sum())
            col_totals[str(ci + 1)] = tot if tot != 0 else 1

    return {
        "description": description or "",
        "cols":        flat_cols,
        "rows":        rows,
        "summary":     summary,
        "cf_cols":     cf_cols,
        "cf_ranges":   cf_ranges,
        "bar_ranges":  {str(i+1): v for i, (col, v) in enumerate(bar_ranges.items())},
        "col_totals":  col_totals,
        "sp_col":      sp_col,
        "numeric_cols": [i for i, col in enumerate(flat_cols) if i > 0 and pd.api.types.is_numeric_dtype(df.iloc[:, i-1])],
    }


def _serialize_section(key, value, sparkline_col=None, cf_cols=None):
    """Recursively serialize any depth of nested dicts → DataFrames."""
    if isinstance(value, pd.DataFrame):
        return {"type": 1, "title": key, "data": _df_to_json_rows(value, sparkline_col, cf_cols)}
    elif isinstance(value, dict):
        children = [
            s for s in (
                _serialize_section(sk, sv, sparkline_col, cf_cols)
                for sk, sv in value.items()
            ) if s is not None
        ]
        return {"type": 2, "title": key, "children": children}
    return None


def _serialize_dict(d, sparkline_col=None, cf_cols=None):
    return [s for s in (_serialize_section(k, v, sparkline_col, cf_cols) for k, v in d.items()) if s]


# ── MAIN FUNCTION ─────────────────────────────

def generate_dashboard(
    dicts: list,
    names: list,
    output: str = "dashboard.html",
    default_theme: str = "Dark Blue",
    rows_per_page: int = 25,
    sparkline_col: str = None,
    cf_cols: list = None,
    title: str = "Dashboard",
    chart_max_rows: int = 5000,
    colorby_max_unique: int = 20,
    embed_plotly: bool = False,
):
    assert len(dicts) == len(names), "dicts and names must have same length!"

    _title = str(title).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    # ── PLOTLY: CDN by default, or inline for true offline (embed_plotly=True) ──
    plotly_tag = '<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>'
    if embed_plotly:
        _code = None
        _cands = ['plotly.min.js',
                  os.path.join(os.path.dirname(os.path.abspath(output)) or '.', 'plotly.min.js')]
        for _c in _cands:
            try:
                if _c and os.path.exists(_c):
                    with open(_c, 'r', encoding='utf-8') as _pf:
                        _code = _pf.read()
                    print(f"📦 Embedding Plotly from {os.path.abspath(_c)}")
                    break
            except Exception:
                pass
        if _code is None:
            try:
                import urllib.request
                print("⏳ embed_plotly: downloading Plotly (~3.5MB) to inline it...")
                _code = urllib.request.urlopen(
                    "https://cdn.plot.ly/plotly-2.27.0.min.js", timeout=60).read().decode('utf-8')
            except Exception as _e:
                print(f"⚠ embed_plotly: could not fetch Plotly ({_e}); falling back to CDN.")
        if _code:
            # Guard against an accidental closing tag inside the minified bundle
            _code = _code.replace('</script>', '<\\/script>')
            plotly_tag = '<scr' + 'ipt>' + _code + '</scr' + 'ipt>'

    # Sanitize names to remove any problematic characters
    safe_names    = [str(n).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;') for n in names]
    all_sections  = [_serialize_dict(d, sparkline_col, cf_cols) for d in dicts]
    sections_json = json.dumps(all_sections, ensure_ascii=True)
    names_json    = json.dumps(safe_names,   ensure_ascii=True)
    theme_json    = json.dumps(default_theme)
    rpp_json      = json.dumps(rows_per_page)
    chart_max_rows_json     = json.dumps(int(chart_max_rows))
    colorby_max_unique_json = json.dumps(int(colorby_max_unique))
    timestamp     = datetime.now().strftime("%d %b %Y, %I:%M %p")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{_title}</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet" onerror="this.remove()"/>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
</style>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'DM Sans',sans-serif;min-height:100vh;transition:background 0.2s,color 0.2s;overflow-x:hidden;}}

/* ── THEMES ── */
body.theme-darkblue  {{--bg:#0e1017;--nav:#13161f;--nav-border:#2a2d3a;--accent:#4f6ef7;--accent-text:#a0aaff;--text1:#ffffff;--text2:#a0a8c0;--text3:#666d88;--row-border:#252a3a;--td:#ffffff;--th:#8892b0;--tbl-border:#2a2d3a;--tab-bg:#1e2236;--tab-border:#2d3560;--surface:#13161f;--row-even:#161a28;--row-odd:#0e1017;--logo-bg:#4f6ef7;--logo-color:#fff;--badge1-bg:rgba(79,110,247,0.15);--badge1-color:#4f6ef7;--badge2-bg:rgba(34,211,165,0.15);--badge2-color:#22d3a5;--toggle-bg:#0e1017;--row-hover:#1e2a4a;--sidebar-bg:#0a0d14;--input-bg:#1a1e2e;--summary-bg:#1a2240;--summary-color:#a0aaff;}}
body.theme-carbon    {{--bg:#111111;--nav:#1a1a1a;--nav-border:#2e2e2e;--accent:#e0e0e0;--accent-text:#ffffff;--text1:#ffffff;--text2:#bbbbbb;--text3:#777777;--row-border:#2a2a2a;--td:#ffffff;--th:#999999;--tbl-border:#2e2e2e;--tab-bg:#222222;--tab-border:#3a3a3a;--surface:#1a1a1a;--row-even:#1a1a1a;--row-odd:#111111;--logo-bg:#e0e0e0;--logo-color:#111;--badge1-bg:rgba(200,200,200,0.15);--badge1-color:#ccc;--badge2-bg:rgba(150,150,150,0.15);--badge2-color:#aaa;--toggle-bg:#111;--row-hover:#2a2a2a;--sidebar-bg:#0a0a0a;--input-bg:#222;--summary-bg:#222;--summary-color:#fff;}}
body.theme-midnight  {{--bg:#0b1512;--nav:#0f1e18;--nav-border:#1a3028;--accent:#22d3a5;--accent-text:#22d3a5;--text1:#ffffff;--text2:#90c8b8;--text3:#4a8070;--row-border:#1a3028;--td:#ffffff;--th:#55aa90;--tbl-border:#1a3028;--tab-bg:#162e26;--tab-border:#1f4035;--surface:#0f1e18;--row-even:#0f1e18;--row-odd:#0b1512;--logo-bg:#22d3a5;--logo-color:#0b1512;--badge1-bg:rgba(34,211,165,0.12);--badge1-color:#22d3a5;--badge2-bg:rgba(34,211,165,0.08);--badge2-color:#1aad89;--toggle-bg:#0b1512;--row-hover:#153525;--sidebar-bg:#070e0b;--input-bg:#102018;--summary-bg:#102a20;--summary-color:#22d3a5;}}
body.theme-slate     {{--bg:#f4f5f7;--nav:#ffffff;--nav-border:#e2e5ec;--accent:#3b5bdb;--accent-text:#3b5bdb;--text1:#060810;--text2:#3a4466;--text3:#7a88aa;--row-border:#d5daea;--td:#060810;--th:#5566aa;--tbl-border:#c8d0e8;--tab-bg:#eef1fc;--tab-border:#c5cde8;--surface:#ffffff;--row-even:#ffffff;--row-odd:#f0f2fa;--logo-bg:#3b5bdb;--logo-color:#fff;--badge1-bg:rgba(59,91,219,0.1);--badge1-color:#3b5bdb;--badge2-bg:rgba(59,91,219,0.06);--badge2-color:#6680e0;--toggle-bg:#e8eaf0;--row-hover:#dde4f8;--sidebar-bg:#e8eaf2;--input-bg:#eef1fc;--summary-bg:#e0e8ff;--summary-color:#3b5bdb;}}
body.theme-crimson   {{--bg:#110c0e;--nav:#1a1014;--nav-border:#2e1820;--accent:#f75a7a;--accent-text:#f75a7a;--text1:#ffffff;--text2:#d4a0b0;--text3:#7a4858;--row-border:#2e1820;--td:#ffffff;--th:#bb6070;--tbl-border:#2e1820;--tab-bg:#221018;--tab-border:#3e1828;--surface:#1a1014;--row-even:#1a1014;--row-odd:#110c0e;--logo-bg:#f75a7a;--logo-color:#fff;--badge1-bg:rgba(247,90,122,0.12);--badge1-color:#f75a7a;--badge2-bg:rgba(247,90,122,0.08);--badge2-color:#e04868;--toggle-bg:#110c0e;--row-hover:#2a1020;--sidebar-bg:#0a0608;--input-bg:#1e0e14;--summary-bg:#220a12;--summary-color:#f75a7a;}}
body.theme-night     {{--bg:#000000;--nav:#080808;--nav-border:#1c1c1c;--accent:#ffffff;--accent-text:#ffffff;--text1:#ffffff;--text2:#cccccc;--text3:#666666;--row-border:#1c1c1c;--td:#ffffff;--th:#666666;--tbl-border:#1c1c1c;--tab-bg:#111111;--tab-border:#333333;--surface:#080808;--row-even:#080808;--row-odd:#000000;--logo-bg:#ffffff;--logo-color:#000;--badge1-bg:rgba(255,255,255,0.1);--badge1-color:#ffffff;--badge2-bg:rgba(255,255,255,0.06);--badge2-color:#cccccc;--toggle-bg:#000000;--row-hover:#1a1a1a;--sidebar-bg:#000;--input-bg:#111;--summary-bg:#111;--summary-color:#fff;}}

/* ── LAYOUT ── */
.app-layout{{display:flex;height:100vh;overflow:hidden;}}
.sidebar{{width:240px;min-width:240px;background:var(--sidebar-bg);border-right:0.5px solid var(--nav-border);display:flex;flex-direction:column;transition:width 0.2s,min-width 0.2s;overflow:hidden;}}
.sidebar.collapsed{{width:0;min-width:0;}}
.main-area{{flex:1;display:flex;flex-direction:column;overflow:hidden;}}

/* ── SIDEBAR ── */
.sidebar-header{{padding:14px 16px 10px;border-bottom:0.5px solid var(--nav-border);display:flex;align-items:center;justify-content:space-between;}}
.sidebar-title{{font-size:11px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.6px;white-space:nowrap;}}
.sidebar-body{{flex:1;overflow-y:auto;padding:8px 0;}}
.sidebar-dict-group{{margin-bottom:4px;}}
.sidebar-dict-label{{padding:6px 16px;font-size:11px;font-weight:600;color:var(--accent);text-transform:uppercase;letter-spacing:0.4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.sidebar-item{{padding:5px 16px 5px 24px;font-size:12px;color:var(--text2);cursor:pointer;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;transition:all 0.1s;border-left:2px solid transparent;}}
.sidebar-item:hover{{color:var(--text1);background:var(--tab-bg);}}
.sidebar-item.active{{color:var(--accent-text);border-left-color:var(--accent);background:var(--tab-bg);}}
.sidebar-body::-webkit-scrollbar{{width:4px;}}
.sidebar-body::-webkit-scrollbar-track{{background:transparent;}}
.sidebar-body::-webkit-scrollbar-thumb{{background:var(--nav-border);border-radius:2px;}}

/* ── TOPNAV ── */
.topnav{{height:52px;display:flex;align-items:center;justify-content:space-between;padding:0 20px;background:var(--nav);border-bottom:0.5px solid var(--nav-border);flex-shrink:0;gap:12px;}}
.topnav-left{{display:flex;align-items:center;gap:12px;}}
.sidebar-toggle{{width:32px;height:32px;border-radius:6px;border:0.5px solid var(--nav-border);background:transparent;color:var(--text2);cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:16px;transition:all 0.15s;}}
.sidebar-toggle:hover{{background:var(--tab-bg);color:var(--text1);}}
.brand{{display:flex;align-items:center;gap:8px;font-size:15px;font-weight:600;color:var(--text1);letter-spacing:-0.3px;}}
.logo{{width:28px;height:28px;border-radius:6px;background:var(--logo-bg);color:var(--logo-color);display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;}}
.brand-accent{{color:var(--accent);}}
.topnav-right{{display:flex;align-items:center;gap:8px;}}
.timestamp{{font-size:11px;color:var(--text3);white-space:nowrap;}}

/* ── GLOBAL SEARCH ── */
.global-search-wrap{{position:relative;}}
.global-search{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:6px 12px 6px 30px;font-size:12px;color:var(--text1);font-family:'DM Sans',sans-serif;width:220px;outline:none;transition:border-color 0.15s;}}
.global-search:focus{{border-color:var(--accent);}}
.global-search::placeholder{{color:var(--text3);}}
.global-search-icon{{position:absolute;left:9px;top:50%;transform:translateY(-50%);color:var(--text3);font-size:13px;pointer-events:none;}}
.global-results{{position:absolute;top:calc(100% + 6px);left:0;right:0;background:var(--nav);border:0.5px solid var(--nav-border);border-radius:8px;z-index:200;max-height:300px;overflow-y:auto;box-shadow:0 8px 24px rgba(0,0,0,0.4);display:none;}}
.global-results.show{{display:block;}}
.global-result-item{{padding:8px 14px;font-size:12px;color:var(--text2);cursor:pointer;border-bottom:0.5px solid var(--nav-border);}}
.global-result-item:last-child{{border-bottom:none;}}
.global-result-item:hover{{background:var(--tab-bg);color:var(--text1);}}
.global-result-path{{font-size:10px;color:var(--text3);margin-top:2px;}}
.global-result-match{{color:var(--accent);font-weight:600;}}
.no-results{{padding:12px 14px;font-size:12px;color:var(--text3);text-align:center;}}

/* ── TOOLBAR ── */
.toolbar{{display:flex;align-items:center;justify-content:space-between;padding:8px 20px;background:var(--nav);border-bottom:0.5px solid var(--nav-border);flex-shrink:0;flex-wrap:wrap;gap:8px;}}
.toolbar-left{{display:flex;align-items:center;gap:8px;flex-wrap:wrap;}}
.toolbar-label{{font-size:11px;color:var(--text3);font-weight:500;text-transform:uppercase;letter-spacing:0.5px;}}
.toggle-wrap{{display:flex;background:var(--toggle-bg);border:0.5px solid var(--nav-border);border-radius:6px;overflow:hidden;}}
.toggle-btn{{padding:5px 12px;font-size:11px;font-weight:500;cursor:pointer;color:var(--text3);background:transparent;border:none;font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.toggle-btn.active{{background:var(--tab-bg);color:var(--accent-text);}}
.theme-switcher{{display:flex;align-items:center;gap:6px;}}
.theme-dot{{width:16px;height:16px;border-radius:50%;cursor:pointer;border:2px solid transparent;transition:transform 0.15s;}}
.theme-dot:hover{{transform:scale(1.2);}}
.theme-dot.active{{border-color:var(--text1);}}
.divider{{width:0.5px;height:20px;background:var(--nav-border);}}

/* ── NAV PILLS ── */
.nav-pills-bar{{padding:0 20px;background:var(--nav);border-bottom:0.5px solid var(--nav-border);display:flex;gap:3px;overflow-x:auto;flex-shrink:0;}}
.nav-pill{{padding:8px 16px;font-size:12px;font-weight:500;cursor:pointer;border-bottom:2px solid transparent;color:var(--text3);white-space:nowrap;transition:all 0.15s;}}
.nav-pill:hover{{color:var(--text2);}}
.nav-pill.active{{color:var(--accent);border-bottom-color:var(--accent);}}

/* ── SCROLL AREA ── */
.scroll-area{{flex:1;overflow-y:auto;padding:20px;}}
.scroll-area::-webkit-scrollbar{{width:6px;}}
.scroll-area::-webkit-scrollbar-track{{background:transparent;}}
.scroll-area::-webkit-scrollbar-thumb{{background:var(--nav-border);border-radius:3px;}}

/* ── SECTION ── */
.section{{margin-bottom:28px;}}
.section-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;flex-wrap:wrap;gap:8px;}}
.section-title{{font-size:14px;font-weight:700;color:var(--text1);display:flex;align-items:center;gap:8px;letter-spacing:-0.2px;padding:8px 12px;background:var(--surface);border-radius:6px;margin-bottom:10px;border-left:4px solid var(--accent);}}
body.theme-slate .section-title{{color:#0a0e1f;background:#f0f2fa;}}
.type-badge{{font-size:9px;padding:2px 8px;border-radius:3px;font-weight:700;letter-spacing:0.5px;text-transform:uppercase;background:#000000;color:#ffffff;border:none;}}
.t1-badge{{background:var(--badge1-bg);color:var(--badge1-color);}}
.t2-badge{{background:var(--badge2-bg);color:var(--badge2-color);}}
.t1-badge{{background:var(--badge1-bg);color:var(--badge1-color);}}
.t2-badge{{background:var(--badge2-bg);color:var(--badge2-color);}}
.section-actions{{display:flex;align-items:center;gap:6px;}}
.action-btn{{padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--tab-bg);color:var(--text2);font-family:'DM Sans',sans-serif;transition:all 0.15s;white-space:nowrap;}}
.action-btn:hover{{color:var(--text1);border-color:var(--accent);}}
.action-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent);}}

/* ── TABLE CONTROLS ── */
.tbl-controls{{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px;gap:8px;flex-wrap:wrap;}}
.tbl-search-wrap{{position:relative;flex:1;max-width:280px;}}
.tbl-search{{width:100%;background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:5px 10px 5px 28px;font-size:12px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;transition:border-color 0.15s;}}
.tbl-search:focus{{border-color:var(--accent);}}
.tbl-search::placeholder{{color:var(--text3);}}
.tbl-search-icon{{position:absolute;left:9px;top:50%;transform:translateY(-50%);color:var(--text3);font-size:12px;}}
.tbl-right-controls{{display:flex;align-items:center;gap:6px;}}
.rpp-select{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:5px;padding:4px 8px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;cursor:pointer;}}
.col-toggle-wrap{{position:relative;}}
.col-toggle-btn{{padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--tab-bg);color:var(--text2);font-family:'DM Sans',sans-serif;}}
.col-toggle-btn:hover{{color:var(--text1);}}
.col-toggle-menu{{position:absolute;top:calc(100%+4px);right:0;background:var(--nav);border:0.5px solid var(--nav-border);border-radius:8px;z-index:50;min-width:160px;padding:6px 0;box-shadow:0 8px 24px rgba(0,0,0,0.4);display:none;}}
.col-toggle-menu.show{{display:block;}}
.col-toggle-item{{display:flex;align-items:center;gap:8px;padding:6px 14px;font-size:12px;color:var(--text2);cursor:pointer;}}
.col-toggle-item:hover{{background:var(--tab-bg);color:var(--text1);}}
.col-toggle-item input{{accent-color:var(--accent);cursor:pointer;}}

/* ── TABLE ── */
.tbl-outer{{border-radius:8px;border:0.5px solid var(--tbl-border);overflow:hidden;}}
.tbl-wrap{{overflow-x:auto;overflow-y:auto;max-height:70vh;}}
.tbl-wrap::-webkit-scrollbar{{width:6px;height:6px;}}
.tbl-wrap::-webkit-scrollbar-track{{background:transparent;}}
.tbl-wrap::-webkit-scrollbar-thumb{{background:var(--nav-border);border-radius:3px;}}
table{{border-collapse:collapse;font-size:12px;min-width:max-content;width:100%;}}
thead{{background:var(--surface);}}
thead th{{padding:9px 14px;text-align:center;font-size:10px;font-weight:600;color:var(--th);text-transform:uppercase;letter-spacing:0.6px;border-bottom:0.5px solid var(--tbl-border);border-right:0.5px solid var(--tbl-border);white-space:nowrap;vertical-align:middle;background:var(--surface);position:sticky;top:0;z-index:3;cursor:pointer;user-select:none;}}
thead th.idx-th{{text-align:left;position:sticky;left:0;z-index:4;}}
thead th:last-child{{border-right:none;}}
thead th:hover{{color:var(--text1);}}
.sort-icon{{margin-left:4px;opacity:0.4;font-size:10px;}}
thead th.sort-asc .sort-icon::after{{content:'↑';opacity:1;}}
thead th.sort-desc .sort-icon::after{{content:'↓';opacity:1;}}
thead th.sort-asc .sort-icon,thead th.sort-desc .sort-icon{{opacity:1;color:var(--accent);}}

tbody tr:nth-child(even){{background:var(--row-even);}}
tbody tr:nth-child(odd){{background:var(--row-odd);}}
tbody tr{{border-bottom:0.5px solid var(--row-border);transition:background 0.1s;}}
tbody tr:last-child{{border-bottom:none;}}
tbody tr:hover{{background:var(--row-hover)!important;}}
tbody td{{padding:9px 14px;color:var(--td);border-right:0.5px solid var(--row-border);white-space:nowrap;font-weight:500;position:relative;}}
tbody td:last-child{{border-right:none;}}
tbody td:first-child{{color:var(--text1);font-weight:700;text-align:left;position:sticky;left:0;z-index:2;background:inherit;border-right:1px solid var(--tbl-border);}}
tbody td:not(:first-child){{font-family:'DM Mono',monospace;font-size:11.5px;text-align:right;}}

/* ── SUMMARY ROW ── */
.summary-row{{background:var(--summary-bg)!important;border-top:1px solid var(--tbl-border)!important;position:sticky;bottom:0;}}
.summary-cell{{color:var(--summary-color)!important;font-weight:700!important;font-family:'DM Mono',monospace;font-size:11.5px;text-align:right;padding:8px 14px!important;}}
.summary-idx{{text-align:left!important;font-family:'DM Sans',sans-serif!important;color:var(--summary-color)!important;position:sticky;left:0;background:var(--summary-bg);}}
/* ── GROUP SUBTOTAL ROW ── */
.subtotal-row{{background:var(--summary-bg);opacity:0.96;}}
.subtotal-cell{{color:var(--text2);font-weight:600;font-family:'DM Mono',monospace;font-size:11px;text-align:right;padding:6px 14px;border-top:1px dashed var(--tbl-border);border-bottom:1px dashed var(--tbl-border);}}
.subtotal-idx{{text-align:left;font-family:'DM Sans',sans-serif;color:var(--accent);font-weight:700;position:sticky;left:0;background:var(--summary-bg);}}
.subtotal-lvl-0 .subtotal-cell{{font-size:11.5px;font-weight:700;border-top:1px solid var(--tbl-border);}}

/* ── DATA BAR ── */
.data-bar-bg{{position:absolute;left:0;top:0;height:100%;opacity:0.18;pointer-events:none;border-radius:0;}}
.data-bar-pos{{background:var(--accent);}}
.data-bar-neg{{background:#f75a7a;}}

/* ── CONDITIONAL FORMAT ── */
.cf-high{{color:#22d3a5!important;font-weight:700!important;}}
.cf-low{{color:#f75a7a!important;font-weight:700!important;}}
.cf-mid{{color:#f7b24f!important;}}

/* ── SPARKLINE ── */
.sparkline-cell{{min-width:80px;}}
canvas.spark{{display:block;}}

/* ── PAGINATION ── */
.pagination{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;background:var(--surface);border-top:0.5px solid var(--tbl-border);flex-wrap:wrap;gap:8px;}}
.page-info{{font-size:11px;color:var(--text1);font-weight:500;}}
.page-btns{{display:flex;gap:3px;}}
.page-btn{{padding:3px 9px;border-radius:5px;font-size:11px;font-weight:600;cursor:pointer;border:0.5px solid var(--tbl-border);background:var(--tab-bg);color:var(--text1);font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.page-btn:hover{{color:#fff;background:var(--accent);border-color:var(--accent);}}
.page-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent);}}
.page-btn:disabled{{opacity:0.3;cursor:not-allowed;}}

/* ── ACCORDION ── */
.acc-item{{border:0.5px solid var(--tbl-border);border-radius:8px;margin-bottom:6px;overflow:hidden;}}
.acc-head{{padding:10px 16px;background:var(--surface);display:flex;align-items:center;justify-content:space-between;cursor:pointer;font-size:12px;font-weight:500;color:var(--text1);transition:background 0.1s;user-select:none;}}
.acc-head:hover{{background:var(--tab-bg);}}
.acc-arrow{{font-size:10px;color:var(--text3);display:inline-block;transition:transform 0.2s;}}
.acc-arrow.open{{transform:rotate(90deg);}}
.acc-body{{display:none;}}
.acc-body.open{{display:block;}}
.acc-body-inner{{padding:12px 16px 16px;}}

/* ── COLLAPSED ── */
.coll-item{{border:0.5px solid var(--tbl-border);border-radius:8px;margin-bottom:8px;overflow:hidden;}}
.coll-head{{padding:11px 16px;background:var(--surface);display:flex;align-items:center;justify-content:space-between;cursor:pointer;transition:background 0.1s;user-select:none;}}
.coll-head:hover{{background:var(--tab-bg);}}
.coll-title{{font-size:13px;font-weight:600;color:var(--text1);display:flex;align-items:center;gap:8px;}}
.coll-arrow{{font-size:10px;color:var(--text3);display:inline-block;transition:transform 0.2s;}}
.coll-arrow.open{{transform:rotate(90deg);}}
.coll-body{{display:none;}}
.coll-body.open{{display:block;}}
.coll-body-inner{{padding:12px 16px 16px;}}
.coll-group-label{{font-size:12px;font-weight:600;color:var(--text1);margin-bottom:10px;padding-top:4px;display:flex;align-items:center;gap:6px;}}

/* ── SUB LABEL ── */
.sub-block{{margin-bottom:14px;}}
.sub-label{{font-size:11px;font-weight:600;color:var(--accent-text);margin-bottom:6px;padding-left:10px;border-left:2px solid var(--tab-border);}}

/* ── FULLSCREEN ── */
.fullscreen-overlay{{display:none;position:fixed;inset:0;z-index:1000;background:var(--bg);flex-direction:column;padding:20px;}}
.fullscreen-overlay.show{{display:flex;}}
.fullscreen-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;}}
.fullscreen-title{{font-size:16px;font-weight:600;color:var(--text1);}}
.fullscreen-close{{padding:6px 14px;border-radius:6px;border:0.5px solid var(--nav-border);background:var(--tab-bg);color:var(--text2);cursor:pointer;font-family:'DM Sans',sans-serif;font-size:12px;}}
.fullscreen-close:hover{{color:var(--text1);}}
.fullscreen-body{{flex:1;overflow:hidden;display:flex;flex-direction:column;}}
.fullscreen-body .tbl-wrap{{max-height:none;flex:1;}}

/* ── TOAST ── */
.toast{{position:fixed;bottom:24px;right:24px;background:var(--accent);color:#fff;padding:10px 18px;border-radius:8px;font-size:12px;font-weight:500;z-index:2000;opacity:0;transform:translateY(10px);transition:all 0.3s;pointer-events:none;}}
.toast.show{{opacity:1;transform:translateY(0);}}

/* ── CHART BUILDER ── */
.chart-builder{{margin-top:12px;background:var(--surface);border:0.5px solid var(--tbl-border);border-radius:10px;overflow:hidden;}}
.chart-builder-header{{padding:10px 16px;border-bottom:0.5px solid var(--tbl-border);display:flex;align-items:center;justify-content:space-between;}}
.chart-builder-title{{font-size:12px;font-weight:600;color:var(--text1);display:flex;align-items:center;gap:6px;}}
.chart-builder-body{{padding:16px;}}
.chart-builder-body.collapsed{{display:none;}}
.chart-edit-btn{{padding:3px 10px;border-radius:5px;font-size:11px;cursor:pointer;border:0.5px solid var(--accent);background:var(--tab-bg);color:var(--accent);font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.chart-edit-btn:hover{{background:var(--accent);color:#fff;}}
.chart-type-row{{display:flex;align-items:center;gap:6px;margin-bottom:14px;flex-wrap:wrap;}}
.chart-type-label{{font-size:11px;color:var(--text3);font-weight:500;text-transform:uppercase;letter-spacing:0.5px;margin-right:4px;}}
.chart-type-btn{{padding:5px 12px;border-radius:6px;font-size:11px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--toggle-bg);color:var(--text2);font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.chart-type-btn:hover{{color:var(--text1);border-color:var(--accent);}}
.chart-type-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent);}}
.chart-axes-row{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:14px;}}
.chart-axis-group{{display:flex;flex-direction:column;gap:6px;}}
.chart-axis-label{{font-size:11px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;}}
.chart-axis-select{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:6px 10px;font-size:12px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;width:100%;}}
.chart-axis-select:focus{{border-color:var(--accent);}}
.chart-y-checks{{display:flex;flex-direction:column;gap:4px;max-height:120px;overflow-y:auto;background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:6px 10px;}}
.chart-y-check{{display:flex;align-items:center;gap:7px;font-size:12px;color:var(--text2);cursor:pointer;padding:2px 0;}}
.chart-y-check:hover{{color:var(--text1);}}
.chart-y-check input{{accent-color:var(--accent);cursor:pointer;}}
.chart-generate-row{{display:flex;align-items:center;gap:8px;flex-wrap:wrap;}}
.chart-generate-btn{{padding:7px 20px;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;border:none;background:var(--accent);color:#fff;font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.chart-generate-btn:hover{{opacity:0.85;}}
.chart-output{{margin-top:12px;border-radius:8px;overflow:hidden;border:0.5px solid var(--tbl-border);background:var(--bg);min-height:400px;}}
.chart-placeholder{{display:flex;align-items:center;justify-content:center;height:400px;color:var(--text3);font-size:13px;}}
.chart-error{{color:#f75a7a;font-size:12px;padding:8px 12px;background:rgba(247,90,122,0.1);border-radius:6px;}}
.chart-label-toggle{{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--text2);cursor:pointer;padding:5px 10px;border-radius:6px;border:0.5px solid var(--nav-border);background:var(--toggle-bg);font-family:'DM Sans',sans-serif;user-select:none;}}
.chart-label-toggle.active{{background:var(--tab-bg);color:var(--accent-text);border-color:var(--accent);}}
.chart-title-row{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px;}}
.chart-title-input{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:7px 10px;font-size:12px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;width:100%;transition:border-color 0.15s;}}
.chart-title-input:focus{{border-color:var(--accent);}}
.chart-title-input::placeholder{{color:var(--text3);}}
.chart-title-label{{font-size:11px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;}}
.chart-axis-label-row{{display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap;}}
.chart-axis-label-group{{display:flex;flex-direction:column;flex:1;min-width:120px;}}
.chart-axis-label-tag{{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:3px;}}
.chart-axis-label-input{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:5px;padding:5px 9px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;width:100%;box-sizing:border-box;transition:border-color 0.15s;}}
.chart-axis-label-input:focus{{border-color:var(--accent);}}
.chart-axis-label-input::placeholder{{color:var(--text3);font-style:italic;}}
.chart-rendered-title{{font-size:14px;font-weight:600;color:var(--text1);margin:12px 0 4px;letter-spacing:-0.2px;}}
.chart-rendered-desc{{font-size:12px;color:var(--text2);margin-bottom:8px;line-height:1.5;}}
.add-chart-btn{{margin-top:10px;width:100%;padding:8px;border-radius:8px;border:1px dashed var(--tbl-border);background:transparent;color:var(--text3);font-size:12px;font-weight:500;cursor:pointer;font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.add-chart-btn:hover{{color:var(--accent);border-color:var(--accent);background:var(--tab-bg);}}
.chart-instance{{position:relative;}}
.chart-instance-remove{{position:absolute;top:10px;right:10px;padding:3px 10px;border-radius:5px;font-size:11px;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--surface);color:var(--text3);font-family:'DM Sans',sans-serif;z-index:10;}}
.chart-instance-remove:hover{{color:#f75a7a;border-color:#f75a7a;}}
.charts-container{{margin-top:8px;}}

/* ── CHART BUILDER v2 CONTROLS ── */
.chart-options-row{{display:flex;align-items:center;gap:6px;margin-bottom:12px;flex-wrap:wrap;}}
.chart-options-row .chart-type-label{{margin-right:2px;}}
.chart-opt-group{{display:flex;align-items:center;gap:4px;background:var(--toggle-bg);border:0.5px solid var(--nav-border);border-radius:6px;overflow:hidden;}}
.chart-opt-btn{{padding:4px 10px;font-size:11px;font-weight:500;cursor:pointer;color:var(--text3);background:transparent;border:none;font-family:'DM Sans',sans-serif;transition:all 0.15s;white-space:nowrap;}}
.chart-opt-btn:hover{{color:var(--text2);}}
.chart-opt-btn.active{{background:var(--tab-bg);color:var(--accent-text);}}
.chart-opt-select{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:5px;padding:4px 8px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;cursor:pointer;}}
.chart-opt-select:focus{{border-color:var(--accent);}}
.chart-opt-toggle{{padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--toggle-bg);color:var(--text3);font-family:'DM Sans',sans-serif;transition:all 0.15s;white-space:nowrap;}}
.chart-opt-toggle:hover{{color:var(--text2);}}
.chart-opt-toggle.active{{background:var(--tab-bg);color:var(--accent-text);border-color:var(--accent);}}
.chart-divider{{width:0.5px;height:18px;background:var(--nav-border);margin:0 4px;}}
/* ── COLUMN FILTERS ── */
.col-filter-icon{{display:inline-flex;align-items:center;justify-content:center;width:14px;height:14px;border-radius:3px;cursor:pointer;margin-left:4px;font-size:9px;color:var(--text3);transition:all 0.15s;vertical-align:middle;}}
.col-filter-icon:hover{{color:var(--accent);}}
.col-filter-icon.active{{color:var(--accent);background:var(--tab-bg);}}
.filter-dropdown{{position:fixed;z-index:9999;background:var(--surface);border:1px solid var(--nav-border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,0.35);min-width:200px;max-width:280px;overflow:hidden;}}
.filter-dropdown-header{{padding:8px 12px;border-bottom:0.5px solid var(--nav-border);font-size:11px;font-weight:600;color:var(--text2);display:flex;align-items:center;justify-content:space-between;}}
.filter-dropdown-body{{max-height:220px;overflow-y:auto;padding:6px 0;}}
.filter-dropdown-footer{{padding:6px 10px;border-top:0.5px solid var(--nav-border);display:flex;gap:6px;justify-content:flex-end;}}
.filter-cat-item{{display:flex;align-items:center;gap:7px;padding:4px 12px;cursor:pointer;font-size:11px;color:var(--text2);transition:background 0.1s;}}
.filter-cat-item:hover{{background:var(--toggle-bg);}}
.filter-cat-item input{{cursor:pointer;accent-color:var(--accent);}}
.filter-num-row{{padding:6px 12px;display:flex;align-items:center;gap:6px;}}
.filter-num-select{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:4px;padding:3px 6px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;}}
.filter-num-input{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:4px;padding:3px 8px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;outline:none;width:90px;}}
.filter-num-input:focus{{border-color:var(--accent);}}
.filter-action-btn{{padding:3px 10px;border-radius:4px;font-size:10px;font-weight:600;cursor:pointer;border:0.5px solid var(--nav-border);font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.filter-apply-btn{{background:var(--accent);color:var(--accent-text);border-color:var(--accent);}}
.filter-apply-btn:hover{{opacity:0.85;}}
.filter-clear-btn{{background:var(--toggle-bg);color:var(--text2);}}
.filter-clear-btn:hover{{color:var(--text1);}}
.filter-search-input{{margin:6px 10px;padding:4px 8px;border-radius:4px;border:0.5px solid var(--nav-border);background:var(--input-bg);color:var(--text1);font-size:11px;font-family:'DM Sans',sans-serif;outline:none;width:calc(100% - 20px);box-sizing:border-box;}}
.filter-search-input:focus{{border-color:var(--accent);}}

/* ── CHART FILTER ROW ── */
.chart-filter-row{{display:flex;gap:10px;margin-bottom:12px;align-items:flex-start;flex-wrap:wrap;}}
.chart-filter-group{{display:flex;flex-direction:column;flex:1;min-width:140px;position:relative;}}
.chart-filter-tag{{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:3px;}}
.chart-filter-trigger{{background:var(--input-bg);border:0.5px solid var(--nav-border);border-radius:5px;padding:5px 9px;font-size:11px;color:var(--text1);font-family:'DM Sans',sans-serif;cursor:pointer;display:flex;align-items:center;justify-content:space-between;gap:6px;transition:border-color 0.15s;user-select:none;}}
.chart-filter-trigger:hover{{border-color:var(--accent);}}
.chart-filter-trigger.active{{border-color:var(--accent);color:var(--accent);}}
.chart-filter-trigger-text{{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}}
.chart-filter-dd{{position:fixed;z-index:9999;background:var(--surface);border:1px solid var(--nav-border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,0.35);width:220px;overflow:hidden;}}
.chart-filter-dd-search{{margin:6px 8px;padding:4px 8px;border-radius:4px;border:0.5px solid var(--nav-border);background:var(--input-bg);color:var(--text1);font-size:11px;font-family:'DM Sans',sans-serif;outline:none;width:calc(100% - 16px);box-sizing:border-box;}}
.chart-filter-dd-search:focus{{border-color:var(--accent);}}
.chart-filter-dd-body{{max-height:180px;overflow-y:auto;padding:4px 0;}}
.chart-filter-dd-item{{display:flex;align-items:center;gap:7px;padding:4px 10px;cursor:pointer;font-size:11px;color:var(--text2);transition:background 0.1s;}}
.chart-filter-dd-item:hover{{background:var(--toggle-bg);}}
.chart-filter-dd-item input{{cursor:pointer;accent-color:var(--accent);flex-shrink:0;}}
.chart-filter-dd-footer{{padding:6px 8px;border-top:0.5px solid var(--nav-border);display:flex;gap:6px;justify-content:space-between;align-items:center;}}
.chart-filter-count{{font-size:10px;color:var(--text3);}}
.chart-filter-apply{{padding:3px 10px;border-radius:4px;font-size:10px;font-weight:600;cursor:pointer;background:var(--accent);color:var(--accent-text);border:none;font-family:'DM Sans',sans-serif;}}
.chart-filter-clear{{padding:3px 10px;border-radius:4px;font-size:10px;font-weight:600;cursor:pointer;background:var(--toggle-bg);color:var(--text2);border:0.5px solid var(--nav-border);font-family:'DM Sans',sans-serif;}}
.cft-multi-row{{display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap;}}
.cft-multi-row-num{{font-size:10px;color:var(--text3);font-weight:600;min-width:14px;}}
.cft-remove-btn{{padding:2px 7px;border-radius:4px;font-size:11px;cursor:pointer;background:transparent;border:0.5px solid var(--nav-border);color:var(--text3);font-family:'DM Sans',sans-serif;flex-shrink:0;transition:all 0.15s;}}
.cft-remove-btn:hover{{color:#f75a7a;border-color:#f75a7a;}}
.cft-multi-group{{display:flex;flex-direction:column;flex:1;min-width:120px;}}

/* ── CHART THEME ROW ── */
.chart-theme-row{{display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap;padding:8px 10px;background:var(--toggle-bg);border-radius:6px;border:0.5px solid var(--nav-border);}}
.chart-theme-label{{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-right:2px;}}
.palette-btn{{width:18px;height:18px;border-radius:4px;cursor:pointer;border:2px solid transparent;transition:all 0.15s;position:relative;overflow:hidden;flex-shrink:0;}}
.palette-btn.active{{border-color:var(--text1);transform:scale(1.15);}}
.palette-btn:hover{{transform:scale(1.1);}}

/* ── SIDEBAR CHART LINKS ── */
.sidebar-chart-link{{display:flex;align-items:center;gap:6px;padding:3px 12px 3px 28px;font-size:11px;color:var(--text3);cursor:pointer;transition:all 0.15s;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.sidebar-chart-link:hover{{color:var(--accent);}}
.sidebar-chart-link::before{{content:'📊';font-size:10px;flex-shrink:0;}}

/* ── FLEXIBLE COMBO PANEL ── */
.combo-axes-panel{{display:none;gap:8px;margin-bottom:10px;}}
.combo-axes-panel.visible{{display:flex;}}
.combo-axis-half{{flex:1;background:var(--toggle-bg);border:0.5px solid var(--nav-border);border-radius:6px;padding:8px 10px;min-width:0;}}
.combo-axis-half-label{{font-size:10px;font-weight:700;color:var(--text2);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;display:flex;align-items:center;gap:6px;}}
.combo-axis-badge{{font-size:9px;padding:2px 6px;border-radius:3px;font-weight:600;}}
.combo-axis-badge.left{{background:#4f8ef720;color:#4f8ef7;}}
.combo-axis-badge.right{{background:#f7b24f20;color:#f7b24f;}}
.combo-type-group{{display:flex;gap:3px;margin-bottom:6px;}}
.combo-type-btn{{padding:3px 8px;border-radius:4px;font-size:10px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--toggle-bg);color:var(--text3);font-family:'DM Sans',sans-serif;transition:all 0.15s;}}
.combo-type-btn:hover{{color:var(--text2);}}
.combo-type-btn.active{{background:var(--tab-bg);color:var(--accent-text);border-color:var(--accent);}}
.combo-y-checks{{display:flex;flex-direction:column;gap:3px;max-height:120px;overflow-y:auto;}}


/* ── MOBILE RESPONSIVE ── */
@media (max-width:768px) {{
  .app-layout{{flex-direction:column;height:auto;}}
  .sidebar{{width:100%!important;height:auto;max-height:200px;overflow-y:auto;border-right:none;border-bottom:0.5px solid var(--nav-border);}}
  .main-area{{flex-direction:column;}}
  .topnav{{flex-wrap:wrap;gap:4px;padding:8px 12px;}}
  .toolbar{{flex-wrap:wrap;gap:4px;}}
  .nav-pills-bar{{overflow-x:auto;}}
  .scroll-area{{padding:12px;}}
  .tbl-controls{{flex-wrap:wrap;}}
  .tbl-right-controls{{flex-wrap:wrap;gap:4px;}}
  .chart-builder{{padding:10px;}}
  .chart-axes-row{{flex-direction:column;}}
  .combo-axes-panel{{flex-direction:column;}}
}}

/* ── COLUMN RESIZE ── */
.tbl-wrap table {{table-layout:auto;}}
thead th{{position:relative;}}
.col-resizer{{position:absolute;right:0;top:0;width:5px;height:100%;cursor:col-resize;z-index:10;background:transparent;}}
.col-resizer:hover{{background:var(--accent);opacity:0.5;}}

/* ── STAT PANEL ── */
.stat-panel{{background:var(--surface);border:0.5px solid var(--nav-border);border-radius:8px;padding:12px 16px;margin-bottom:10px;display:none;}}
.stat-panel.visible{{display:block;}}
.stat-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:8px;}}
.stat-item{{background:var(--toggle-bg);border-radius:6px;padding:8px 10px;}}
.stat-label{{font-size:9px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:2px;}}
.stat-value{{font-size:13px;font-weight:700;color:var(--text1);}}
.stat-col-name{{font-size:11px;font-weight:600;color:var(--accent);margin-bottom:6px;}}

/* ── FILTER PRESETS ── */
.preset-wrap{{display:flex;align-items:center;gap:6px;flex-wrap:wrap;}}
.preset-tag{{padding:3px 8px;border-radius:12px;font-size:10px;font-weight:600;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--toggle-bg);color:var(--text2);font-family:'DM Sans',sans-serif;white-space:nowrap;transition:all 0.15s;}}
.preset-tag:hover{{border-color:var(--accent);color:var(--accent);}}
.preset-tag.active{{background:var(--accent);color:#fff;border-color:var(--accent);}}

/* ── COMPUTED COLS ── */
.computed-col-hdr{{background:rgba(167,139,250,0.12)!important;color:#a78bfa;font-size:10px;border-left:0.5px dashed var(--nav-border)!important;cursor:default;}}
.computed-col-cell{{background:rgba(167,139,250,0.06)!important;color:var(--text2);font-size:11px;border-left:0.5px dashed var(--nav-border)!important;}}

/* ── SHORTCUTS MODAL ── */
.shortcuts-modal{{position:fixed;inset:0;background:rgba(0,0,0,0.6);z-index:99999;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(4px);}}
.shortcuts-box{{background:var(--surface);border:0.5px solid var(--nav-border);border-radius:12px;padding:24px 28px;max-width:480px;width:90%;max-height:80vh;overflow-y:auto;}}
.shortcuts-title{{font-size:16px;font-weight:700;color:var(--text1);margin-bottom:16px;}}
.shortcut-row{{display:flex;align-items:center;justify-content:space-between;padding:5px 0;border-bottom:0.5px solid var(--nav-border);}}
.shortcut-key{{font-family:monospace;background:var(--toggle-bg);border:0.5px solid var(--nav-border);border-radius:4px;padding:2px 7px;font-size:12px;color:var(--accent);}}
.shortcut-desc{{font-size:12px;color:var(--text2);}}

/* ── COMPACT MODE ── */
body.compact-mode td, body.compact-mode th {{padding:3px 8px!important;font-size:11px!important;}}
body.compact-mode .tbl-controls {{margin-bottom:4px;}}
body.compact-mode .section {{margin-bottom:16px;}}

/* ── COND FMT cells ── */
.cond-green{{background:rgba(34,211,165,0.18)!important;color:#22d3a5;font-weight:600;}}
.cond-red{{background:rgba(247,90,122,0.18)!important;color:#f75a7a;font-weight:600;}}
.cond-yellow{{background:rgba(247,178,79,0.18)!important;color:#f7b24f;font-weight:600;}}

/* ── % OF TOTAL / RUNNING TOTAL extra cols ── */
.extra-col{{background:var(--toggle-bg)!important;color:var(--text3);font-size:10px;border-left:0.5px dashed var(--nav-border)!important;}}
.extra-col-hdr{{background:var(--toggle-bg)!important;color:var(--text3);font-size:10px;font-weight:600;border-left:0.5px dashed var(--nav-border)!important;}}

/* ── TABLE DESCRIPTION ── */
.tbl-description{{font-size:11px;color:var(--text3);padding:4px 0 8px 0;line-height:1.5;}}

/* ── KPI SCORECARDS ── */
.kpi-row{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px;}}
.kpi-card{{flex:1;min-width:130px;background:var(--surface);border:0.5px solid var(--tbl-border);border-left:3px solid var(--accent);border-radius:8px;padding:10px 14px;}}
.kpi-card-label{{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.kpi-card-value{{font-size:18px;font-weight:700;color:var(--text1);font-family:'DM Mono',monospace;}}
.kpi-card-sub{{font-size:10px;color:var(--text3);margin-top:2px;}}

/* ── ROW SPARKLINE ── */
.row-spark{{display:inline-block;vertical-align:middle;}}
.row-spark rect{{fill:var(--accent);}}

/* ── CAVEMAN HELP PANEL ── */
.caveman-fab{{position:fixed;bottom:20px;left:20px;width:42px;height:42px;border-radius:50%;background:var(--accent);color:#fff;border:none;cursor:pointer;font-size:20px;z-index:1500;box-shadow:0 4px 12px rgba(0,0,0,0.3);}}
.caveman-fab:hover{{opacity:0.9;}}

/* ── PIVOT BUILDER ── */
.pivot-panel{{margin-top:10px;}}
.pivot-builder{{background:var(--surface);border:0.5px solid var(--tbl-border);border-radius:10px;padding:14px 16px;}}
.pivot-builder-title{{font-size:13px;font-weight:700;color:var(--text1);margin-bottom:10px;}}
.pivot-controls{{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end;margin-bottom:6px;}}
.pivot-field{{display:flex;flex-direction:column;gap:4px;}}
.pivot-field label{{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:0.5px;}}
.pivot-result-label{{font-size:12px;font-weight:600;color:var(--accent);margin:10px 0 6px;}}
.pivot-multi{{min-width:150px;min-height:90px;height:auto;padding:4px;}}
.pivot-multi option{{padding:2px 4px;}}

/* ── COMPARE / DUAL-CHART OVERLAY ── */
.cmp-overlay{{position:fixed;inset:0;background:rgba(0,0,0,0.55);z-index:9000;display:none;overflow:auto;padding:32px 16px;}}
.cmp-overlay.open{{display:block;}}
.cmp-modal{{max-width:1100px;margin:0 auto;background:var(--surface);border:0.5px solid var(--tbl-border);border-radius:12px;padding:18px 20px;box-shadow:0 12px 40px rgba(0,0,0,0.4);}}
.cmp-modal-head{{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;}}
.cmp-modal-title{{font-size:15px;font-weight:700;color:var(--text1);}}
.cmp-slot{{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;padding:10px;border:0.5px dashed var(--tbl-border);border-radius:8px;margin-bottom:8px;}}
.cmp-slot .pivot-field select{{min-width:150px;}}
.cmp-actions{{display:flex;gap:10px;align-items:center;margin-top:10px;flex-wrap:wrap;}}
.cmp-out{{margin-top:14px;}}

/* ── PRINT CSS (basic) ── */
@media print {{
  .sidebar,.topnav,.toolbar,.nav-pills-bar,.tbl-controls,.pagination,.action-btn,.section-actions,.fullscreen-overlay,.toast{{display:none!important;}}
  .app-layout{{display:block!important;height:auto!important;}}
  .main-area{{overflow:visible!important;}}
  .scroll-area{{overflow:visible!important;height:auto!important;padding:0!important;}}
  .tbl-wrap{{overflow:visible!important;max-height:none!important;}}
}}
</style>
<!-- Plotly JS — CDN by default, or inlined when embed_plotly=True (offline) -->
{plotly_tag}
</head>
<body>

<div class="app-layout">

  <!-- ── SIDEBAR ── -->
  <div class="sidebar" id="sidebar">
    <div class="sidebar-header">
      <span class="sidebar-title">Tables</span>
    </div>
    <div class="sidebar-body" id="sidebar-body"></div>
  </div>

  <!-- ── MAIN ── -->
  <div class="main-area">

    <!-- TOPNAV -->
    <div class="topnav">
      <div class="topnav-left">
        <button class="sidebar-toggle" onclick="toggleSidebar()" title="Toggle Sidebar">☰</button>
        <div class="brand"><div class="logo">{_title[0]}</div>{_title}</div>
      </div>
      <div class="topnav-right">
        <div class="global-search-wrap">
          <span class="global-search-icon">🔍</span>
          <input class="global-search" id="global-search" placeholder="Global search…" oninput="globalSearch(this.value)" onfocus="showGlobalResults()" autocomplete="off"/>
          <div class="global-results" id="global-results"></div>
        </div>
        <span class="timestamp">Updated: {timestamp}</span>
      </div>
    </div>

    <!-- NAV PILLS -->
    <div class="nav-pills-bar" id="nav-pills"></div>

    <!-- TOOLBAR -->
    <div class="toolbar">
      <div class="toolbar-left">
        <span class="toolbar-label">View</span>
        <div class="toggle-wrap">
          <button class="toggle-btn active" id="btn-expand"   onclick="setView('expand')">⊞ Expanded</button>
          <button class="toggle-btn"        id="btn-collapse" onclick="setView('collapse')">☰ Collapsed</button>
        </div>
        <div class="divider"></div>
        <span class="toolbar-label">Numbers</span>
        <div class="toggle-wrap">
          <button class="toggle-btn active" id="fmt-actual" onclick="setNumFmt('actual')">Actual</button>
          <button class="toggle-btn"        id="fmt-k"      onclick="setNumFmt('k')">K</button>
          <button class="toggle-btn"        id="fmt-m"      onclick="setNumFmt('m')">M</button>
          <button class="toggle-btn"        id="fmt-b"      onclick="setNumFmt('b')">B</button>
        </div>
        <div class="divider"></div>
        <span class="toolbar-label">Format</span>
        <div class="toggle-wrap">
          <button class="toggle-btn active" id="colfmt-excel" onclick="setColFmt('excel')" title="Excel format: rounded integers + PCT% (shortcut: E)">Excel</button>
          <button class="toggle-btn"        id="colfmt-raw"   onclick="setColFmt('raw')"   title="Raw values from dataframe (shortcut: R)">Raw</button>
        </div>
        <div class="divider"></div>
        <button class="action-btn" onclick="openCompareBuilder()" title="Compare a metric across tables / periods">⇄ Compare</button>
        <button class="action-btn" onclick="exportCharts()" style="color:#22d3a5;border-color:#22d3a5;">📊 Export Charts</button>
        <button class="action-btn" id="darklight-btn" onclick="toggleDarkLight()" title="Toggle Dark/Light mode">☀</button>
        <button class="action-btn" id="compact-btn"   onclick="toggleCompact()"   title="Compact row height">⊟</button>
        <button class="action-btn" onclick="toggleShortcuts()" title="Keyboard shortcuts (?)">⌨</button>
        <div class="divider"></div>
        <span class="toolbar-label">Theme</span>
        <div class="theme-switcher">
          <div class="theme-dot" style="background:#4f6ef7;"                      data-theme="Dark Blue"      title="Dark Blue"      onclick="setTheme('Dark Blue')"></div>
          <div class="theme-dot" style="background:#aaaaaa;"                      data-theme="Carbon"         title="Carbon"         onclick="setTheme('Carbon')"></div>
          <div class="theme-dot" style="background:#22d3a5;"                      data-theme="Midnight Green" title="Midnight Green"  onclick="setTheme('Midnight Green')"></div>
          <div class="theme-dot" style="background:#3b5bdb;"                      data-theme="Slate Light"    title="Slate Light"     onclick="setTheme('Slate Light')"></div>
          <div class="theme-dot" style="background:#f75a7a;"                      data-theme="Crimson"        title="Crimson"         onclick="setTheme('Crimson')"></div>
          <div class="theme-dot" style="background:#000;border:1.5px solid #444;" data-theme="Night"          title="Night"           onclick="setTheme('Night')"></div>
        </div>
      </div>
    </div>

    <!-- SCROLL AREA -->
    <div class="scroll-area" id="scroll-area"></div>

  </div><!-- end main-area -->
</div><!-- end app-layout -->

<!-- FULLSCREEN OVERLAY -->
<div class="fullscreen-overlay" id="fullscreen-overlay">
  <div class="fullscreen-header">
    <span class="fullscreen-title" id="fullscreen-title"></span>
    <button class="fullscreen-close" onclick="closeFullscreen()">✕ Close</button>
  </div>
  <div class="fullscreen-body" id="fullscreen-body"></div>
</div>

<!-- COMPARE / DUAL-CHART OVERLAY (filled dynamically) -->
<div class="cmp-overlay" id="cmp-overlay" onclick="if(event.target===this)closeCompareBuilder()"></div>

<!-- TOAST -->
<div class="toast" id="toast"></div>

<!-- CHART BUILDER PANEL (injected per table) -->
<script>
// ── GRAND TOTAL KEYWORDS ──
const TOTAL_KEYWORDS = ["grand total","total","subtotal","sub total","overall","overall total","net total","sum","gt"];

function isGrandTotal(val) {{
  return TOTAL_KEYWORDS.includes(String(val).toLowerCase().trim());
}}

// ── CHART BUILDER ──
let chartInstanceCount = {{}};

function buildChartBuilder(tableId, data, instanceId) {{
  const cols        = data.cols;
  const numericCols = data.numeric_cols || [];
  const iid         = instanceId;

  return `
  <div class="chart-builder chart-instance" id="chartbuilder-${{iid}}">
    <div class="chart-builder-header">
      <div class="chart-builder-title">📊 Chart ${{iid.split('-inst-')[1] || ''}}</div>
      <div style="display:flex;gap:6px;align-items:center;">
        <button class="chart-edit-btn" id="edit-btn-${{iid}}" onclick="toggleChartBuilder('${{iid}}')">⚙ Edit</button>
        <button class="chart-edit-btn" id="chart-toggle-btn-${{iid}}" onclick="toggleChartOutput('${{iid}}')">▼ Chart</button>
        <button class="chart-instance-remove" onclick="removeChartInstance('${{iid}}')">✕ Remove</button>
      </div>
    </div>
    <div class="chart-builder-body collapsed" id="builder-body-${{iid}}">
      <div class="chart-title-row">
        <div>
          <div class="chart-title-label">Chart Title</div>
          <input class="chart-title-input" id="chart-title-${{iid}}" placeholder="e.g. Bank Wise Success Rate" type="text"/>
        </div>
        <div>
          <div class="chart-title-label">Description</div>
          <input class="chart-title-input" id="chart-desc-${{iid}}" placeholder="e.g. Top banks by success %" type="text"/>
        </div>
      </div>
      <!-- AXIS LABEL INPUTS -->
      <div class="chart-axis-label-row" id="axis-label-row-${{iid}}">
        <div class="chart-axis-label-group">
          <div class="chart-axis-label-tag">X Axis Label</div>
          <input class="chart-axis-label-input" id="xlabel-custom-${{iid}}" placeholder="Auto (from column name)" type="text" oninput="_autoRegenerate('${{iid}}')"/>
        </div>
        <div class="chart-axis-label-group" id="ylabel-custom-group-${{iid}}">
          <div class="chart-axis-label-tag">Y Axis Label</div>
          <input class="chart-axis-label-input" id="ylabel-custom-${{iid}}" placeholder="Auto (from column name)" type="text" oninput="_autoRegenerate('${{iid}}')"/>
        </div>
        <div class="chart-axis-label-group" id="ylabel2-custom-group-${{iid}}" style="display:none;">
          <div class="chart-axis-label-tag">Right Y Label</div>
          <input class="chart-axis-label-input" id="ylabel2-custom-${{iid}}" placeholder="Auto (from column name)" type="text" oninput="_autoRegenerate('${{iid}}')"/>
        </div>
      </div>

      <div class="chart-type-row">
        <span class="chart-type-label">Chart Type</span>
        <button class="chart-type-btn active" data-ctype="bar"     onclick="selectChartType('${{iid}}','bar',this)">▊ Bar</button>
        <button class="chart-type-btn"        data-ctype="line"    onclick="selectChartType('${{iid}}','line',this)">📈 Line</button>
        <button class="chart-type-btn"        data-ctype="area"    onclick="selectChartType('${{iid}}','area',this)">🏔 Area</button>
        <button class="chart-type-btn"        data-ctype="scatter" onclick="selectChartType('${{iid}}','scatter',this)">⬤ Scatter</button>
        <button class="chart-type-btn"        data-ctype="pie"     onclick="selectChartType('${{iid}}','pie',this)">🥧 Pie</button>
        <button class="chart-type-btn"        data-ctype="donut"   onclick="selectChartType('${{iid}}','donut',this)">🍩 Donut</button>
        <button class="chart-type-btn"        data-ctype="combo"      onclick="selectChartType('${{iid}}','combo',this)">📊 Combo</button>
        <button class="chart-type-btn"        data-ctype="waterfall"  onclick="selectChartType('${{iid}}','waterfall',this)">📉 Waterfall</button>
        <button class="chart-type-btn"        data-ctype="funnel"     onclick="selectChartType('${{iid}}','funnel',this)">🔽 Funnel</button>
        <button class="chart-type-btn"        data-ctype="heatmap"    onclick="selectChartType('${{iid}}','heatmap',this)">🟩 Heatmap</button>
        <button class="chart-type-btn"        data-ctype="bubble"     onclick="selectChartType('${{iid}}','bubble',this)">🫧 Bubble</button>
        <button class="chart-type-btn"        data-ctype="box"        onclick="selectChartType('${{iid}}','box',this)">📦 Box</button>
        <button class="chart-type-btn"        data-ctype="treemap"    onclick="selectChartType('${{iid}}','treemap',this)">🌳 Treemap</button>
        <button class="chart-type-btn"        data-ctype="sunburst"   onclick="selectChartType('${{iid}}','sunburst',this)">☀ Sunburst</button>
        <button class="chart-type-btn"        data-ctype="gauge"      onclick="selectChartType('${{iid}}','gauge',this)">🎯 Gauge</button>
      </div>

      <div class="chart-options-row" id="bar-opts-${{iid}}">
        <span class="chart-type-label">Bar Mode</span>
        <div class="chart-opt-group">
          <button class="chart-opt-btn active" onclick="setChartOpt('${{iid}}','barmode','group',this)">Group</button>
          <button class="chart-opt-btn"        onclick="setChartOpt('${{iid}}','barmode','stack',this)">Stack</button>
          <button class="chart-opt-btn"        onclick="setChartOpt('${{iid}}','barmode','overlay',this)">Overlay</button>
        </div>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Orient</span>
        <div class="chart-opt-group">
          <button class="chart-opt-btn active" onclick="setChartOpt('${{iid}}','orient','v',this)">Vertical</button>
          <button class="chart-opt-btn"        onclick="setChartOpt('${{iid}}','orient','h',this)">Horizontal</button>
        </div>
        <div class="chart-divider"></div>
        <button class="chart-opt-toggle" id="corners-${{iid}}" onclick="toggleChartOpt('${{iid}}','corners')">⬜ Corners: OFF</button>
      </div>

      <div class="chart-options-row">
        <span class="chart-type-label">Sort</span>
        <select class="chart-opt-select" id="sort-${{iid}}">
          <option value="original">Original</option>
          <option value="asc">Ascending</option>
          <option value="desc">Descending</option>
          <option value="az">A → Z</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Show</span>
        <select class="chart-opt-select" id="topn-${{iid}}">
          <option value="all">All</option>
          <option value="5">Top 5</option>
          <option value="10">Top 10</option>
          <option value="15">Top 15</option>
          <option value="b5">Bottom 5</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-type-label">X Axis</span>
        <select class="chart-opt-select" id="xaxis-type-${{iid}}" onchange="_autoRegenerate('${{iid}}')" title="X axis scale — Auto detects category/numeric, override if needed">
          <option value="auto">Auto</option>
          <option value="category">Category</option>
          <option value="numeric">Numeric</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Y Scale</span>
        <select class="chart-opt-select" id="yscale-${{iid}}">
          <option value="linear">Linear</option>
          <option value="log">Log</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Hover</span>
        <select class="chart-opt-select" id="hovermode-${{iid}}">
          <option value="closest">Closest</option>
          <option value="x unified">Unified</option>
        </select>
      </div>

      <div class="chart-options-row">
        <button class="chart-opt-toggle" id="pattern-${{iid}}" onclick="toggleChartOpt('${{iid}}','pattern')">▤ Pattern: OFF</button>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Color</span>
        <select class="chart-opt-select" id="colormode-${{iid}}">
          <option value="flat">Flat</option>
          <option value="byvalue">By Value</option>
        </select>
        <div class="chart-divider"></div>
        <button class="chart-opt-toggle" id="opacity-${{iid}}" onclick="toggleChartOpt('${{iid}}','opacity')">◐ Opacity: OFF</button>
        <div class="chart-divider"></div>
        <button class="chart-opt-toggle" id="rangeslider-${{iid}}" onclick="toggleChartOpt('${{iid}}','rangeslider')">↔ Range Slider: OFF</button>
        <div class="chart-divider"></div>
        <button class="chart-opt-toggle" id="annotate-${{iid}}" onclick="toggleChartOpt('${{iid}}','annotate')">📌 Annotate: OFF</button>
        <div class="chart-divider"></div>
        <button class="chart-opt-toggle" id="condcolor-${{iid}}" onclick="toggleChartOpt('${{iid}}','condcolor')">🔴 Cond. Color: OFF</button>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Ref Line</span>
        <input class="chart-axis-label-input" id="refline-${{iid}}" placeholder="e.g. 0.90" type="number" step="any" style="width:80px;" oninput="_autoRegenerate('${{iid}}')"/>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Trend</span>
        <select class="chart-opt-select" id="trend-${{iid}}" onchange="_autoRegenerate('${{iid}}')" title="Overlay a trendline / moving average on line, bar, scatter or area charts">
          <option value="none">None</option>
          <option value="mean">Mean</option>
          <option value="linear">Linear fit</option>
          <option value="ma3">Moving avg (3)</option>
          <option value="ma5">Moving avg (5)</option>
          <option value="ma7">Moving avg (7)</option>
        </select>
      </div>
      <div class="chart-axes-row" id="axes-${{iid}}">
        <div class="chart-axis-group">
          <span class="chart-axis-label" id="xlabel-${{iid}}">X Axis (Dimension)</span>
          <select class="chart-axis-select" id="xcol-${{iid}}">
            ${{cols.map((c,i) => '<option value="'+i+'">'+String(c)+'</option>').join('')}}
          </select>
        </div>
        <div class="chart-axis-group" id="ygroup-${{iid}}">
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;">
            <span class="chart-axis-label" id="ylabel-${{iid}}">Y Axis (Metrics)</span>
            <div style="display:flex;gap:4px;" id="yaxis-btns-${{iid}}">
              <button style="padding:2px 8px;border-radius:4px;font-size:10px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--tab-bg);color:var(--text2);font-family:'DM Sans',sans-serif;" onclick="selectAllY('${{iid}}',true)">All</button>
              <button style="padding:2px 8px;border-radius:4px;font-size:10px;font-weight:500;cursor:pointer;border:0.5px solid var(--nav-border);background:var(--tab-bg);color:var(--text2);font-family:'DM Sans',sans-serif;" onclick="selectAllY('${{iid}}',false)">None</button>
            </div>
          </div>
          <div class="chart-y-checks" id="ycols-${{iid}}">
            ${{numericCols.map(i => '<label class="chart-y-check"><input type="checkbox" value="'+i+'" checked> '+String(cols[i])+'</label>').join('')}}
          </div>
        </div>
        <div class="chart-axis-group" id="colorby-group-${{iid}}">
          <span class="chart-axis-label">Color By <span style="font-size:9px;color:var(--text3);">(optional — groups bars by category)</span></span>
          <select class="chart-axis-select" id="colorbycol-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
            <option value="-1">— None —</option>
            ${{cols.map((c,i) => '<option value="'+i+'">'+String(c)+'</option>').join('')}}
          </select>
        </div>
      </div>

      <!-- FLEXIBLE COMBO PANEL — shown only when Combo selected -->
      <div class="combo-axes-panel" id="combo-panel-${{iid}}">
        <div class="combo-axis-half">
          <div class="combo-axis-half-label">
            Left Axis
            <span class="combo-axis-badge left">Y1</span>
          </div>
          <div class="combo-type-group" id="combo-left-type-${{iid}}">
            <button class="combo-type-btn active" onclick="_setComboType('${{iid}}','left','bar',this)">▊ Bar</button>
            <button class="combo-type-btn"        onclick="_setComboType('${{iid}}','left','line',this)">📈 Line</button>
            <button class="combo-type-btn"        onclick="_setComboType('${{iid}}','left','area',this)">🏔 Area</button>
          </div>
          <div class="combo-y-checks" id="combo-left-cols-${{iid}}" data-iid="${{iid}}">
            ${{numericCols.map(i => '<label class="chart-y-check"><input type="checkbox" value="'+i+'" '+(i===numericCols[0]?'checked':'')+' data-combo-side="left" onchange="_comboColChange(this)"> '+String(cols[i])+'</label>').join('')}}
          </div>
        </div>
        <div class="combo-axis-half">
          <div class="combo-axis-half-label">
            Right Axis
            <span class="combo-axis-badge right">Y2</span>
          </div>
          <div class="combo-type-group" id="combo-right-type-${{iid}}">
            <button class="combo-type-btn"        onclick="_setComboType('${{iid}}','right','bar',this)">▊ Bar</button>
            <button class="combo-type-btn active" onclick="_setComboType('${{iid}}','right','line',this)">📈 Line</button>
            <button class="combo-type-btn"        onclick="_setComboType('${{iid}}','right','area',this)">🏔 Area</button>
          </div>
          <div class="combo-y-checks" id="combo-right-cols-${{iid}}" data-iid="${{iid}}">
            ${{numericCols.map(i => '<label class="chart-y-check"><input type="checkbox" value="'+i+'" '+(i===numericCols[1]||numericCols.length===1?'checked':'')+' data-combo-side="right" onchange="_comboColChange(this)"> '+String(cols[i])+'</label>').join('')}}
          </div>
        </div>
      </div>
      <!-- CHART FILTER — MULTI ROW -->
      <div id="cft-container-${{iid}}">
        <div id="cft-rows-${{iid}}"></div>
        <div style="margin-bottom:10px;">
          <button class="chart-opt-toggle" onclick="_cftAddRow('${{iid}}')">+ Add Filter</button>
        </div>
      </div>

      <!-- CHART THEME ROW -->
      <div class="chart-theme-row">
        <span class="chart-theme-label">Palette</span>
        <button class=\"palette-btn active\" id=\"palette-PayU-${{iid}}\" title=\"PayU\" style=\"background:linear-gradient(135deg,#4f8ef7 50%,#22d3a5 50%);\" onclick=\"_setPalette('${{iid}}','PayU',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Vibrant-${{iid}}\" title=\"Vibrant\" style=\"background:linear-gradient(135deg,#ff3366 50%,#ff9500 50%);\" onclick=\"_setPalette('${{iid}}','Vibrant',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Pastel-${{iid}}\" title=\"Pastel\" style=\"background:linear-gradient(135deg,#a8d8ea 50%,#aa96da 50%);\" onclick=\"_setPalette('${{iid}}','Pastel',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Mono-${{iid}}\" title=\"Mono\" style=\"background:linear-gradient(135deg,#e2e8f0 50%,#94a3b8 50%);\" onclick=\"_setPalette('${{iid}}','Mono',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Sunset-${{iid}}\" title=\"Sunset\" style=\"background:linear-gradient(135deg,#ff6b6b 50%,#ffa500 50%);\" onclick=\"_setPalette('${{iid}}','Sunset',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Ocean-${{iid}}\" title=\"Ocean\" style=\"background:linear-gradient(135deg,#0077b6 50%,#00b4d8 50%);\" onclick=\"_setPalette('${{iid}}','Ocean',this)\"></button>
        <button class=\"palette-btn\" id=\"palette-Forest-${{iid}}\" title=\"Forest\" style=\"background:linear-gradient(135deg,#2d6a4f 50%,#52b788 50%);\" onclick=\"_setPalette('${{iid}}','Forest',this)\"></button>
        <div class="chart-divider"></div>
        <span class="chart-theme-label">BG</span>
        <select class="chart-opt-select" id="chartbg-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
          <option value="theme">Theme</option>
          <option value="white">White</option>
          <option value="dark">Dark</option>
          <option value="transparent">Transparent</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-theme-label">Grid</span>
        <select class="chart-opt-select" id="chartgrid-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
          <option value="on">On</option>
          <option value="off">Off</option>
          <option value="h">H Only</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-theme-label">Font</span>
        <select class="chart-opt-select" id="chartfont-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
          <option value="medium">Medium</option>
          <option value="small">Small</option>
          <option value="large">Large</option>
        </select>
        <div class="chart-divider"></div>
        <span class="chart-theme-label">Border</span>
        <select class="chart-opt-select" id="chartborder-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
          <option value="none">None</option>
          <option value="thin">Thin</option>
          <option value="thick">Thick</option>
        </select>
      </div>
      <div class="chart-generate-row">
        <button class="chart-generate-btn" onclick="generateChart('${{iid}}','${{tableId}}')">▶ Generate</button>
        <button class="chart-label-toggle" id="label-toggle-${{iid}}" onclick="toggleLabels('${{iid}}')">🏷 Labels: OFF</button>
        <button class="chart-opt-toggle" id="pin-${{iid}}" onclick="_togglePin('${{iid}}')">📌 Pin: OFF</button>
        <button class="chart-opt-toggle" onclick="_copyChartImage('${{iid}}')" title="Copy chart as PNG image">📋 Copy Image</button>
        <div class="chart-divider"></div>
        <span class="chart-type-label">Agg</span>
        <select class="chart-opt-select" id="aggmode-${{iid}}" onchange="_autoRegenerate('${{iid}}')">
          <option value="none">None</option>
          <option value="sum">Sum</option>
          <option value="avg">Avg</option>
          <option value="count">Count</option>
          <option value="max">Max</option>
          <option value="min">Min</option>
        </select>
        <span style="font-size:11px;color:var(--text3);">Grand Total auto-excluded</span>
      </div>
    </div>
    <div id="chart-output-wrap-${{iid}}">
      <div id="chart-title-display-${{iid}}" style="padding:0 16px;"></div>
      <div class="chart-output" id="chartout-${{iid}}">
        <div class="chart-placeholder">Click ⚙ Edit → configure → Generate</div>
      </div>
    </div>
  </div>`;
}}

// ── PALETTE STATE ──
// ── TIME-LIKE COLUMN DETECTION ──
const TIME_KEYWORDS = ['year','month','date','day','week','quarter',
                       'period','fy','fiscal','hour','time','dt','yr',
                       'year_month','yearmonth','yyyymm','mmyyyy'];

function _isTimeCol(colName) {{
  const lower = String(colName).toLowerCase();
  return TIME_KEYWORDS.some(kw => lower === kw || lower.includes(kw));
}}

// ID-like column names that should be treated as categories even if numeric
const ID_KEYWORDS = ['id','code','mapping','plant','store','sku','no','num','number',
                     'key','ref','pin','zip','postal','branch','region','zone','area',
                     'flag','type','class','cat','seg','grp','group'];

function _isIdCol(colName) {{
  const lower = String(colName).toLowerCase().replace(/[_\\s-]/g,' ');
  return ID_KEYWORDS.some(kw => lower === kw || lower.split(' ').includes(kw) || lower.endsWith(' ' + kw) || lower.startsWith(kw + ' '));
}}

// ── DATE / CHRONOLOGICAL HELPERS (shared: table sort + chart axis) ──
function _isDateStr(v) {{
  return /^\\d{{4}}-\\d{{2}}(-\\d{{2}})?$/.test(String(v));
}}
function _valsLookDate(vals) {{
  const s = (vals||[]).slice(0, 50).map(v => String(v)).filter(v => v && v !== 'null' && v !== 'None');
  if (!s.length) return false;
  return s.filter(_isDateStr).length / s.length > 0.8;
}}
function _datesHaveDay(vals) {{
  return (vals||[]).slice(0, 50).map(v => String(v)).some(v => /^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(v));
}}
function _chronoKey(v) {{
  const s = String(v);
  if (_isDateStr(s)) return s;
  const M = {{jan:1,feb:2,mar:3,apr:4,may:5,jun:6,jul:7,aug:8,sep:9,oct:10,nov:11,dec:12}};
  const mi = M[s.toLowerCase().slice(0,3)];
  if (mi) return 'M' + String(mi).padStart(2,'0');
  const n = parseFloat(s);
  return (isNaN(n) || /^\\d{{4}}-/.test(s)) ? s : n;
}}
function _chronoCmp(a, b) {{
  const ka = _chronoKey(a), kb = _chronoKey(b);
  return ka < kb ? -1 : ka > kb ? 1 : 0;
}}

// ── SMART ROUND FOR AGGREGATED VALUES ──
function _smartRound(v, aggMode) {{
  if (v === null || v === undefined || isNaN(v)) return v;
  if (aggMode === 'count') return Math.round(v);
  const abs = Math.abs(v);
  if (abs >= 10000) return Math.round(v);
  if (abs >= 100)   return Math.round(v * 10) / 10;
  if (abs >= 1)     return Math.round(v * 100) / 100;
  return Math.round(v * 10000) / 10000;  // PCT range 0-1
}}

const PALETTES = {{
  'PayU':    ['#4f8ef7','#22d3a5','#f7b24f','#f75a7a','#a78bfa','#34d399','#60a5fa','#fb923c','#e879f9','#4ade80'],
  'Vibrant': ['#ff3366','#ff9500','#30d5c8','#7b2fff','#00cc44','#ff6600','#e91e63','#2196f3','#4caf50','#ff5722'],
  'Pastel':  ['#a8d8ea','#aa96da','#fcbad3','#ffffd2','#b5ead7','#ffdac1','#c7ceea','#e2f0cb','#ffd3b6','#ffaaa5'],
  'Mono':    ['#e2e8f0','#cbd5e1','#94a3b8','#64748b','#475569','#334155','#1e293b','#0f172a','#f8fafc','#f1f5f9'],
  'Sunset':  ['#ff6b6b','#ffa500','#ffd700','#ff69b4','#ff4500','#dc143c','#ff8c00','#ff1493','#fa8072','#e9967a'],
  'Ocean':   ['#0077b6','#00b4d8','#90e0ef','#48cae4','#023e8a','#0096c7','#caf0f8','#ade8f4','#00b4d8','#0077b6'],
  'Forest':  ['#2d6a4f','#40916c','#52b788','#74c69d','#95d5b2','#b7e4c7','#d8f3dc','#1b4332','#081c15','#40916c'],
}};
const paletteState = {{}};

function _setPalette(iid, name, btn) {{
  paletteState[iid] = name;
  const row = btn.closest('.chart-theme-row');
  if (row) row.querySelectorAll('.palette-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  _autoRegenerate(iid);
}}

function _getColors(iid) {{
  const name = paletteState[iid] || 'PayU';
  return PALETTES[name] || PALETTES['PayU'];
}}

// ── CHART FILTER STATE (multi-row) ──
// chartFilterState[iid] = [ {{colIdx, allowed:Set|null}}, ... ]
const chartFilterState = {{}};
let   _cftRowCounter   = {{}};   // per-iid row ID counter

function _getCftRows(iid) {{
  if (!chartFilterState[iid]) chartFilterState[iid] = [];
  return chartFilterState[iid];
}}

// ── OPEN/CLOSE DROPDOWN TRACKER ──
let _openCftDd = null;
function _closeCftDd() {{
  if (_openCftDd) {{ _openCftDd.remove(); _openCftDd = null; }}
}}
document.addEventListener('click', e => {{
  if (_openCftDd && !_openCftDd.contains(e.target) &&
      !e.target.closest('.chart-filter-trigger') &&
      !e.target.closest('.cft-remove-btn')) _closeCftDd();
}});

// ── ADD A FILTER ROW ──
function _cftAddRow(iid) {{
  const rows = _getCftRows(iid);
  if (rows.length >= 5) {{ showToast('Maximum 5 filters per chart'); return; }}
  if (!_cftRowCounter[iid]) _cftRowCounter[iid] = 0;
  const rid = iid + '-r' + (_cftRowCounter[iid]++);
  rows.push({{ rid, colIdx: -1, allowed: null }});
  _cftRender(iid);
}}

function _cftRemoveRow(iid, rid) {{
  chartFilterState[iid] = (chartFilterState[iid]||[]).filter(r => r.rid !== rid);
  _cftRender(iid);
  _autoRegenerate(iid);
}}

// ── RENDER ALL FILTER ROWS ──
function _cftRender(iid) {{
  const container = document.getElementById('cft-rows-' + iid);
  if (!container) return;
  const rows = _getCftRows(iid);
  container.innerHTML = rows.map((row, idx) => _cftRowHTML(iid, row, idx)).join('');
}}

function _cftRowHTML(iid, row, idx) {{
  const colLabel  = row.colIdx >= 0 && row._colName ? row._colName : 'Column...';
  const colActive = row.colIdx >= 0;
  const valLabel  = !colActive ? 'Select column first'
    : row.allowed === null ? 'All values'
    : [...row.allowed].slice(0,3).join(', ') + (row.allowed.size > 3 ? ' +' + (row.allowed.size-3) + ' more' : '');
  const valActive = colActive && row.allowed !== null;

  return '<div class="cft-multi-row" id="cft-row-' + row.rid + '">' +
    '<span class="cft-multi-row-num">' + (idx+1) + '</span>' +
    '<div class="cft-multi-group">' +
      '<div class="chart-filter-trigger' + (colActive?' active':'') + '" onclick="_cftOpenColDd(' + JSON.stringify(iid) + ',' + JSON.stringify(row.rid) + ',this)">' +
        '<span class="chart-filter-trigger-text">' + colLabel + '</span><span>▾</span>' +
      '</div>' +
    '</div>' +
    '<div class="cft-multi-group" style="' + (!colActive ? 'opacity:0.4;pointer-events:none;' : '') + '">' +
      '<div class="chart-filter-trigger' + (valActive?' active':'') + '" onclick="_cftOpenValDd(' + JSON.stringify(iid) + ',' + JSON.stringify(row.rid) + ',this)">' +
        '<span class="chart-filter-trigger-text">' + valLabel + '</span><span>▾</span>' +
      '</div>' +
    '</div>' +
    '<button class="cft-remove-btn" onclick="_cftRemoveRow(' + JSON.stringify(iid) + ',' + JSON.stringify(row.rid) + ')">✕</button>' +
  '</div>';
}}

// ── OPEN COLUMN DROPDOWN ──
function _cftOpenColDd(iid, rid, triggerEl) {{
  _closeCftDd();
  const tableId   = iid.split('-inst-')[0];
  const tableData = findTableData(tableId);
  if (!tableData) return;

  const xColIdx = parseInt((document.getElementById('xcol-' + iid)||{{}}).value ?? -1);
  const cols    = tableData.cols;
  const rows    = tableData.rows;

  // Get already-used colIdxs (exclude current row)
  const usedIdxs = new Set(
    (_getCftRows(iid)).filter(r => r.rid !== rid && r.colIdx >= 0).map(r => r.colIdx)
  );

  const catCols = cols.map((col, i) => ({{col: String(col), idx: i}}))
    .filter(o => o.idx !== xColIdx && !usedIdxs.has(o.idx) && _colType(o.idx, rows) === 'categorical');

  const dd    = document.createElement('div');
  dd.className = 'chart-filter-dd';
  const rect   = triggerEl.getBoundingClientRect();
  dd.style.top  = (rect.bottom + 4) + 'px';
  dd.style.left = rect.left + 'px';
  _openCftDd = dd;

  dd.innerHTML =
    '<input class="chart-filter-dd-search" placeholder="Search columns..." oninput="_cftColSearch(this)">' +
    '<div class="chart-filter-dd-body">' +
    catCols.map(o =>
      '<div class="chart-filter-dd-item" data-iid="' + iid + '" data-rid="' + rid + '" data-idx="' + o.idx + '" data-col="' + o.col.replace(/"/g,'&quot;') + '" onclick="_cftSelectCol(this)">' +
      '<span>' + o.col + '</span></div>'
    ).join('') +
    (catCols.length === 0 ? '<div style="padding:8px 10px;font-size:11px;color:var(--text3);">No more categorical columns</div>' : '') +
    '</div>';

  document.body.appendChild(dd);
  const ddRect = dd.getBoundingClientRect();
  if (ddRect.bottom > window.innerHeight) dd.style.top = (rect.top - ddRect.height - 4) + 'px';
  dd.querySelector('.chart-filter-dd-search').focus();
}}

function _cftColSearch(input) {{
  input.closest('.chart-filter-dd').querySelectorAll('.chart-filter-dd-item').forEach(el => {{
    el.style.display = el.textContent.toLowerCase().includes(input.value.toLowerCase()) ? '' : 'none';
  }});
}}

function _cftSelectCol(el) {{
  const iid    = el.dataset.iid;
  const rid    = el.dataset.rid;
  const colIdx = parseInt(el.dataset.idx);
  const colName = el.dataset.col;
  _closeCftDd();

  const row = (_getCftRows(iid)).find(r => r.rid === rid);
  if (!row) return;
  // Reset allowed if column changed
  if (row.colIdx !== colIdx) row.allowed = null;
  row.colIdx   = colIdx;
  row._colName = colName;
  _cftRender(iid);
  _autoRegenerate(iid);
}}

// ── OPEN VALUES DROPDOWN ──
function _cftOpenValDd(iid, rid, triggerEl) {{
  _closeCftDd();
  const row = (_getCftRows(iid)).find(r => r.rid === rid);
  if (!row || row.colIdx < 0) return;

  const tableId   = iid.split('-inst-')[0];
  const tableData = findTableData(tableId);
  if (!tableData) return;

  const allVals = [...new Set(tableData.rows.map(r => String(r[row.colIdx])).filter(v => v && v !== 'null' && v !== 'None'))].sort();
  const allowed = row.allowed || new Set(allVals);
  const allChecked = allowed.size === allVals.length;

  const dd    = document.createElement('div');
  dd.className = 'chart-filter-dd';
  dd.dataset.iid = iid;
  dd.dataset.rid = rid;
  const rect   = triggerEl.getBoundingClientRect();
  dd.style.top  = (rect.bottom + 4) + 'px';
  dd.style.left = rect.left + 'px';
  _openCftDd = dd;

  dd.innerHTML =
    '<input class="chart-filter-dd-search" placeholder="Search values..." oninput="_cftValSearch(this)">' +
    '<div class="chart-filter-dd-body" id="cft-vbody-' + rid + '">' +
    '<div class="chart-filter-dd-item">' +
      '<input type="checkbox" id="cft-all-' + rid + '" ' + (allChecked ? 'checked' : '') + ' onchange="_cftToggleAll(this)">' +
      '<label for="cft-all-' + rid + '" style="font-weight:600;cursor:pointer;">Select All</label>' +
    '</div>' +
    allVals.map(v =>
      '<div class="chart-filter-dd-item" data-val="' + v.replace(/"/g,'&quot;') + '">' +
        '<input type="checkbox" value="' + v.replace(/"/g,'&quot;') + '" ' + (allowed.has(v) ? 'checked' : '') + ' onchange="_cftValCheck(this)">' +
        '<label style="cursor:pointer;">' + v + '</label>' +
      '</div>'
    ).join('') +
    '</div>' +
    '<div class="chart-filter-dd-footer">' +
      '<span class="chart-filter-count" id="cft-vcount-' + rid + '">' + allowed.size + ' of ' + allVals.length + ' selected</span>' +
      '<div style="display:flex;gap:6px;">' +
        '<button class="chart-filter-clear" onclick="_cftValClear(this)">Clear</button>' +
        '<button class="chart-filter-apply" onclick="_cftValApply(this)">Apply</button>' +
      '</div>' +
    '</div>';

  document.body.appendChild(dd);
  const ddRect = dd.getBoundingClientRect();
  if (ddRect.bottom > window.innerHeight) dd.style.top = (rect.top - ddRect.height - 4) + 'px';
  dd.querySelector('.chart-filter-dd-search').focus();
}}

function _cftValSearch(input) {{
  const body = input.closest('.chart-filter-dd').querySelector('[id^="cft-vbody-"]');
  if (!body) return;
  const q = input.value.toLowerCase();
  body.querySelectorAll('.chart-filter-dd-item[data-val]').forEach(el => {{
    el.style.display = el.dataset.val.toLowerCase().includes(q) ? '' : 'none';
  }});
}}

function _cftToggleAll(cb) {{
  const body = cb.closest('.chart-filter-dd').querySelector('[id^="cft-vbody-"]');
  if (!body) return;
  body.querySelectorAll('input[type=checkbox][value]').forEach(c => c.checked = cb.checked);
  _cftUpdateCount(cb.closest('.chart-filter-dd'));
}}

function _cftValCheck(cb) {{
  const dd    = cb.closest('.chart-filter-dd');
  const body  = dd.querySelector('[id^="cft-vbody-"]');
  const allCb = dd.querySelector('[id^="cft-all-"]');
  if (body && allCb) allCb.checked = [...body.querySelectorAll('input[value]')].every(c => c.checked);
  _cftUpdateCount(dd);
}}

function _cftUpdateCount(dd) {{
  const body  = dd.querySelector('[id^="cft-vbody-"]');
  const count = dd.querySelector('[id^="cft-vcount-"]');
  if (!body || !count) return;
  const total   = body.querySelectorAll('input[value]').length;
  const checked = body.querySelectorAll('input[value]:checked').length;
  count.textContent = checked + ' of ' + total + ' selected';
}}

function _cftValApply(btn) {{
  const dd   = btn.closest('.chart-filter-dd');
  const iid  = dd.dataset.iid;
  const rid  = dd.dataset.rid;
  const body = dd.querySelector('[id^="cft-vbody-"]');
  const checked = new Set([...body.querySelectorAll('input[value]:checked')].map(cb => cb.value));
  const row = (_getCftRows(iid)).find(r => r.rid === rid);
  if (row) row.allowed = checked.size > 0 ? checked : null;
  _closeCftDd();
  _cftRender(iid);
  _autoRegenerate(iid);
}}

function _cftValClear(btn) {{
  const dd  = btn.closest('.chart-filter-dd');
  const iid = dd.dataset.iid;
  const rid = dd.dataset.rid;
  const row = (_getCftRows(iid)).find(r => r.rid === rid);
  if (row) row.allowed = null;
  _closeCftDd();
  _cftRender(iid);
  _autoRegenerate(iid);
}}

// ── APPLY ALL CHART FILTERS TO ROWS ──
function _applyChartFilter(iid, rows) {{
  const filters = _getCftRows(iid).filter(r => r.colIdx >= 0 && r.allowed !== null);
  if (filters.length === 0) return rows;
  return rows.filter(row =>
    filters.every(flt => flt.allowed.has(String(row[flt.colIdx])))
  );
}}
// ══════════════════════════════════════════════
// BATCH 2 FEATURES
// ══════════════════════════════════════════════

// ── COLUMN RESIZE ──
let _resizeState = null;
function _colResizeStart(e, colIdx, tableId) {{
  e.preventDefault(); e.stopPropagation();
  const th = e.target.closest('th');
  if (!th) return;
  _resizeState = {{ th, startX: e.clientX, startW: th.offsetWidth, colIdx, tableId }};
  document.addEventListener('mousemove', _colResizeMove);
  document.addEventListener('mouseup',   _colResizeEnd);
}}
function _colResizeMove(e) {{
  if (!_resizeState) return;
  const w = Math.max(50, _resizeState.startW + e.clientX - _resizeState.startX);
  _resizeState.th.style.minWidth = w + 'px';
  _resizeState.th.style.width    = w + 'px';
}}
function _colResizeEnd() {{
  _resizeState = null;
  document.removeEventListener('mousemove', _colResizeMove);
  document.removeEventListener('mouseup',   _colResizeEnd);
}}

// ── STATISTICAL SUMMARY PANEL ──
const _statPanelState = {{}};  // tableId → visible

function toggleStatPanel(tableId) {{
  _statPanelState[tableId] = !_statPanelState[tableId];
  const panel = document.getElementById('stat-panel-' + tableId);
  if (!panel) return;
  panel.classList.toggle('visible', !!_statPanelState[tableId]);
  if (_statPanelState[tableId]) _buildStatPanel(tableId);
}}

function _buildStatPanel(tableId) {{
  const panel = document.getElementById('stat-panel-' + tableId);
  if (!panel) return;
  const data = findTableData(tableId);
  if (!data) return;
  const numCols = (data.numeric_cols || []).slice(0, 8);  // max 8 cols
  if (numCols.length === 0) {{ panel.innerHTML = '<div style="color:var(--text3);font-size:11px;">No numeric columns</div>'; return; }}

  panel.innerHTML = numCols.map(ci => {{
    const vals = data.rows.map(r => parseFloat(r[ci])).filter(v => !isNaN(v));
    if (vals.length === 0) return '';
    const sum  = vals.reduce((a,b)=>a+b,0);
    const mean = sum / vals.length;
    const sorted = [...vals].sort((a,b)=>a-b);
    const median = sorted.length % 2 === 0
      ? (sorted[sorted.length/2-1] + sorted[sorted.length/2]) / 2
      : sorted[Math.floor(sorted.length/2)];
    const variance = vals.reduce((a,v)=>a+(v-mean)**2,0)/vals.length;
    const std = Math.sqrt(variance);
    const colName = String(data.cols[ci]);
    return `<div>
      <div class="stat-col-name">${{colName}}</div>
      <div class="stat-grid">
        <div class="stat-item"><div class="stat-label">Count</div><div class="stat-value">${{vals.length}}</div></div>
        <div class="stat-item"><div class="stat-label">Sum</div><div class="stat-value">${{fmtNum(String(Math.round(sum)))}}</div></div>
        <div class="stat-item"><div class="stat-label">Mean</div><div class="stat-value">${{mean.toFixed(2)}}</div></div>
        <div class="stat-item"><div class="stat-label">Median</div><div class="stat-value">${{median.toFixed(2)}}</div></div>
        <div class="stat-item"><div class="stat-label">Std Dev</div><div class="stat-value">${{std.toFixed(2)}}</div></div>
        <div class="stat-item"><div class="stat-label">Min</div><div class="stat-value">${{fmtNum(String(sorted[0]))}}</div></div>
        <div class="stat-item"><div class="stat-label">Max</div><div class="stat-value">${{fmtNum(String(sorted[sorted.length-1]))}}</div></div>
        <div class="stat-item"><div class="stat-label">P25</div><div class="stat-value">${{sorted[Math.floor(sorted.length*0.25)].toFixed(2)}}</div></div>
      </div>
    </div>`;
  }}).join('<hr style="border-color:var(--nav-border);margin:10px 0;">');
}}

// ── FILTER PRESETS ──
const _filterPresets = {{}};  // tableId → [{{name, filters}}, ...]

function saveFilterPreset(tableId) {{
  const st = getState(tableId);
  if (!st.filters || Object.keys(st.filters).length === 0) {{ showToast('No active filters to save'); return; }}
  const name = prompt('Preset name:');
  if (!name) return;
  if (!_filterPresets[tableId]) _filterPresets[tableId] = [];
  // Deep copy filters (Sets → Arrays for storage)
  const savedFilters = {{}};
  for (const [k, flt] of Object.entries(st.filters)) {{
    savedFilters[k] = flt.type === 'cat' ? {{...flt, allowed: [...flt.allowed]}} : {{...flt}};
  }}
  _filterPresets[tableId].push({{ name, filters: savedFilters }});
  _renderPresets(tableId);
  showToast('✅ Preset "' + name + '" saved');
}}

function _renderPresets(tableId) {{
  const wrap = document.getElementById('preset-wrap-' + tableId);
  if (!wrap) return;
  const presets = _filterPresets[tableId] || [];
  wrap.dataset.tid = tableId;
  wrap.innerHTML = presets.map((p,i) =>
    '<span class="preset-tag" data-tid="' + tableId + '" data-idx="' + i + '" onclick="_applyPresetEl(this)" oncontextmenu="event.preventDefault();_deletePresetEl(this)" title="Click to apply | Right-click to delete">' + p.name + '</span>'
  ).join('');
}}

function _applyPreset(tableId, idx) {{
  const preset = (_filterPresets[tableId] || [])[idx];
  if (!preset) return;
  const st = getState(tableId);
  st.filters = {{}};
  for (const [k, flt] of Object.entries(preset.filters)) {{
    st.filters[k] = flt.type === 'cat' ? {{...flt, allowed: new Set(flt.allowed)}} : {{...flt}};
  }}
  st.page = 1;
  refreshTable(tableId);
  const iids = _chartInstancesByTable[tableId] || [];
  iids.forEach(iid => _autoRegenerate(iid));
  showToast('Applied preset: ' + preset.name);
}}

function _applyPresetEl(el)  {{ _applyPreset(el.dataset.tid, parseInt(el.dataset.idx)); }}
function _deletePresetEl(el) {{ _deletePreset(el.dataset.tid, parseInt(el.dataset.idx)); }}

function _deletePreset(tableId, idx) {{
  if (!_filterPresets[tableId]) return;
  _filterPresets[tableId].splice(idx, 1);
  _renderPresets(tableId);
}}

// ── COMPUTED COLUMNS ──
const _computedCols = {{}};  // tableId → [{{name, expr}}, ...]

function addComputedCol(tableId) {{
  const data = findTableData(tableId);
  if (!data) return;
  const colNames = data.cols.join(', ');
  const name = prompt('Column name:');
  if (!name) return;
  const expr = prompt('Formula — use the EXACT column names shown below.\\nExample: REVENUE / SALES QUANTITY * 100\\n\\nAvailable columns:\\n' + colNames);
  if (!expr) return;
  if (!_computedCols[tableId]) _computedCols[tableId] = [];
  _computedCols[tableId].push({{ name: name, expr: expr }});
  refreshTable(tableId);
  _persistSoon();
  showToast('✅ Computed column "' + name + '" added');
}}

function _removeComputedCol(tableId, idx) {{
  if (_computedCols[tableId] && _computedCols[tableId][idx]) {{
    const nm = _computedCols[tableId][idx].name;
    _computedCols[tableId].splice(idx, 1);
    refreshTable(tableId);
    _persistSoon();
    showToast('Removed computed column "' + nm + '"');
  }}
}}

function _evalComputedRow(expr, row, cols) {{
  try {{
    // Map each column to a numeric context slot (c0, c1, …)
    const ctx = {{}};
    cols.forEach((col, i) => {{ const n = parseFloat(row[i]); ctx['c' + i] = isNaN(n) ? 0 : n; }});
    // Substitute the EXACT column names (longest first to avoid partial overlaps).
    // Two-phase via a null-char placeholder so a column literally named "c"/"ctx"/"c0"
    // can't corrupt the generated "(ctx.cN)" reference text. Column headers never contain a NUL char.
    let e = String(expr);
    const pairs = cols.map((col, i) => ({{ name: String(col), i }}))
                      .sort((a, b) => b.name.length - a.name.length);
    pairs.forEach(p => {{ if (p.name) e = e.split(p.name).join('\\u0000' + p.i + '\\u0000'); }});
    pairs.forEach(p => {{ e = e.split('\\u0000' + p.i + '\\u0000').join('(ctx.c' + p.i + ')'); }});
    const result = new Function('ctx', 'return (' + e + ')')(ctx);
    return (typeof result === 'number' && isFinite(result)) ? Math.round(result * 10000) / 10000 : '—';
  }} catch(err) {{ return '—'; }}
}}

// ── KEYBOARD SHORTCUTS MODAL ──
const SHORTCUTS = [
  ['E', 'Excel number format'],
  ['R', 'Raw number format'],
  ['Esc', 'Exit fullscreen / close modal'],
  ['?', 'Show this shortcuts panel'],
  ['Shift+Click column', 'Multi-column sort'],
  ['Ctrl+Click column', 'Multi-column sort (same as Shift)'],
];

let _shortcutsOpen = false;
function toggleShortcuts() {{
  _shortcutsOpen = !_shortcutsOpen;
  let modal = document.getElementById('shortcuts-modal');
  if (_shortcutsOpen) {{
    if (!modal) {{
      modal = document.createElement('div');
      modal.id = 'shortcuts-modal';
      modal.className = 'shortcuts-modal';
      modal.onclick = e => {{ if (e.target === modal) toggleShortcuts(); }};
      modal.innerHTML = `
        <div class="shortcuts-box">
          <div class="shortcuts-title">⌨ Keyboard Shortcuts</div>
          ${{SHORTCUTS.map(([k,d]) => `<div class="shortcut-row"><span class="shortcut-key">${{k}}</span><span class="shortcut-desc">${{d}}</span></div>`).join('')}}
          <div style="margin-top:14px;text-align:right;"><button class="action-btn" onclick="toggleShortcuts()">Close</button></div>
        </div>`;
      document.body.appendChild(modal);
    }}
    modal.style.display = 'flex';
  }} else if (modal) {{
    modal.style.display = 'none';
  }}
}}

// Add ? key listener
document.addEventListener('keydown', e => {{
  if (e.key === '?' && !e.target.matches('input,textarea,select')) toggleShortcuts();
  if (e.key === 'Escape' && _shortcutsOpen) toggleShortcuts();
}});

// ── CLICK BAR → FILTER TABLE ──
function _onChartClick(iid, tableId, data) {{
  if (!data || !data.points || data.points.length === 0) return;
  const pt = data.points[0];
  const xVal = pt.x !== undefined ? String(pt.x) : (pt.label !== undefined ? String(pt.label) : null);
  if (!xVal) return;

  // Find which column is the X axis
  const xColEl = document.getElementById('xcol-' + iid);
  if (!xColEl) return;
  const xColIdx = parseInt(xColEl.value);

  // Apply as a categorical filter on that column
  const st = getState(tableId);
  if (!st.filters) st.filters = {{}};
  // Toggle: if already filtered to this value, clear; else set
  const existing = st.filters[xColIdx];
  if (existing && existing.type === 'cat' && existing.allowed.size === 1 && existing.allowed.has(xVal)) {{
    delete st.filters[xColIdx];
    showToast('Filter cleared: ' + xVal);
  }} else {{
    st.filters[xColIdx] = {{ type: 'cat', allowed: new Set([xVal]) }};
    showToast('Filtered to: ' + xVal + ' (click again to clear)');
  }}
  st.page = 1;
  refreshTable(tableId);
}}

// ── COPY CHART AS IMAGE ──
function _copyChartImage(iid) {{
  const plotEl = document.getElementById('plotly-' + iid);
  if (!plotEl) {{ showToast('Generate a chart first'); return; }}
  Plotly.toImage(plotEl, {{format:'png', width:1200, height:500}}).then(dataUrl => {{
    fetch(dataUrl).then(r=>r.blob()).then(blob=>{{
      const item = new ClipboardItem({{'image/png': blob}});
      navigator.clipboard.write([item]).then(()=>showToast('✅ Chart copied to clipboard')).catch(()=>{{
        // Fallback: open in new tab
        const a = document.createElement('a'); a.href=dataUrl; a.download='chart.png'; a.click();
        showToast('Chart saved as PNG');
      }});
    }});
  }});
}}

// ── PIN CHART ──
function _togglePin(iid) {{
  const opts = _getOpt(iid);
  opts.pinned = !opts.pinned;
  const btn = document.getElementById('pin-' + iid);
  if (btn) {{
    btn.classList.toggle('active', opts.pinned);
    btn.textContent = opts.pinned ? '📌 Pin: ON' : '📌 Pin: OFF';
  }}
  const builder = document.getElementById('chartbuilder-' + iid);
  if (builder) {{
    // When pinned — collapse the builder but keep chart visible
    if (opts.pinned) {{
      const editBtn = builder.querySelector('[id^="edit-btn-"]');
      if (editBtn) {{
        const panel = document.getElementById('chartpanel-' + iid);
        if (panel && panel.style.display !== 'none') editBtn.click();
      }}
      builder.style.border = '1px solid var(--accent)';
      builder.style.boxShadow = '0 0 0 2px rgba(79,110,247,0.2)';
    }} else {{
      builder.style.border = '';
      builder.style.boxShadow = '';
    }}
  }}
}}

// ── FLEXIBLE COMBO STATE ──
const comboTypeState = {{}};  // comboTypeState[iid] = {{ left:'bar', right:'line' }}

function _getComboTypes(iid) {{
  if (!comboTypeState[iid]) comboTypeState[iid] = {{ left:'bar', right:'line' }};
  return comboTypeState[iid];
}}

function _comboColChange(input) {{
  // Find the iid from the parent container's data-iid attribute
  const container = input.closest('[data-iid]');
  if (container) _autoRegenerate(container.dataset.iid);
}}

function _setComboType(iid, side, ctype, btn) {{
  _getComboTypes(iid)[side] = ctype;
  const group = document.getElementById('combo-' + side + '-type-' + iid);
  if (group) group.querySelectorAll('.combo-type-btn').forEach(b => b.classList.toggle('active', b === btn));
  _autoRegenerate(iid);
}}

function _getComboYCols(iid, side) {{
  const container = document.getElementById('combo-' + side + '-cols-' + iid);
  if (!container) return [];
  return Array.from(container.querySelectorAll('input[data-combo-side="' + side + '"]:checked')).map(i => parseInt(i.value));
}}

function selectAllY(iid, checked) {{
  const ycols = document.getElementById('ycols-' + iid);
  if (!ycols) return;
  ycols.querySelectorAll('input[type="checkbox"]').forEach(inp => inp.checked = checked);
}}

// ── LABELS TOGGLE ──
const labelState = {{}};
function toggleLabels(iid) {{
  labelState[iid] = !labelState[iid];
  const btn = document.getElementById('label-toggle-' + iid);
  if (btn) {{
    btn.classList.toggle('active', labelState[iid]);
    btn.textContent = '🏷 Labels: ' + (labelState[iid] ? 'ON' : 'OFF');
  }}
  // Auto regenerate chart if already generated
  const chartOut = document.getElementById('chartout-' + iid);
  if (chartOut && chartOut.querySelector('[id^="plotly-"]')) {{
    const tableId = iid.split('-inst-')[0];
    generateChart(iid, tableId);
  }}
}}

const chartTypeState = {{}};

// ── CHART REGISTRY — for sidebar links ──
const chartRegistry = {{}};           // chartRegistry[tableId] = [{{iid, title}}]
const _chartInstancesByTable = {{}};  // _chartInstancesByTable[tableId] = [iid, ...]

function _registerChart(tableId, iid) {{
  if (!_chartInstancesByTable[tableId]) _chartInstancesByTable[tableId] = [];
  if (!_chartInstancesByTable[tableId].includes(iid)) _chartInstancesByTable[tableId].push(iid);
}}

function _unregisterChart(tableId, iid) {{
  if (_chartInstancesByTable[tableId]) {{
    _chartInstancesByTable[tableId] = _chartInstancesByTable[tableId].filter(i => i !== iid);
  }}
  if (chartRegistry[tableId]) {{
    chartRegistry[tableId] = chartRegistry[tableId].filter(e => e.iid !== iid);
  }}
  _refreshSidebar();
}}

function _updateChartRegistry(tableId, iid) {{
  if (!chartRegistry[tableId]) chartRegistry[tableId] = [];
  const titleEl = document.getElementById('chart-title-' + iid);
  const idx     = (_chartInstancesByTable[tableId] || []).indexOf(iid);
  const label   = (titleEl && titleEl.value.trim()) ? titleEl.value.trim() : 'Chart ' + (idx + 1);
  const existing = chartRegistry[tableId].find(e => e.iid === iid);
  if (existing) {{ existing.title = label; }}
  else          {{ chartRegistry[tableId].push({{ iid, title: label }}); }}
  _refreshSidebar();
}}

function _refreshSidebar() {{
  buildSidebar();
}}

// ── CHART OPTIONS STATE ──
const chartOptState = {{}};


function _getOpt(iid) {{
  if (!chartOptState[iid]) {{
    chartOptState[iid] = {{
      barmode: 'group', orient: 'v', corners: false,
      pattern: false, opacity: false, rangeslider: false, annotate: false,
      condcolor: false, pinned: false
    }};
  }}
  return chartOptState[iid];
}}

function setChartOpt(iid, key, val, btn) {{
  _getOpt(iid)[key] = val;
  if (btn) {{
    btn.parentElement.querySelectorAll('.chart-opt-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
  }}
  _autoRegenerate(iid);
}}

function toggleChartOpt(iid, key) {{
  const st  = _getOpt(iid);
  st[key]   = !st[key];
  const btn = document.getElementById(key + '-' + iid);
  if (btn) {{
    btn.classList.toggle('active', st[key]);
    const label = key.charAt(0).toUpperCase() + key.slice(1);
    const icons = {{corners:'⬜',pattern:'▤',opacity:'◐',rangeslider:'↔',annotate:'📌'}};
    btn.textContent = (icons[key]||'') + ' ' + label + ': ' + (st[key] ? 'ON' : 'OFF');
  }}
  _autoRegenerate(iid);
}}

function _autoRegenerate(iid) {{
  const chartOut = document.getElementById('chartout-' + iid);
  if (chartOut && chartOut.querySelector('[id^="plotly-"]')) {{
    const tableId = iid.split('-inst-')[0];
    generateChart(iid, tableId);
  }}
}}

// ── UPDATE AXIS LABEL PLACEHOLDERS ──
function _updateAxisPlaceholders(iid, tableData) {{
  if (!tableData) return;
  const cols = tableData.cols;
  const xSel = document.getElementById('xcol-' + iid);
  if (!xSel) return;
  const xIdx    = parseInt(xSel.value);
  const xName   = xIdx >= 0 ? String(cols[xIdx]) : '';
  const ycols   = document.getElementById('ycols-' + iid);
  const yIdxs   = ycols ? Array.from(ycols.querySelectorAll('input:checked')).map(i=>parseInt(i.value)) : [];
  const yName   = yIdxs.map(i => String(cols[i])).join(', ') || 'Y Axis';

  const xEl  = document.getElementById('xlabel-custom-'  + iid);
  const yEl  = document.getElementById('ylabel-custom-'  + iid);
  if (xEl && !xEl.value) xEl.placeholder = xName || 'Auto';
  if (yEl && !yEl.value) yEl.placeholder = yName || 'Auto';
}}

function selectChartType(iid, ctype, btn) {{
  chartTypeState[iid] = ctype;
  const builder = document.getElementById('chartbuilder-' + iid);
  builder.querySelectorAll('.chart-type-btn').forEach(b => b.classList.toggle('active', b === btn));

  const isPie    = ctype === 'pie' || ctype === 'donut';
  const isCombo  = ctype === 'combo';
  const isBar    = ctype === 'bar';
  const isSpecial = ctype === 'waterfall' || ctype === 'funnel' || ctype === 'heatmap';
  document.getElementById('xlabel-' + iid).textContent = isPie ? 'Labels (Dimension)' : 'X Axis (Dimension)';
  document.getElementById('ylabel-' + iid).textContent = isPie ? 'Values (pick one)' : isCombo ? 'Y Axis (pick 2+: 1st=Bar, 2nd=Line)' : 'Y Axis (Metrics)';

  // Show/hide bar options row
  const barOpts = document.getElementById('bar-opts-' + iid);
  if (barOpts) barOpts.style.display = (isBar || isCombo) ? 'flex' : 'none';
  // Special types: hide color by, bar opts already hidden
  if (isSpecial) {{
    const cb = document.getElementById('colorby-group-' + iid);
    if (cb) cb.style.display = 'none';
  }}

  // Show/hide Color By group (only for bar, line, area, scatter)
  const colorByGroup = document.getElementById('colorby-group-' + iid);
  if (colorByGroup) colorByGroup.style.display = (isPie || isCombo) ? 'none' : '';

  // Show/hide flexible combo panel + standard Y group
  const comboPanel = document.getElementById('combo-panel-' + iid);
  const ygroup     = document.getElementById('ygroup-' + iid);
  if (comboPanel) comboPanel.classList.toggle('visible', isCombo);
  if (ygroup)     ygroup.style.display = isCombo ? 'none' : '';

  const yaxisBtns = document.getElementById('yaxis-btns-' + iid);
  if (yaxisBtns) yaxisBtns.style.display = isPie ? 'none' : 'flex';
  const ycolsGroup = document.getElementById('ycols-' + iid);
  if (ycolsGroup) {{
    ycolsGroup.querySelectorAll('input').forEach(inp => {{
      if (isPie) {{
        inp.type    = 'radio';
        inp.name    = 'ypie-' + iid;
        inp.checked = false;
      }} else {{
        inp.type    = 'checkbox';
        inp.checked = true;
      }}
    }});
    if (isPie) {{
      const first = ycolsGroup.querySelector('input');
      if (first) first.checked = true;
    }}
  }}

  // Show/hide axis label row (hidden for Pie/Donut), Right Y only for Combo
  const axisLabelRow = document.getElementById('axis-label-row-'       + iid);
  const ylabel2Grp   = document.getElementById('ylabel2-custom-group-' + iid);
  const ylabelGrp    = document.getElementById('ylabel-custom-group-'  + iid);
  if (axisLabelRow) axisLabelRow.style.display = isPie ? 'none' : 'flex';
  if (ylabel2Grp)   ylabel2Grp.style.display   = isCombo ? '' : 'none';
  if (ylabelGrp)    ylabelGrp.style.display     = '';
}}

// ── GENERATE CHART ──
function generateChart(iid, tableId) {{
  if (typeof Plotly === 'undefined') {{
    document.getElementById('chartout-' + iid).innerHTML =
      `<div class="chart-error" style="margin:20px;padding:16px;">
        ⚠ Plotly not loaded! Ensure internet connection — Plotly loads from CDN.
      </div>`;
    return;
  }}

  const tableDataRaw = findTableData(tableId);
  if (!tableDataRaw) return;

  const ctype      = chartTypeState[iid] || 'bar';
  const isPie      = ctype === 'pie' || ctype === 'donut';
  const isCombo    = ctype === 'combo';
  const isWaterfall = ctype === 'waterfall';
  const isFunnel   = ctype === 'funnel';
  const isHeatmap  = ctype === 'heatmap';
  const isBubble   = ctype === 'bubble';
  const isBox      = ctype === 'box';
  const isTreemap  = ctype === 'treemap';
  const isSunburst = ctype === 'sunburst';
  const isGauge    = ctype === 'gauge';
  const isSpecial  = isWaterfall || isFunnel || isHeatmap || isBubble || isBox || isTreemap || isSunburst || isGauge;
  const showLabels = !!labelState[iid];
  const opts       = _getOpt(iid);
  const COLORS     = _getColors(iid);
  const PATTERNS   = ['/','\\\\','x','-','|','+','.'];

  // ── READ UI OPTIONS ──
  const xColIdx    = parseInt(document.getElementById('xcol-' + iid).value);
  const ycols      = document.getElementById('ycols-' + iid);
  const yColIdxs   = Array.from(ycols.querySelectorAll('input:checked')).map(i => parseInt(i.value));
  const colorByEl  = document.getElementById('colorbycol-' + iid);
  const colorByIdx = colorByEl ? parseInt(colorByEl.value) : -1;
  const useColorBy = !isPie && !isCombo && colorByIdx >= 0 && colorByIdx !== xColIdx;
  const sortMode   = (document.getElementById('sort-'      + iid)||{{}}).value || 'original';
  const topN       = (document.getElementById('topn-'      + iid)||{{}}).value || 'all';
  const yScale     = (document.getElementById('yscale-'    + iid)||{{}}).value || 'linear';
  const hoverMode  = (document.getElementById('hovermode-' + iid)||{{}}).value || 'closest';
  const colorMode  = (document.getElementById('colormode-' + iid)||{{}}).value || 'flat';
  const aggMode    = (document.getElementById('aggmode-'   + iid)||{{}}).value || 'none';
  const chartBg    = (document.getElementById('chartbg-'   + iid)||{{}}).value || 'theme';
  const chartGrid  = (document.getElementById('chartgrid-' + iid)||{{}}).value || 'on';
  const chartFont  = (document.getElementById('chartfont-' + iid)||{{}}).value || 'medium';
  const chartBorderW = (document.getElementById('chartborder-' + iid)||{{}}).value || 'none';

  if (yColIdxs.length === 0) {{ showToast('Select at least one Y axis column!'); return; }}
  // Combo uses its own left/right panels — no check on yColIdxs needed

  // ── FILTERED ROWS (uses column filters + search) ──
  let rows = applyFilters(tableId, tableDataRaw.rows);
  // also apply search filter
  const st = getState(tableId);
  if (st.search) {{
    const q = st.search.toLowerCase();
    rows = rows.filter(r => r.some(c => String(c).toLowerCase().includes(q)));
  }}

  // Apply chart-level filter (on top of table column filters)
  rows = _applyChartFilter(iid, rows);

  // Strip null X values (Bug 1 fix)
  rows = rows.filter(r => {{
    const v = r[xColIdx];
    return v !== null && v !== '' && String(v) !== 'null' && String(v) !== 'None';
  }});

  // Filter Grand Total
  rows = rows.filter(row => !isGrandTotal(row[xColIdx]));

  // ── EMPTY GUARD (Bug 9) ──
  if (rows.length === 0) {{
    document.getElementById('chartout-' + iid).innerHTML =
      `<div style="padding:40px;text-align:center;color:var(--text3);">
        📭 No data to chart.<br>
        <span style="font-size:11px;">All rows filtered out — adjust your table filters.</span>
      </div>`;
    return;
  }}

  // ── AUTO CHRONOLOGICAL SORT for time/date X (only when Sort = Original) ──
  // Category axes keep data order; our upstream groupby can emit month-then-year.
  {{
    const _xcn = String(tableDataRaw.cols[xColIdx]);
    const _xsample = rows.map(r => r[xColIdx]);
    if (sortMode === 'original' && (_isTimeCol(_xcn) || _valsLookDate(_xsample))) {{
      rows = [...rows].sort((a,b) => _chronoCmp(a[xColIdx], b[xColIdx]));
    }}
  }}

  // ── AGGREGATION ──
  function _aggregate(rowsIn, xIdx, yIdx, colorIdx) {{
    if (aggMode === 'none') return rowsIn;
    const key = r => colorIdx >= 0 ? String(r[xIdx]) + '||' + String(r[colorIdx]) : String(r[xIdx]);
    const grouped = {{}};
    rowsIn.forEach(r => {{
      const k = key(r);
      if (!grouped[k]) grouped[k] = {{ xVal: r[xIdx], colorVal: colorIdx>=0?r[colorIdx]:null, vals: [] }};
      const n = parseFloat(r[yIdx]);
      if (!isNaN(n)) grouped[k].vals.push(n);
    }});
    return Object.values(grouped).map(g => {{
      const vals = g.vals;
      let yVal = null;
      if (vals.length > 0) {{
        let raw;
        if (aggMode === 'sum')   raw = vals.reduce((a,b)=>a+b,0);
        else if (aggMode === 'avg')   raw = vals.reduce((a,b)=>a+b,0)/vals.length;
        else if (aggMode === 'count') raw = vals.length;
        else if (aggMode === 'max')   raw = Math.max(...vals);
        else if (aggMode === 'min')   raw = Math.min(...vals);
        yVal = _smartRound(raw, aggMode);
      }}
      const synth = [...rowsIn[0]];
      synth[xIdx] = g.xVal;
      synth[yIdx] = yVal;
      if (colorIdx >= 0) synth[colorIdx] = g.colorVal;
      return synth;
    }});
  }}

  // ── SORT ──
  if (sortMode !== 'original' && yColIdxs.length > 0) {{
    const syi = yColIdxs[0];
    const _isTimeX = _isTimeCol(String(tableDataRaw.cols[xColIdx])) || _valsLookDate(rows.slice(0,50).map(r => r[xColIdx]));
    if (_isTimeX) {{
      // Time/date X axis — always sort chronologically; desc reverses order
      rows = [...rows].sort((a,b) => _chronoCmp(a[xColIdx], b[xColIdx]));
      if (sortMode === 'desc') rows.reverse();
    }} else {{
      if (sortMode === 'asc')  rows = [...rows].sort((a,b)=>(parseFloat(a[syi])||0)-(parseFloat(b[syi])||0));
      if (sortMode === 'desc') rows = [...rows].sort((a,b)=>(parseFloat(b[syi])||0)-(parseFloat(a[syi])||0));
      if (sortMode === 'az')   rows = [...rows].sort((a,b)=>String(a[xColIdx]).localeCompare(String(b[xColIdx])));
    }}
  }}

  // ── TOP N ──
  if (topN !== 'all') {{
    if (topN === 'b5') {{
      rows = [...rows].sort((a,b)=>(parseFloat(a[yColIdxs[0]])||0)-(parseFloat(b[yColIdxs[0]])||0)).slice(0,5);
    }} else {{
      const n = parseInt(topN);
      if (!isNaN(n)) rows = [...rows].sort((a,b)=>(parseFloat(b[yColIdxs[0]])||0)-(parseFloat(a[yColIdxs[0]])||0)).slice(0,n);
    }}
  }}

  // ── CHART MAX ROWS CAP (Bug 11) — only when NOT aggregating ──
  // With Aggregation ON, rows collapse to unique X (x Color By), so the raw cap is
  // unnecessary AND harmful (it would truncate before _aggregate sums all rows).
  if (aggMode === 'none' && rows.length > CHART_MAX_ROWS) {{
    const _tot = rows.length;
    showToast('Showing first ' + CHART_MAX_ROWS.toLocaleString() + ' of ' + _tot.toLocaleString() + ' rows. Turn on Aggregation to use all rows, or raise chart_max_rows.');
    rows = rows.slice(0, CHART_MAX_ROWS);
  }}

  // Warn Sum on PCT col
  if (aggMode === 'sum') {{
    yColIdxs.forEach(yi => {{
      if (isPctCol(String(tableDataRaw.cols[yi]))) showToast('⚠ Sum on PCT column may not be meaningful — consider Avg');
    }});
  }}

  let xVals = rows.map(r => r[xColIdx]);
  const _scType = rows.length > 2000 ? 'scattergl' : 'scatter';  // WebGL for big scatters

  // ── THEME COLORS ──
  const cssProp  = k => getComputedStyle(document.body).getPropertyValue(k).trim();
  const themeBg  = cssProp('--bg');
  const text1    = cssProp('--text1');
  const border   = cssProp('--tbl-border');
  const bgMap    = {{ theme: themeBg, white:'#ffffff', dark:'#0d1117', transparent:'rgba(0,0,0,0)' }};
  const bg       = bgMap[chartBg] || themeBg;
  const fontText = chartBg === 'white' ? '#1a1a1a' : chartBg === 'dark' ? '#e2e8f0' : text1;
  const fontSizeMap = {{ small:10, medium:12, large:14 }};
  const fontSize    = fontSizeMap[chartFont] || 12;
  const borderW     = {{ none:0, thin:0.5, thick:1.5 }}[chartBorderW] || 0;
  const gridColor   = chartGrid === 'off' ? 'transparent' : (chartBg === 'white' ? '#e2e8f0' : border);
  const xGridColor  = (chartGrid === 'h' || chartGrid === 'off') ? 'transparent' : gridColor;

  // ── MARKER BASE ──
  function _markerBase(ti, yVals) {{
    const m = {{ color: COLORS[ti % COLORS.length], opacity: opts.opacity ? 0.7 : 1 }};
    if (borderW > 0) m.line = {{ color: fontText, width: borderW }};
    if (opts.pattern) m.pattern = {{ shape: PATTERNS[ti % PATTERNS.length], solidity:0.5 }};
    if (colorMode === 'byvalue' && !useColorBy) {{
      m.color     = yVals.map(v => v);
      m.colorscale = 'Blues';
      m.showscale  = true;
      m.colorbar   = {{ tickfont:{{ color:fontText }}, outlinewidth:0 }};
    }}
    return m;
  }}

  function _textBase(fmtVals) {{
    return {{
      text:         showLabels ? fmtVals : [],
      texttemplate: showLabels ? '%{{text}}' : '',
      textposition: 'outside',
      cliponaxis:   false,
      textfont:     {{ color:fontText, size:fontSize, family:'DM Sans,sans-serif' }},
    }};
  }}

  // ── BUILD TRACES ──
  let traces = [];

  if (useColorBy) {{
    const yIdx = yColIdxs[0];
    // Color By — group by X + ColorBy
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, colorByIdx) : rows;

    // Get unique ColorBy values
    let uniqueColorVals = [...new Set(aggRows.map(r => String(r[colorByIdx])).filter(v => v && v !== 'null' && v !== 'None'))];
    if (uniqueColorVals.length > COLORBY_MAX_UNIQUE) {{
      showToast(`⚠ Color By has ${{uniqueColorVals.length}} unique values — capped at ${{COLORBY_MAX_UNIQUE}}.`);
      uniqueColorVals = uniqueColorVals.slice(0, COLORBY_MAX_UNIQUE);
    }}

    // Get unique X values (preserving order from data)
    const uniqueX = [...new Set(aggRows.map(r => r[xColIdx]).filter(v => v !== null && v !== '' && String(v) !== 'null'))];
    const isHoriz = opts.orient === 'h' && ctype === 'bar';

    traces = uniqueColorVals.map((uval, ti) => {{
      const yMap = {{}};
      aggRows.filter(r => String(r[colorByIdx]) === uval).forEach(r => {{ yMap[r[xColIdx]] = parseFloat(r[yIdx]); }});
      const yVals   = uniqueX.map(x => isNaN(yMap[x]) ? null : yMap[x]);
      const fmtVals = yVals.map(v => v !== null ? fmtNum(String(v)) : '');
      const base    = {{ name: String(uval), marker: _markerBase(ti, yVals), ..._textBase(fmtVals) }};

      if (ctype === 'bar') {{
        if (isHoriz) {{
          const t = {{ ...base, type:'bar', y:uniqueX, x:yVals, orientation:'h' }};
          if (!showLabels) {{ t.text=[]; t.texttemplate=''; }}
          return t;
        }}
        return {{ ...base, type:'bar', x:uniqueX, y:yVals }};
      }}
      if (ctype === 'line')    return {{ ...base, type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', x:uniqueX, y:yVals, textposition:'top center' }};
      if (ctype === 'scatter') return {{ ...base, type:_scType, mode:showLabels?'markers+text':'markers', x:uniqueX, y:yVals, textposition:'top center' }};
      if (ctype === 'area')    return {{ ...base, type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', fill:'tozeroy', x:uniqueX, y:yVals, textposition:'top center' }};
      return {{ ...base, type:'bar', x:uniqueX, y:yVals }};
    }});

  }} else if (isPie) {{
    const yIdx  = yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    aggRows = aggRows.filter(r => r[xColIdx] !== null && String(r[xColIdx]) !== '');
    const yVals = aggRows.map(r => {{ const v = parseFloat(r[yIdx]); return isNaN(v) ? null : v; }});
    traces = [{{
      type:'pie', labels:aggRows.map(r=>r[xColIdx]), values:yVals,
      name:String(tableDataRaw.cols[yIdx]), hole:ctype==='donut'?0.45:0,
      textinfo:showLabels?'label+value+percent':'label+percent',
      hoverinfo:'label+value+percent', textposition:'outside', automargin:true,
      marker:{{ colors:COLORS, line:{{ color:bg, width:2 }} }}
    }}];

  }} else if (isCombo) {{
    const isHoriz      = opts.orient === 'h';
    const comboTypes   = _getComboTypes(iid);
    const leftYIdxs    = _getComboYCols(iid, 'left');
    const rightYIdxs   = _getComboYCols(iid, 'right');

    if (leftYIdxs.length === 0 && rightYIdxs.length === 0) {{
      showToast('Select at least one Y column on left or right axis'); return;
    }}

    function _buildComboTrace(yIdx, side, subIdx, ctype) {{
      let aggRows   = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
      const yVals   = aggRows.map(r => {{ const v=parseFloat(r[yIdx]); return isNaN(v)?null:v; }});
      const fmtVals = yVals.map(v => v!==null?fmtNum(String(v)):'');
      const colName = String(tableDataRaw.cols[yIdx]);
      const colorIdx = side === 'left' ? subIdx : leftYIdxs.length + subIdx;
      const xData   = aggRows.map(r=>r[xColIdx]);
      const axisKey = side === 'right' ? 'y2' : 'y';

      let t;
      if (ctype === 'bar') {{
        t = {{ type:'bar', name:colName, yaxis:axisKey, marker:_markerBase(colorIdx,yVals), ..._textBase(fmtVals) }};
        if (isHoriz) {{ t.y=xData; t.x=yVals; t.orientation='h'; delete t.yaxis; t.xaxis = side==='right'?'x2':'x'; if(!showLabels){{t.text=[];t.texttemplate='';}} }}
        else         {{ t.x=xData; t.y=yVals; }}
      }} else if (ctype === 'area') {{
        t = {{ type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', fill:'tozeroy', name:colName, yaxis:axisKey,
          marker:{{ color:COLORS[colorIdx%COLORS.length], size:6 }}, line:{{ color:COLORS[colorIdx%COLORS.length], width:2.5 }},
          ..._textBase(fmtVals), textposition:'top center' }};
        if (isHoriz) {{ t.y=xData; t.x=yVals; t.xaxis=side==='right'?'x2':'x'; delete t.yaxis; }}
        else         {{ t.x=xData; t.y=yVals; }}
      }} else {{
        // line (default)
        t = {{ type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', name:colName, yaxis:axisKey,
          marker:{{ color:COLORS[colorIdx%COLORS.length], size:7 }}, line:{{ color:COLORS[colorIdx%COLORS.length], width:2.5 }},
          ..._textBase(fmtVals), textposition:'top center' }};
        if (isHoriz) {{ t.y=xData; t.x=yVals; t.xaxis=side==='right'?'x2':'x'; delete t.yaxis; }}
        else         {{ t.x=xData; t.y=yVals; }}
      }}
      return t;
    }}

    traces = [
      ...leftYIdxs.map((yIdx, si)  => _buildComboTrace(yIdx, 'left',  si, comboTypes.left)),
      ...rightYIdxs.map((yIdx, si) => _buildComboTrace(yIdx, 'right', si, comboTypes.right)),
    ];

  }} else if (isTreemap) {{
    const yIdx  = yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    const labels = aggRows.map(r=>String(r[xColIdx]));
    const vals   = aggRows.map(r=>{{const v=parseFloat(r[yIdx]);return isNaN(v)?0:Math.abs(v);}});
    traces = [{{
      type: 'treemap',
      labels: labels,
      parents: labels.map(()=>''),
      values:  vals,
      texttemplate: '%{{label}}<br>%{{value:.3s}}',
      hovertemplate: '<b>%{{label}}</b><br>'+String(tableDataRaw.cols[yIdx])+': %{{value}}<br>%{{percentRoot:.1%}}<extra></extra>',
      marker: {{ colorscale:'Blues', showscale:false }},
      textfont: {{ family:'DM Sans,sans-serif', size:fontSize }},
    }}];

  }} else if (isSunburst) {{
    const yIdx  = yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    const labels = aggRows.map(r=>String(r[xColIdx]));
    const vals   = aggRows.map(r=>{{const v=parseFloat(r[yIdx]);return isNaN(v)?0:Math.abs(v);}});
    traces = [{{
      type: 'sunburst',
      labels: labels,
      parents: labels.map(()=>''),
      values:  vals,
      hovertemplate: '<b>%{{label}}</b><br>'+String(tableDataRaw.cols[yIdx])+': %{{value}}<br>%{{percentRoot:.1%}}<extra></extra>',
      marker: {{ colors: COLORS }},
      textfont: {{ family:'DM Sans,sans-serif', size:fontSize, color:fontText }},
      leaf: {{ opacity:0.8 }},
    }}];

  }} else if (isGauge) {{
    // Gauge: shows first Y value as a speedometer (best for single-row or aggregated single value)
    const yIdx   = yColIdxs[0];
    let aggRows  = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    const gaugeVal = aggRows.length > 0 ? (parseFloat(aggRows[0][yIdx]) || 0) : 0;
    const allVals  = aggRows.map(r=>parseFloat(r[yIdx])||0).filter(v=>!isNaN(v));
    const gaugeMax = Math.max(...allVals) * 1.2 || 100;
    const gaugeMin = Math.min(0, Math.min(...allVals));
    const colName  = String(tableDataRaw.cols[yIdx]);
    traces = [{{
      type: 'indicator',
      mode: 'gauge+number+delta',
      value: gaugeVal,
      title: {{ text: colName, font: {{ color: fontText, family: 'DM Sans,sans-serif' }} }},
      number: {{ font: {{ color: fontText, family: 'DM Sans,sans-serif' }} }},
      delta: {{ reference: gaugeMax * 0.8 }},
      gauge: {{
        axis: {{ range:[gaugeMin, gaugeMax], tickfont:{{ color:fontText }}, tickcolor:gridColor }},
        bar:  {{ color: COLORS[0] }},
        bgcolor: bg,
        bordercolor: gridColor,
        steps: [
          {{ range:[gaugeMin, gaugeMax*0.5],  color:'rgba(247,90,122,0.15)' }},
          {{ range:[gaugeMax*0.5, gaugeMax*0.8], color:'rgba(247,178,79,0.15)' }},
          {{ range:[gaugeMax*0.8, gaugeMax],  color:'rgba(34,211,165,0.15)' }},
        ],
        threshold: {{ line:{{ color:COLORS[0], width:4 }}, thickness:0.75, value:gaugeMax*0.8 }},
      }},
    }}];

  }} else if (isBubble) {{
    // Bubble: X=col, Y=col, Size=col (3rd Y column)
    const xIdx  = xColIdx;
    const yIdx  = yColIdxs[0];
    const szIdx = yColIdxs[1] >= 0 ? yColIdxs[1] : yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xIdx, yIdx, -1) : rows;
    const xVals  = aggRows.map(r=>r[xIdx]);
    const yVals  = aggRows.map(r=>{{const v=parseFloat(r[yIdx]);return isNaN(v)?null:v;}});
    const szVals = aggRows.map(r=>{{const v=parseFloat(r[szIdx]);return isNaN(v)?1:Math.abs(v);}});
    const szMax  = Math.max(...szVals.filter(v=>v!==null)) || 1;
    traces = [{{
      type:'scatter', mode:'markers',
      x: xVals, y: yVals,
      name: String(tableDataRaw.cols[yIdx]),
      marker:{{
        size: szVals.map(v=>5+((v/szMax)*45)),
        color: COLORS[0], opacity:0.7,
        line:{{ color:bg, width:1 }},
        sizemode:'diameter',
      }},
      text: showLabels ? xVals.map((x,i)=>String(x)+'<br>'+fmtNum(String(yVals[i]))) : undefined,
      hovertemplate:'<b>%{{x}}</b><br>'+String(tableDataRaw.cols[yIdx])+': %{{y}}<br>Size: %{{marker.size:.1f}}<extra></extra>',
    }}];

  }} else if (isBox) {{
    // Box plot: X=category col, Y=numeric — one box per X category
    const yIdx = yColIdxs[0];
    const uniqueX = [...new Set(rows.map(r=>String(r[xColIdx])))];
    traces = uniqueX.map((xv,ti)=>{{
      const yVals = rows.filter(r=>String(r[xColIdx])===xv).map(r=>{{const v=parseFloat(r[yIdx]);return isNaN(v)?null:v;}}).filter(v=>v!==null);
      return {{
        type:'box', name:xv, y:yVals,
        marker:{{ color:COLORS[ti%COLORS.length] }},
        boxmean:true,
      }};
    }});

  }} else if (isWaterfall) {{
    const yIdx  = yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    const yVals = aggRows.map(r => {{ const v=parseFloat(r[yIdx]); return isNaN(v)?0:v; }});
    const xData = aggRows.map(r=>r[xColIdx]);
    traces = [{{
      type: 'waterfall',
      x: xData, y: yVals,
      name: String(tableDataRaw.cols[yIdx]),
      orientation: 'v',
      connector: {{ line: {{ color: border }} }},
      increasing:  {{ marker: {{ color: '#22d3a5' }} }},
      decreasing:  {{ marker: {{ color: '#f75a7a' }} }},
      totals:      {{ marker: {{ color: COLORS[0] }} }},
      textposition: showLabels ? 'outside' : 'none',
      text: showLabels ? yVals.map(v => fmtNum(String(v))) : [],
      textfont: {{ color: fontText, size: fontSize, family: 'DM Sans,sans-serif' }},
    }}];

  }} else if (isFunnel) {{
    const yIdx  = yColIdxs[0];
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
    const yVals = aggRows.map(r => {{ const v=parseFloat(r[yIdx]); return isNaN(v)?0:v; }});
    const xData = aggRows.map(r=>r[xColIdx]);
    traces = [{{
      type: 'funnel',
      y: xData, x: yVals,
      name: String(tableDataRaw.cols[yIdx]),
      textinfo: showLabels ? 'value+percent initial' : 'percent initial',
      textposition: 'inside',
      textfont: {{ color: '#fff', size: fontSize, family: 'DM Sans,sans-serif' }},
      marker: {{ color: COLORS, line: {{ width: 1, color: bg }} }},
      connector: {{ line: {{ color: border, dash: 'dot', width: 1 }} }},
    }}];

  }} else if (isHeatmap) {{
    // Heatmap: X axis = X column, Y axis = second categorical col, Color = first numeric col
    const yIdx   = yColIdxs[0];
    // Find a second categorical column for Y axis (not X axis col)
    const catCols = tableDataRaw.cols.map((col,i)=>i).filter(i => i !== xColIdx && _colType(i, tableDataRaw.rows) === 'categorical');
    const yAxisColIdx = catCols[0] >= 0 ? catCols[0] : -1;
    if (yAxisColIdx < 0) {{
      showToast('Heatmap needs a second categorical column for the Y axis');
      return;
    }}
    let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, yAxisColIdx) : rows;
    const uniqueX = [...new Set(aggRows.map(r=>String(r[xColIdx])))];
    const uniqueY = [...new Set(aggRows.map(r=>String(r[yAxisColIdx])))];
    const zMap = {{}};
    aggRows.forEach(r => {{
      const key = String(r[xColIdx]) + '|||' + String(r[yAxisColIdx]);
      const v   = parseFloat(r[yIdx]);
      zMap[key] = isNaN(v) ? null : v;
    }});
    const zVals = uniqueY.map(yv => uniqueX.map(xv => zMap[xv+'|||'+yv] ?? null));
    traces = [{{
      type: 'heatmap',
      x: uniqueX, y: uniqueY, z: zVals,
      colorscale: 'Blues',
      showscale: true,
      text: showLabels ? zVals.map(row => row.map(v => v !== null ? fmtNum(String(v)) : '')) : undefined,
      texttemplate: showLabels ? '%{{text}}' : undefined,
      hovertemplate: '<b>%{{x}}</b> × <b>%{{y}}</b><br>' + String(tableDataRaw.cols[yIdx]) + ': %{{z}}<extra></extra>',
      colorbar: {{ tickfont: {{ color: fontText }}, outlinewidth: 0 }},
    }}];

  }} else {{
    const isHoriz = opts.orient === 'h' && ctype === 'bar';
    traces = yColIdxs.map((yIdx, ti) => {{
      let aggRows = aggMode !== 'none' ? _aggregate(rows, xColIdx, yIdx, -1) : rows;
      const yVals   = aggRows.map(r => {{ const v=parseFloat(r[yIdx]); return isNaN(v)?null:v; }});
      const fmtVals = yVals.map(v => v!==null?fmtNum(String(v)):'');
      const base    = {{ name:String(tableDataRaw.cols[yIdx]), marker:_markerBase(ti,yVals), ..._textBase(fmtVals) }};
      const xData   = aggRows.map(r=>r[xColIdx]);
      if (ctype === 'bar') {{
        if (isHoriz) {{
          const t = {{ ...base, type:'bar', y:xData, x:yVals, orientation:'h' }};
          if (!showLabels) {{ t.text=[]; t.texttemplate=''; }}
          return t;
        }}
        return {{ ...base, type:'bar', x:xData, y:yVals }};
      }}
      if (ctype === 'line')    return {{ ...base, type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', x:xData, y:yVals, textposition:'top center' }};
      if (ctype === 'scatter') return {{ ...base, type:_scType, mode:showLabels?'markers+text':'markers', x:xData, y:yVals, textposition:'top center' }};
      if (ctype === 'area')    return {{ ...base, type:'scatter', mode:showLabels?'lines+markers+text':'lines+markers', fill:'tozeroy', x:xData, y:yVals, textposition:'top center' }};
      return {{ ...base, type:'bar', x:xData, y:yVals }};
    }});
  }}

  // ── LAYOUT ──
  const isHorizLayout = opts.orient === 'h' && (ctype === 'bar' || isCombo);
  const xColName = String(tableDataRaw.cols[xColIdx]);
  const yColName = yColIdxs.map(i => String(tableDataRaw.cols[i])).join(', ');

  // X axis type — read manual override first, then auto-detect
  const xAxisTypeEl  = document.getElementById('xaxis-type-' + iid);
  const xAxisTypeOverride = xAxisTypeEl ? xAxisTypeEl.value : 'auto';

  let xAxisType;
  if (xAxisTypeOverride === 'category') {{
    xAxisType = 'category';
  }} else if (xAxisTypeOverride === 'numeric') {{
    xAxisType = '-';
  }} else {{
    // Auto mode:
    //  - real date strings (YYYY-MM / YYYY-MM-DD) -> 'date' (Plotly auto-orders + time scale)
    //  - other time-like / ID-like names, or bar/line/area/scatter -> 'category'
    //  - otherwise numeric
    const isBarLineArea = (ctype === 'bar' || ctype === 'line' || ctype === 'area' || ctype === 'scatter');
    const _xLooksDate   = _valsLookDate(xVals);
    if (!isPie && !isHeatmap && !isGauge) {{
      if (_xLooksDate && !isHorizLayout) {{
        xAxisType = 'date';
      }} else if (_isTimeCol(xColName) || _isIdCol(xColName) || isBarLineArea) {{
        xAxisType = 'category';
      }} else {{
        xAxisType = '-';
      }}
    }} else {{
      xAxisType = '-';
    }}
  }}

  // ── CUSTOM AXIS LABELS (user overrides, fall back to auto col names) ──
  const _xLblEl   = document.getElementById('xlabel-custom-'  + iid);
  const _yLblEl   = document.getElementById('ylabel-custom-'  + iid);
  const _y2LblEl  = document.getElementById('ylabel2-custom-' + iid);
  const xAxisLabel  = (_xLblEl  && _xLblEl.value.trim())  ? _xLblEl.value.trim()  : xColName;
  const yAxisLabel  = (_yLblEl  && _yLblEl.value.trim())  ? _yLblEl.value.trim()  : yColName;
  const y2AxisLabel = (_y2LblEl && _y2LblEl.value.trim()) ? _y2LblEl.value.trim() : null;

  const baseLayout = {{
    paper_bgcolor: bg, plot_bgcolor: bg,
    font:   {{ color:fontText, family:'DM Sans,sans-serif', size:fontSize }},
    margin: {{ t:60, b:isHorizLayout?60:80, l:isHorizLayout?130:70, r:isCombo?70:40 }},
    legend: {{ bgcolor:'transparent', font:{{ color:fontText, size:fontSize }} }},
    hovermode: hoverMode,
  }};

  let layout;
  if (isPie) {{
    layout = baseLayout;
  }} else if (isCombo) {{
    const leftCols  = _getComboYCols(iid, 'left');
    const rightCols = _getComboYCols(iid, 'right');
    const leftAutoTitle  = leftCols.map(i  => String(tableDataRaw.cols[i])).join(', ')  || 'Left Axis';
    const rightAutoTitle = rightCols.map(i => String(tableDataRaw.cols[i])).join(', ') || 'Right Axis';
    const leftTitle  = (_yLblEl  && _yLblEl.value.trim())  ? _yLblEl.value.trim()  : leftAutoTitle;
    const rightTitle = (_y2LblEl && _y2LblEl.value.trim()) ? _y2LblEl.value.trim() : rightAutoTitle;
    layout = Object.assign(baseLayout, {{
      xaxis:  {{ title:{{ text:isHorizLayout?leftTitle:xAxisLabel }}, tickfont:{{ color:fontText,size:fontSize }}, gridcolor:xGridColor, linecolor:gridColor,
               tickangle: isHorizLayout?0:(xVals&&xVals.some&&xVals.some(v=>String(v).length>8)?-45:-35),
               ticklen:4, type:xAxisType,
               nticks: (xVals && xVals.length > 20) ? 20 : undefined,
               automargin: true }},
      yaxis:  {{ title:{{ text:isHorizLayout?xAxisLabel:leftTitle }}, tickfont:{{ color:fontText,size:fontSize }}, gridcolor:gridColor, linecolor:gridColor, type:yScale }},
      barmode: opts.barmode,
    }});
    if (isHorizLayout) {{
      layout.xaxis2 = {{ title:{{ text:rightTitle }}, overlaying:'x', side:'top', tickfont:{{ color:fontText,size:fontSize }}, gridcolor:'transparent' }};
    }} else {{
      layout.yaxis2 = {{ title:{{ text:rightTitle }}, overlaying:'y', side:'right', tickfont:{{ color:fontText,size:fontSize }}, gridcolor:'transparent' }};
    }}
  }} else {{
    layout = Object.assign(baseLayout, {{
      xaxis: {{ title:{{ text:isHorizLayout?yAxisLabel:xAxisLabel }}, tickfont:{{ color:fontText,size:fontSize }}, gridcolor:xGridColor, linecolor:gridColor,
               // Bug 3: auto-rotate for long labels, increase bottom margin
               tickangle: isHorizLayout?0:(xVals&&xVals.some&&xVals.some(v=>String(v).length>8)?-45:-35),
               ticklen:4, type:xAxisType,
               // Auto-limit visible labels when >20 data points to prevent congestion
               nticks: (xVals && xVals.length > 20) ? 20 : undefined,
               automargin: true }},
      yaxis: {{ title:{{ text:isHorizLayout?xAxisLabel:yAxisLabel }}, tickfont:{{ color:fontText,size:fontSize }}, gridcolor:gridColor, linecolor:gridColor, type:(yScale==='log'&&traces.some(t=>(t.y||t.x||[]).some(v=>parseFloat(v)<=0)))?'linear':yScale,  // Bug 12: negative values incompatible with log
               // Bug 2: if all Y values identical, add ±10% padding so chart isn't a flat line
               range: (()=>{{ if(isPie||isCombo) return undefined; const ys=traces.flatMap(t=>t.y||t.x||[]).map(v=>parseFloat(v)).filter(v=>!isNaN(v)); if(!ys.length) return undefined; const mn=Math.min(...ys),mx=Math.max(...ys); if(mn===mx){{ const pad=Math.abs(mx)*0.1||1; return [mn-pad,mx+pad]; }} return undefined; }})() }},
      barmode: ctype==='bar' ? opts.barmode : 'relative',
    }});
  }}

  // ── REFERENCE LINE ──
  const refLineEl  = document.getElementById('refline-' + iid);
  const refLineVal = refLineEl ? parseFloat(refLineEl.value) : NaN;
  if (!isNaN(refLineVal) && !isPie && !isHeatmap && !isFunnel) {{
    if (!layout.shapes) layout.shapes = [];
    layout.shapes.push({{
      type: 'line', xref: 'paper', x0: 0, x1: 1,
      yref: 'y', y0: refLineVal, y1: refLineVal,
      line: {{ color: '#f7b24f', width: 2, dash: 'dash' }},
    }});
    if (!layout.annotations) layout.annotations = [];
    layout.annotations.push({{
      xref: 'paper', x: 1.01, xanchor: 'left',
      yref: 'y', y: refLineVal,
      text: fmtNum(String(refLineVal)),
      showarrow: false,
      font: {{ color: '#f7b24f', size: 11, family: 'DM Sans,sans-serif' }},
    }});
  }}

  // ── CONDITIONAL COLORS (red/green threshold) ──
  if (opts.condcolor && !isPie && !isHeatmap && !isFunnel && !isNaN(refLineVal)) {{
    traces.forEach(trace => {{
      if (trace.type === 'bar') {{
        const vals = (isHorizLayout ? trace.x : trace.y) || [];
        trace.marker = {{ ...trace.marker,
          color: vals.map(v => (parseFloat(v)||0) >= refLineVal ? '#22d3a5' : '#f75a7a'),
        }};
      }}
    }});
  }}

  // Range slider
  if (opts.rangeslider && !isPie) {{
    if (isHorizLayout) layout.yaxis.rangeslider = {{ visible:true }};
    else               {{ layout.xaxis.rangeslider = {{ visible:true }}; layout.margin.b=40; }}
  }}

  // Auto annotate (Bug 9 guard: check non-null)
  if (opts.annotate && !isPie && yColIdxs.length > 0) {{
    const firstYIdx = yColIdxs[0];
    const yVals     = rows.map(r => parseFloat(r[firstYIdx])).filter(v => !isNaN(v));
    if (yVals.length > 0) {{
      const allY  = rows.map(r => {{ const v=parseFloat(r[firstYIdx]); return isNaN(v)?null:v; }});
      const maxIdx = allY.indexOf(Math.max(...yVals));
      const minIdx = allY.indexOf(Math.min(...yVals));
      layout.annotations = [];
      if (maxIdx >= 0) layout.annotations.push({{ x:isHorizLayout?allY[maxIdx]:xVals[maxIdx], y:isHorizLayout?xVals[maxIdx]:allY[maxIdx], text:'▲ Max: '+fmtNum(String(allY[maxIdx])), showarrow:true, arrowhead:2, arrowcolor:'#22d3a5', font:{{ color:'#22d3a5',size:11,family:'DM Sans,sans-serif' }}, bgcolor:bg, bordercolor:'#22d3a5', borderwidth:1, borderpad:4 }});
      if (minIdx >= 0 && minIdx !== maxIdx) layout.annotations.push({{ x:isHorizLayout?allY[minIdx]:xVals[minIdx], y:isHorizLayout?xVals[minIdx]:allY[minIdx], text:'▼ Min: '+fmtNum(String(allY[minIdx])), showarrow:true, arrowhead:2, arrowcolor:'#f75a7a', font:{{ color:'#f75a7a',size:11,family:'DM Sans,sans-serif' }}, bgcolor:bg, bordercolor:'#f75a7a', borderwidth:1, borderpad:4 }});
    }}
  }}

  const config = {{
    responsive:true, displaylogo:false, scrollZoom:true,
    toImageButtonOptions:{{ format:'png', filename:'chart' }},
    modeBarButtonsToRemove:['select2d','lasso2d','autoScale2d']
  }};

  // Title + description
  const titleVal = (document.getElementById('chart-title-' + iid)||{{}}).value?.trim() || '';
  const descVal  = (document.getElementById('chart-desc-'  + iid)||{{}}).value?.trim() || '';
  const titleDiv = document.getElementById('chart-title-display-' + iid);
  if (titleDiv) {{
    titleDiv.innerHTML = '';
    if (titleVal) {{ const t=document.createElement('div'); t.className='chart-rendered-title'; t.textContent=titleVal; titleDiv.appendChild(t); }}
    if (descVal)  {{ const d=document.createElement('div'); d.className='chart-rendered-desc';  d.textContent=descVal;  titleDiv.appendChild(d); }}
  }}

  // Update sidebar chart registry
  _updateChartRegistry(tableId, iid);

  // Show filter info badge if filters active
  const outDiv = document.getElementById('chartout-' + iid);
  const filteredCount = rows.length;
  const totalCount    = tableDataRaw.rows.filter(r => !isGrandTotal(r[xColIdx])).length;
  const cftActive = _getCftRows(iid).some(r => r.colIdx >= 0 && r.allowed !== null);
  const cftBadgeParts = _getCftRows(iid)
    .filter(r => r.colIdx >= 0 && r.allowed !== null)
    .map(r => (r._colName||String(tableDataRaw.cols[r.colIdx])) + ' = ' + [...r.allowed].slice(0,3).join(', ') + (r.allowed.size > 3 ? ' +' + (r.allowed.size-3) + ' more' : ''));
  let filterBadge = '';
  if (filteredCount < totalCount || cftActive) {{
    filterBadge = '<div style="font-size:10px;color:var(--accent);padding:2px 12px 4px;display:flex;gap:8px;flex-wrap:wrap;">';
    if (filteredCount < totalCount) filterBadge += '<span>⚠ Filtered data (' + filteredCount + ' of ' + totalCount + ' rows)</span>';
    cftBadgeParts.forEach(p => {{ filterBadge += '<span>🔵 ' + p + '</span>'; }});
    filterBadge += '</div>';
  }}
  outDiv.innerHTML = filterBadge + '<div id="plotly-' + iid + '" style="width:100%;height:420px;"></div>';
  const plotDiv = document.getElementById('plotly-' + iid);
  // Date axis: parse + keep YYYY-MM / YYYY-MM-DD tick labels (avoid 2021-04 -> 2021)
  if (xAxisType === 'date' && layout.xaxis) {{
    layout.xaxis.type = 'date';
    layout.xaxis.tickformat = _datesHaveDay(xVals) ? '%Y-%m-%d' : '%Y-%m';
    layout.xaxis.nticks = undefined;
  }}
  // ── TRENDLINE / MOVING AVERAGE OVERLAY ──
  const _trendMode = (document.getElementById('trend-' + iid)||{{}}).value || 'none';
  if (_trendMode !== 'none' && !isPie && !isCombo && !isSpecial && traces.length > 0) {{
    const baseT = traces[0];
    const xs = isHorizLayout ? baseT.y : baseT.x;
    const ys = (isHorizLayout ? baseT.x : baseT.y || []).map(v => parseFloat(v));
    const tline = _computeTrend(_trendMode, ys);
    if (tline && xs) {{
      const tt = {{ type:'scatter', mode:'lines', name:_trendLabel(_trendMode),
        line:{{ color:'#f7b24f', width:2, dash:'dot' }}, hoverinfo:'skip' }};
      if (isHorizLayout) {{ tt.y = xs; tt.x = tline; }} else {{ tt.x = xs; tt.y = tline; }}
      traces.push(tt);
    }}
  }}

  Plotly.newPlot(plotDiv, traces, layout, config);
  // Click bar → filter table
  plotDiv.on('plotly_click', d => _onChartClick(iid, tableId, d));
  // Snapshot config so charts can be restored after a refresh
  try {{ _snapshotChart(iid, tableId); }} catch(e) {{}}
}}

// ── MULTI CHART OPEN / ADD / REMOVE ──
function findTableData(tableId) {{
  // Derived tables (pivots, comparisons) live in their own registry
  if (typeof _derivedTables !== 'undefined' && _derivedTables[tableId]) return _derivedTables[tableId];
  // Extract dictIdx from tableId prefix d[n]-tbl-...
  const dictMatch = tableId.match(/^d(\\d+)-/);
  const tidDictIdx = dictMatch ? parseInt(dictMatch[1]) : null;

  function searchSections(sections, parentTid, dictIdx) {{
    for (const sec of sections) {{
      const _prefix = parentTid ? parentTid + '-' : 'd' + dictIdx + '-tbl-';
      const tid = _prefix + sec.title.normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-zA-Z0-9]+/g,'-').replace(/^-+|-+$/g,'').replace(/[^a-z0-9-]/gi,'') || 'sec';
      if (sec.type === 1 && tid === tableId) return sec.data;
      if (sec.type === 2) {{
        const found = searchSections(sec.children, tid, dictIdx);
        if (found) return found;
      }}
    }}
    return null;
  }}
  // Search only the matching dict if we have a prefix, else search all
  if (tidDictIdx !== null) {{
    const found = searchSections(ALL_SECTIONS[tidDictIdx], null, tidDictIdx);
    if (found) return found;
  }} else {{
    for (let di = 0; di < ALL_SECTIONS.length; di++) {{
      const found = searchSections(ALL_SECTIONS[di], null, di);
      if (found) return found;
    }}
  }}
  return null;
}}

function openChartBuilder(tableId) {{
  let container = document.getElementById('charts-container-' + tableId);
  if (!container) {{
    const tblContainer = document.getElementById('tblcontainer-' + tableId);
    if (!tblContainer) return;
    container = document.createElement('div');
    container.className = 'charts-container';
    container.id = 'charts-container-' + tableId;
    tblContainer.parentElement.appendChild(container);

    // Add "+ Add Chart" button
    const addBtn = document.createElement('button');
    addBtn.className = 'add-chart-btn';
    addBtn.textContent = '+ Add Another Chart';
    addBtn.onclick = () => addChartInstance(tableId);
    container.appendChild(addBtn);
  }}
  addChartInstance(tableId);
}}

function addChartInstance(tableId) {{
  const tableData = findTableData(tableId);
  if (!tableData) {{ showToast('Could not find table data!'); return; }}

  chartInstanceCount[tableId] = (chartInstanceCount[tableId] || 0) + 1;
  const iid = tableId + '-inst-' + chartInstanceCount[tableId];

  chartTypeState[iid] = 'bar';
  labelState[iid]     = false;
  chartOptState[iid]  = {{
    barmode:'group', orient:'v',
    pattern:false, opacity:false, rangeslider:false, annotate:false,
    condcolor:false, pinned:false
  }};

  _registerChart(tableId, iid);
  chartFilterState[iid] = [];  // multi-filter rows
  _cftRowCounter[iid]   = 0;

  const container = document.getElementById('charts-container-' + tableId);
  const addBtn    = container.querySelector('.add-chart-btn');

  const div = document.createElement('div');
  div.innerHTML = buildChartBuilder(tableId, tableData, iid);
  container.insertBefore(div.firstElementChild, addBtn);
  div.firstElementChild && div.firstElementChild.scrollIntoView({{behavior:'smooth', block:'nearest'}});
}}

function removeChartInstance(iid) {{
  const el      = document.getElementById('chartbuilder-' + iid);
  const tableId = iid.split('-inst-')[0];
  if (el) el.remove();
  _unregisterChart(tableId, iid);
}}

function toggleChartBuilder(iid) {{
  const body = document.getElementById('builder-body-' + iid);
  const btn  = document.getElementById('edit-btn-' + iid);
  if (!body) return;
  const collapsed = body.classList.toggle('collapsed');
  if (btn) btn.textContent = collapsed ? '⚙ Edit' : '▲ Collapse';
}}

function toggleChartOutput(iid) {{
  const wrap = document.getElementById('chart-output-wrap-' + iid);
  const btn  = document.getElementById('chart-toggle-btn-' + iid);
  if (!wrap) return;
  const hidden = wrap.style.display === 'none';
  wrap.style.display = hidden ? '' : 'none';
  if (btn) btn.textContent = hidden ? '▼ Chart' : '▶ Chart';
}}
</script>

<script>
/* ── DATA ── */
const ALL_SECTIONS        = {sections_json};
const NAV_NAMES           = {names_json};
const DEFAULT_RPP         = {rpp_json};
const CHART_MAX_ROWS      = {chart_max_rows_json};
const COLORBY_MAX_UNIQUE  = {colorby_max_unique_json};
</script>

<script>
/* ── RUNTIME ── */
const THEME_MAP = {{
  "Dark Blue":"theme-darkblue","Carbon":"theme-carbon",
  "Midnight Green":"theme-midnight","Slate Light":"theme-slate",
  "Crimson":"theme-crimson","Night":"theme-night"
}};

let currentDict  = 0;
let currentView  = "expand";
let currentTheme = {theme_json};
let currentNumFmt = "actual"; // actual | k | m | b
let currentColFmt = "excel";  // excel  | raw

// ── COLUMN FORMAT (Excel vs Raw) ──
function isPctCol(colName) {{
  const name = String(colName).toLowerCase();
  return name.includes('pct') || name.includes('%');
}}

function fmtCol(val, colName) {{
  // Raw mode — return as-is
  if (currentColFmt === 'raw') return val;

  const n = parseFloat(val);
  if (isNaN(n)) return val; // non-numeric — return as-is

  // PCT column formatting
  if (isPctCol(colName)) {{
    let pct = n;
    if (n >= 0 && n <= 1)   pct = n * 100;      // 0.78 → 78
    else if (n > 100)        return val;          // >100 not a pct — show raw
    return pct.toFixed(2) + '%';
  }}

  // Regular numeric — round to 0 decimals + thousand separator
  return Math.round(n).toLocaleString('en-IN');
}}

function setColFmt(fmt) {{
  currentColFmt = fmt;
  document.getElementById('colfmt-excel').classList.toggle('active', fmt === 'excel');
  document.getElementById('colfmt-raw').classList.toggle('active',   fmt === 'raw');

  document.querySelectorAll('td[data-raw][data-col]').forEach(td => {{
    const raw     = td.getAttribute('data-raw');
    const colName = td.getAttribute('data-col');
    let   display = '';

    if (fmt === 'raw') {{
      // Show exact raw value
      display = raw;
    }} else {{
      // Excel mode
      if (isPctCol(colName) || _isTimeCol(colName)) {{
        // PCT + time columns — always use raw value (ignore K/M/B)
        display = isPctCol(colName) ? fmtCol(raw, colName) : raw;
      }} else {{
        // Regular numeric — apply fmtCol (round + commas) then K/M/B on top
        const n = parseFloat(raw);
        if (!isNaN(n)) {{
          display = fmtNum(raw); // K/M/B takes priority for non-PCT
          if (currentNumFmt === 'actual') {{
            // No K/M/B — use Excel rounding + commas
            display = Math.round(n).toLocaleString('en-IN');
          }}
        }} else {{
          display = raw;
        }}
      }}
    }}

    // Update cell text safely (preserve data-bar divs)
    const dataBar = td.querySelector('.data-bar-bg');
    const cfSpan  = td.querySelector('.cf-high,.cf-low,.cf-mid');
    if (cfSpan)       cfSpan.textContent = display;
    else if (dataBar) {{
      Array.from(td.childNodes).forEach(n => {{ if (n.nodeType === 3) n.textContent = display; }});
    }} else {{
      td.textContent = display;
    }}
  }});

  // Re-render KPI scorecards in the new format
  _refreshKpis();
  // Re-render active charts
  Object.keys(chartTypeState).forEach(iid => {{
    const chartOut = document.getElementById('chartout-' + iid);
    if (chartOut && chartOut.querySelector('[id^="plotly-"]')) {{
      generateChart(iid, iid.split('-inst-')[0]);
    }}
  }});
}}

// Combined format: apply col format first then num format (K/M/B)
function applyNumFmt(colFormatted, raw, colName) {{
  // PCT + time columns — don't apply K/M/B
  if ((isPctCol(colName) || _isTimeCol(colName)) && currentColFmt === 'excel') return colFormatted;
  // Apply K/M/B on top of raw numeric value
  return fmtNum(raw);
}}

// ── NUMBER FORMAT ──
function fmtNum(val) {{
  const n = parseFloat(val);
  if (isNaN(n)) return val; // not a number — return as-is
  if (currentNumFmt === "actual") return val;

  const abs = Math.abs(n);

  if (currentNumFmt === "k") {{
    if (abs < 1000) return val;
    return (n / 1000).toFixed(2) + "K";
  }}
  if (currentNumFmt === "m") {{
    if (abs < 1000) return val;
    if (abs < 1000000) return (n / 1000).toFixed(2) + "K";
    return (n / 1000000).toFixed(2) + "M";
  }}
  if (currentNumFmt === "b") {{
    if (abs < 1000) return val;
    if (abs < 1000000) return (n / 1000).toFixed(2) + "K";
    if (abs < 1000000000) return (n / 1000000).toFixed(2) + "M";
    return (n / 1000000000).toFixed(2) + "B";
  }}
  return val;
}}

function setNumFmt(fmt) {{
  currentNumFmt = fmt;
  ["actual","k","m","b"].forEach(f => {{
    document.getElementById("fmt-" + f).classList.toggle("active", f === fmt);
  }});
  // Update all numeric non-PCT cells in-place
  document.querySelectorAll('td[data-raw][data-col]').forEach(td => {{
    const raw     = td.getAttribute('data-raw');
    const colName = td.getAttribute('data-col');
    // PCT + time cols handled by col format — skip K/M/B for them
    if ((isPctCol(colName) || _isTimeCol(colName)) && currentColFmt === 'excel') return;
    const formatted = fmtNum(raw);
    const dataBar = td.querySelector('.data-bar-bg');
    const cfSpan  = td.querySelector('.cf-high,.cf-low,.cf-mid');
    if (cfSpan) {{ cfSpan.textContent = formatted; }}
    else if (dataBar) {{
      Array.from(td.childNodes).forEach(node => {{
        if (node.nodeType === 3) node.textContent = formatted;
      }});
    }} else {{ td.textContent = formatted; }}
  }});
  // Re-render KPI scorecards in the new format
  _refreshKpis();
  // Re-render active charts
  Object.keys(chartTypeState).forEach(iid => {{
    const chartOut = document.getElementById('chartout-' + iid);
    if (chartOut && chartOut.querySelector('[id^="plotly-"]')) {{
      const tableId = iid.split('-inst-')[0];
      generateChart(iid, tableId);
    }}
  }});
}}

// Per-table state
const tableState = {{}};

// ── THEME ──
function setTheme(name) {{
  currentTheme = name;
  document.body.className = THEME_MAP[name];
  document.querySelectorAll('.theme-dot').forEach(d => d.classList.toggle('active', d.dataset.theme === name));
  // Re-render all existing charts so they pick up new theme colors
  setTimeout(() => {{
    Object.keys(chartTypeState).forEach(iid => {{
      const outDiv = document.getElementById('chartout-' + iid);
      if (outDiv && outDiv.querySelector('[id^="plotly-"]')) {{
        const tableId = iid.split('-inst-')[0];
        generateChart(iid, tableId);
      }}
    }});
  }}, 50);
}}

// ── SIDEBAR ──
// Resize every currently-visible Plotly chart to fit its container (call after a
// layout change that doesn't fire a window 'resize' event, e.g. sidebar collapse).
function _resizeAllPlots() {{
  if (typeof Plotly === 'undefined') return;
  document.querySelectorAll('[id^="plotly-"]').forEach(el => {{
    if (el.offsetParent === null) return;   // skip hidden
    try {{ Plotly.Plots.resize(el); }} catch(e) {{}}
  }});
}}

function toggleSidebar() {{
  document.getElementById('sidebar').classList.toggle('collapsed');
  // The main area widens via CSS over 0.2s — resize charts after the transition so
  // Plotly recomputes width and the graph fills the freed space.
  setTimeout(_resizeAllPlots, 230);
}}

function buildSidebar() {{
  const body = document.getElementById('sidebar-body');

  function _chartLinks(tableId) {{
    const charts = chartRegistry[tableId] || [];
    return charts.map(e => `<div class="sidebar-chart-link" onclick="document.getElementById('chart-output-wrap-${{e.iid}}')?.scrollIntoView({{behavior:'smooth',block:'nearest'}})">${{e.title}}</div>`).join('');
  }}

  function renderSidebarItems(sections, di, indent) {{
    return sections.map(sec => {{
      const pad = indent * 8;
      if (sec.type === 1) {{
        const tid      = 'd' + di + '-tbl-' + sec.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
        const links    = _chartLinks(tid);
        return `<div class="sidebar-item" style="padding-left:${{16+pad}}px;" data-dict="${{di}}" data-section="${{sec.title}}" onclick="jumpToSection(${{di}},'${{sec.title}}')">${{sec.title}}</div>${{links}}`;
      }} else {{
        const nested = renderSidebarItems(sec.children, di, indent + 1);
        return `<div class="sidebar-item" style="padding-left:${{16+pad}}px;font-weight:600;color:var(--text2);" data-dict="${{di}}" data-section="${{sec.title}}" onclick="jumpToSection(${{di}},'${{sec.title}}')">${{sec.title}}</div>
        ${{nested}}`;
      }}
    }}).join('');
  }}

  body.innerHTML = ALL_SECTIONS.map((sections, di) => {{
    return `<div class="sidebar-dict-group">
      <div class="sidebar-dict-label">${{NAV_NAMES[di]}}</div>
      ${{renderSidebarItems(sections, di, 0)}}
    </div>`;
  }}).join('');
}}

function jumpToSection(dictIdx, title) {{
  if (dictIdx !== currentDict) {{
    currentDict = dictIdx;
    document.querySelectorAll('.nav-pill').forEach((p,i) => p.classList.toggle('active', i === dictIdx));
    renderBody();
  }}
  setTimeout(() => {{
    const secId = 'sec-d' + dictIdx + '-' + title.replace(/[^a-zA-Z0-9]+/g,'-');
    const el = document.getElementById(secId);
    if (el) {{
      el.scrollIntoView({{behavior:'smooth', block:'start'}});
      // If collapsed section — expand it first
      const head = el.querySelector('.coll-head');
      const body = el.querySelector('.coll-body');
      if (head && body && !body.classList.contains('open')) head.click();
    }}
  }}, 120);
  document.querySelectorAll('.sidebar-item').forEach(el => {{
    el.classList.toggle('active', el.dataset.dict == dictIdx && el.dataset.section === title);
  }});
}}

// ── NAV ──
function buildNav() {{
  document.getElementById('nav-pills').innerHTML = NAV_NAMES.map((name,i) =>
    '<div class="nav-pill ' + (i===0?'active':'') + '" onclick="switchDict(' + i + ')">' + name + '</div>'
  ).join('');
}}

function switchDict(i) {{
  currentDict = i;
  document.querySelectorAll('.nav-pill').forEach((p,idx) => p.classList.toggle('active', idx===i));
  renderBody();
  try {{ history.replaceState(null,'','#tab'+i); }} catch(e) {{}}
  _persistSoon();
}}

// ── VIEW ──
function setView(v) {{
  currentView = v;
  document.getElementById('btn-expand').classList.toggle('active',   v==='expand');
  document.getElementById('btn-collapse').classList.toggle('active', v==='collapse');
  renderBody();
}}

// ── TABLE STATE ──
function getState(tableId) {{
  if (!tableState[tableId]) {{
    tableState[tableId] = {{
      search: '', page: 1, rpp: DEFAULT_RPP,
      sortCol: -1, sortDir: 0,
      sortKeys: [],   // multi-col sort: [{{col, dir}}, ...]
      showPct: false, showRunning: false, condFmt: false,
      hiddenCols: new Set(), databarOn: false,
      showSummary: true, sparkOn: false,
      groupByCols: [],   // hierarchical row grouping + subtotals (empty = off)
      filters: {{}}
    }};
  }}
  return tableState[tableId];
}}

// ── COLUMN TYPE DETECTION ──
const MONTH_NAMES = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec'];
function _colType(colIdx, rows) {{
  const vals = rows.map(r => r[colIdx]).filter(v => v !== null && v !== '' && v !== 'None' && v !== 'null');
  if (vals.length === 0) return 'categorical';
  const sample = vals.slice(0, 30);
  // Month name check
  const monthCount = sample.filter(v => MONTH_NAMES.includes(String(v).toLowerCase().slice(0,3))).length;
  if (monthCount / sample.length > 0.8) return 'categorical';
  // Date check
  const dateCount = sample.filter(v => !isNaN(Date.parse(v))).length;
  if (dateCount / sample.length > 0.8) {{
    const unique = new Set(vals.map(v => String(v)));
    return unique.size <= 30 ? 'categorical' : 'date';
  }}
  // Numeric check
  const numCount = sample.filter(v => !isNaN(parseFloat(v))).length;
  if (numCount / sample.length > 0.8) return 'numeric';
  return 'categorical';
}}

// ── APPLY FILTERS ──
function applyFilters(tableId, rows) {{
  const st = getState(tableId);
  if (!st.filters || Object.keys(st.filters).length === 0) return rows;
  return rows.filter(row => {{
    for (const [colIdx, flt] of Object.entries(st.filters)) {{
      const ci  = parseInt(colIdx);
      const val = row[ci];
      if (flt.type === 'cat') {{
        if (!flt.allowed.has(String(val))) return false;
      }} else if (flt.type === 'date') {{
        const ts  = Date.parse(val);
        const min = flt.min ? Date.parse(flt.min) : null;
        const max = flt.max ? Date.parse(flt.max) : null;
        if (min !== null && !isNaN(min) && ts < min) return false;
        if (max !== null && !isNaN(max) && ts > max) return false;
      }} else if (flt.type === 'num') {{
        const n = parseFloat(val);
        if (isNaN(n)) return false;
        const op = flt.op;
        if (op === '='  && n !== flt.v1) return false;
        if (op === '!=' && n === flt.v1) return false;
        if (op === '>'  && n <= flt.v1)  return false;
        if (op === '>=' && n <  flt.v1)  return false;
        if (op === '<'  && n >= flt.v1)  return false;
        if (op === '<=' && n >  flt.v1)  return false;
        if (op === 'between' && (n < flt.v1 || n > flt.v2)) return false;
      }}
    }}
    return true;
  }});
}}

function hasActiveFilters(tableId) {{
  const st = getState(tableId);
  return st.filters && Object.keys(st.filters).length > 0;
}}

// ── FILTER DROPDOWN STATE ──
let _openFilterDropdown = null;
function closeAllFilterDropdowns() {{
  if (_openFilterDropdown) {{
    _openFilterDropdown.remove();
    _openFilterDropdown = null;
  }}
}}
document.addEventListener('click', e => {{
  if (_openFilterDropdown && !_openFilterDropdown.contains(e.target) && !e.target.classList.contains('col-filter-icon')) {{
    closeAllFilterDropdowns();
  }}
}});

function openFilterDropdown(tableId, colIdx, iconEl, colName, rows) {{
  closeAllFilterDropdowns();
  const ctype     = _colType(colIdx, rows);
  const st        = getState(tableId);
  const activeFlt = st.filters[colIdx];
  const rect      = iconEl.getBoundingClientRect();

  const dd        = document.createElement('div');
  dd.className    = 'filter-dropdown';
  dd.style.top    = (rect.bottom + 4) + 'px';
  dd.style.left   = Math.min(rect.left, window.innerWidth - 290) + 'px';
  _openFilterDropdown = dd;

  if (ctype === 'categorical') {{
    const allVals = [...new Set(rows.map(r => String(r[colIdx])).filter(v => v !== '' && v !== 'null' && v !== 'None'))].sort();
    const allowed = activeFlt ? activeFlt.allowed : new Set(allVals);
    dd.innerHTML = `
      <div class="filter-dropdown-header">
        <span>${{colName}}</span>
        <span style="font-size:10px;color:var(--text3)">${{allVals.length}} unique</span>
      </div>
      <input class="filter-search-input" id="flt-search-${{tableId}}-${{colIdx}}" placeholder="Search values..." data-tid="${{tableId}}" data-ci="${{colIdx}}" oninput="_filterCatSearch(this)">
      <div class="filter-dropdown-body" id="flt-body-${{tableId}}-${{colIdx}}">
        <div class="filter-cat-item">
          <input type="checkbox" id="flt-all-${{tableId}}-${{colIdx}}" ${{allowed.size===allVals.length?'checked':''}} data-tid="${{tableId}}" data-ci="${{colIdx}}" onchange="_toggleAllCat(this)">
          <label for="flt-all-${{tableId}}-${{colIdx}}" style="font-weight:600;">Select All</label>
        </div>
        ${{allVals.map(v => '<div class="filter-cat-item" data-val="'+v+'">' +
          '<input type="checkbox" value="'+v+'" '+(allowed.has(v)?'checked':'')+' data-tid="'+tableId+'" data-ci="'+colIdx+'" onchange="_catCheckChange(this)">' +
          '<label>'+v+'</label>' +
        '</div>').join('')}}
      </div>
      <div class="filter-dropdown-footer">
        <button class="filter-action-btn filter-clear-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_clearFilter(this.dataset.tid,this.dataset.ci)">Clear</button>
        <button class="filter-action-btn filter-apply-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_applyCatFilter(this.dataset.tid,this.dataset.ci)">Apply</button>
      </div>`;
  }} else if (ctype === 'date') {{
    const allDates = [...new Set(rows.map(r => String(r[colIdx])).filter(v => v && v !== 'null'))].sort((a,b) => Date.parse(a)-Date.parse(b));
    const minDate  = activeFlt ? activeFlt.min : '';
    const maxDate  = activeFlt ? activeFlt.max : '';
    dd.innerHTML = `
      <div class="filter-dropdown-header"><span>${{colName}}</span><span style="font-size:10px;color:var(--text3)">Date Range</span></div>
      <div class="filter-num-row"><span style="font-size:11px;color:var(--text3);width:35px;">From</span>
        <input class="filter-num-input" id="flt-dmin-${{tableId}}-${{colIdx}}" style="width:130px;" placeholder="${{allDates[0]||''}}" value="${{minDate}}">
      </div>
      <div class="filter-num-row"><span style="font-size:11px;color:var(--text3);width:35px;">To</span>
        <input class="filter-num-input" id="flt-dmax-${{tableId}}-${{colIdx}}" style="width:130px;" placeholder="${{allDates[allDates.length-1]||''}}" value="${{maxDate}}">
      </div>
      <div class="filter-dropdown-footer">
        <button class="filter-action-btn filter-clear-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_clearFilter(this.dataset.tid,this.dataset.ci)">Clear</button>
        <button class="filter-action-btn filter-apply-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_applyDateFilter(this.dataset.tid,this.dataset.ci)">Apply</button>
      </div>`;
  }} else {{
    // Numeric
    const numVals  = rows.map(r => parseFloat(r[colIdx])).filter(v => !isNaN(v));
    const minV     = numVals.length ? Math.min(...numVals) : 0;
    const maxV     = numVals.length ? Math.max(...numVals) : 0;
    const curOp    = activeFlt ? activeFlt.op    : 'between';
    const curV1    = activeFlt ? activeFlt.v1    : minV;
    const curV2    = activeFlt ? activeFlt.v2    : maxV;
    dd.innerHTML = `
      <div class="filter-dropdown-header"><span>${{colName}}</span><span style="font-size:10px;color:var(--text3)">Numeric Filter</span></div>
      <div class="filter-num-row">
        <select class="filter-num-select" id="flt-op-${{tableId}}-${{colIdx}}" onchange="_numOpChange('${{tableId}}',${{colIdx}})">
          <option value="between" ${{curOp==='between'?'selected':''}}>Between</option>
          <option value=">="  ${{curOp==='>='?'selected':''}}>≥</option>
          <option value="<="  ${{curOp==='<='?'selected':''}}>≤</option>
          <option value=">"   ${{curOp==='>'?'selected':''}}>></option>
          <option value="<"   ${{curOp==='<'?'selected':''}}>< </option>
          <option value="="   ${{curOp==='='?'selected':''}}>= (exact)</option>
          <option value="!="  ${{curOp==='!='?'selected':''}}>≠</option>
        </select>
      </div>
      <div class="filter-num-row" id="flt-v1-row-${{tableId}}-${{colIdx}}">
        <span style="font-size:11px;color:var(--text3);width:35px;" id="flt-v1-lbl-${{tableId}}-${{colIdx}}">${{curOp==='between'?'Min':'Val'}}</span>
        <input class="filter-num-input" id="flt-v1-${{tableId}}-${{colIdx}}" type="number" value="${{curV1}}" step="any">
      </div>
      <div class="filter-num-row" id="flt-v2-row-${{tableId}}-${{colIdx}}" style="${{curOp!=='between'?'display:none':''}}">
        <span style="font-size:11px;color:var(--text3);width:35px;">Max</span>
        <input class="filter-num-input" id="flt-v2-${{tableId}}-${{colIdx}}" type="number" value="${{curV2}}" step="any">
      </div>
      <div style="padding:4px 12px;font-size:10px;color:var(--text3)">Range: ${{minV.toLocaleString()}} – ${{maxV.toLocaleString()}}</div>
      <div class="filter-dropdown-footer">
        <button class="filter-action-btn filter-clear-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_clearFilter(this.dataset.tid,this.dataset.ci)">Clear</button>
        <button class="filter-action-btn filter-apply-btn" data-tid="${{tableId}}" data-ci="${{colIdx}}" onclick="_applyNumFilter(this.dataset.tid,this.dataset.ci)">Apply</button>
      </div>`;
  }}

  document.body.appendChild(dd);
  // Reposition if off-screen bottom
  const ddRect = dd.getBoundingClientRect();
  if (ddRect.bottom > window.innerHeight) dd.style.top = (rect.top - ddRect.height - 4) + 'px';
}}

function _filterCatSearch(input) {{
  const tableId = input.dataset.tid;
  const colIdx  = input.dataset.ci;
  const q       = input.value.toLowerCase();
  const body    = document.getElementById('flt-body-' + tableId + '-' + colIdx);
  if (!body) return;
  body.querySelectorAll('.filter-cat-item[data-val]').forEach(el => {{
    el.style.display = el.dataset.val.toLowerCase().includes(q) ? '' : 'none';
  }});
}}

function _toggleAllCat(cb) {{
  const tableId = cb.dataset.tid;
  const colIdx  = cb.dataset.ci;
  const body    = document.getElementById('flt-body-' + tableId + '-' + colIdx);
  if (!body) return;
  body.querySelectorAll('input[type=checkbox]').forEach(c => c.checked = cb.checked);
}}

function _catCheckChange(cb) {{
  const tableId = cb.dataset.tid;
  const colIdx  = cb.dataset.ci;
  const body    = document.getElementById('flt-body-' + tableId + '-' + colIdx);
  const allCb   = document.getElementById('flt-all-' + tableId + '-' + colIdx);
  if (body && allCb) allCb.checked = [...body.querySelectorAll('input[value]')].every(c => c.checked);
}}

function _numOpChange(tableId, colIdx) {{
  const op     = document.getElementById('flt-op-' + tableId + '-' + colIdx).value;
  const v2Row  = document.getElementById('flt-v2-row-' + tableId + '-' + colIdx);
  const v1Lbl  = document.getElementById('flt-v1-lbl-' + tableId + '-' + colIdx);
  if (v2Row) v2Row.style.display = op === 'between' ? '' : 'none';
  if (v1Lbl) v1Lbl.textContent   = op === 'between' ? 'Min' : 'Val';
}}

function _applyCatFilter(tableId, colIdx) {{
  const body    = document.getElementById('flt-body-' + tableId + '-' + colIdx);
  const checked = new Set([...body.querySelectorAll('input[value]:checked')].map(cb => cb.value));
  const st      = getState(tableId);
  const allRows = findTableData(tableId);
  if (!allRows) return closeAllFilterDropdowns();
  const allVals = new Set(allRows.rows.map(r => String(r[colIdx])));
  // Only store filter if not all selected
  if (checked.size === allVals.size) {{
    delete st.filters[colIdx];
  }} else {{
    st.filters[colIdx] = {{ type: 'cat', allowed: checked }};
  }}
  closeAllFilterDropdowns();
  _onFilterChange(tableId);
}}

function _applyDateFilter(tableId, colIdx) {{
  const minVal = document.getElementById('flt-dmin-' + tableId + '-' + colIdx).value.trim();
  const maxVal = document.getElementById('flt-dmax-' + tableId + '-' + colIdx).value.trim();
  const st     = getState(tableId);
  if (!minVal && !maxVal) {{ delete st.filters[colIdx]; }}
  else {{ st.filters[colIdx] = {{ type: 'date', min: minVal, max: maxVal }}; }}
  closeAllFilterDropdowns();
  _onFilterChange(tableId);
}}

function _applyNumFilter(tableId, colIdx) {{
  const op  = document.getElementById('flt-op-'  + tableId + '-' + colIdx).value;
  const v1  = parseFloat(document.getElementById('flt-v1-' + tableId + '-' + colIdx).value);
  const v2  = parseFloat(document.getElementById('flt-v2-' + tableId + '-' + colIdx)?.value);
  const st  = getState(tableId);
  if (isNaN(v1)) {{ delete st.filters[colIdx]; }}
  else {{ st.filters[colIdx] = {{ type: 'num', op, v1, v2: isNaN(v2) ? v1 : v2 }}; }}
  closeAllFilterDropdowns();
  _onFilterChange(tableId);
}}

function _clearFilter(tableId, colIdx) {{
  const st = getState(tableId);
  delete st.filters[colIdx];
  closeAllFilterDropdowns();
  _onFilterChange(tableId);
}}

function _onFilterChange(tableId) {{
  getState(tableId).page = 1;
  refreshTable(tableId);
  // Auto re-render chart if one exists
  const instances = _chartInstancesByTable[tableId] || [];
  instances.forEach(iid => _autoRegenerate(iid));
  _refreshSidebar();
}}

// ── TABLE BUILDER ──
function buildTable(data, tableId, inFullscreen) {{
  const state   = getState(tableId);
  const cols    = data.cols;
  const allRows = data.rows;
  const summary = data.summary;

  // Apply column filters first, then search
  let rows = applyFilters(tableId, allRows);
  if (state.search) {{
    const q = state.search.toLowerCase();
    rows = rows.filter(r => r.some(c => String(c).toLowerCase().includes(q)));
  }}

  // Sort — multi-column (Shift+Click) or single
  const _sks = (state.sortKeys && state.sortKeys.length > 0)
    ? state.sortKeys
    : (state.sortCol >= 0 && state.sortDir !== 0 ? [{{col:state.sortCol,dir:state.sortDir}}] : []);
  if (_sks.length > 0) {{
    rows = [...rows].sort((a, b) => {{
      for (const sk of _sks) {{
        const cmp = _chronoCmp(a[sk.col], b[sk.col]);
        if (cmp !== 0) return cmp * sk.dir;
      }}
      return 0;
    }});
  }}

  // Group-by: make the group columns the PRIMARY (stable) sort, in order, so groups
  // stay contiguous and nest hierarchically. JS sort is stable → any user sort above
  // is preserved as the within-group order.
  const _gcols = (state.groupByCols || []).filter(c => c >= 0 && c < cols.length);
  if (_gcols.length > 0) {{
    rows = [...rows].sort((a, b) => {{
      for (const gci of _gcols) {{ const c = _chronoCmp(a[gci], b[gci]); if (c !== 0) return c; }}
      return 0;
    }});
  }}

  // Pagination
  const total    = rows.length;
  const rpp      = state.rpp;
  const pages    = Math.max(1, Math.ceil(total / rpp));
  state.page     = Math.min(state.page, pages);
  const pageRows = rows.slice((state.page-1)*rpp, state.page*rpp);

  // Visible cols
  const visibleCols = cols.map((_,i) => !state.hiddenCols.has(i));

  // Header
  let thead = '<thead><tr>';
  cols.forEach((col, i) => {{
    if (!visibleCols[i]) return;
    const sc      = state.sortCol === i;
    const hasFlt  = !!(state.filters && state.filters[i]);
    const cls     = ['sortable', i===0?'idx-th':'', sc&&state.sortDir===1?'sort-asc':'', sc&&state.sortDir===-1?'sort-desc':''].filter(Boolean).join(' ');
    const fltIcon = `<span class="col-filter-icon${{hasFlt?' active':''}}" title="Filter" onclick="event.stopPropagation();openFilterDropdown('${{tableId}}',${{i}},this,'${{String(col).replace(/'/g,'\\'')}}',(findTableData('${{tableId}}')||{{}}).rows||[])">▾</span>`;
    const skNum = (state.sortKeys||[]).findIndex(k=>k.col===i);
    const skLabel = skNum >= 0 ? `<sup style="color:var(--accent);font-size:9px;">${{skNum+1}}</sup>` : '';
    thead += `<th class="${{cls}}" onclick="sortTable('${{tableId}}',${{i}},event)" style="min-width:60px;">${{_escHtml(col)}}${{fltIcon}}${{skLabel}}<span class="sort-icon"></span><div class="col-resizer" onmousedown="_colResizeStart(event,${{i}},'${{tableId}}')"></div></th>`;
  }});
  // Extra computed columns headers
  const numCols = cols.map((c,i)=>i).filter(i=>i>0&&data.numeric_cols&&data.numeric_cols.includes(i));
  if (state.showPct && numCols.length > 0)     thead += `<th class="extra-col-hdr">% of Total</th>`;
  if (state.showRunning && numCols.length > 0) thead += `<th class="extra-col-hdr">Running Total</th>`;
  // Computed columns
  const compCols = _computedCols[tableId] || [];
  compCols.forEach((cc, ci) => {{ thead += `<th class="computed-col-hdr" title="${{_escHtml(cc.expr)}}">${{_escHtml(cc.name)}} ƒ <span style="cursor:pointer;color:#f75a7a;margin-left:3px;font-weight:700;" title="Delete this computed column" onclick="event.stopPropagation();_removeComputedCol('${{tableId}}',${{ci}})">✕</span></th>`; }});
  // Sparkline column (row-wise mini chart across numeric columns)
  if (state.sparkOn && numCols.length > 1) thead += `<th class="extra-col-hdr">✨ Spark</th>`;
  thead += '</tr></thead>';

  // Prefix sums for the Running Total column (computed once → avoids O(n²) per page)
  let _runPrefix = null;
  if (state.showRunning && numCols[0] !== undefined) {{
    const _fni = numCols[0];
    _runPrefix = new Array(rows.length);
    let _acc = 0;
    for (let _i = 0; _i < rows.length; _i++) {{ _acc += parseFloat(rows[_i]?.[_fni]) || 0; _runPrefix[_i] = _acc; }}
  }}

  // Group subtotals — for hierarchical grouping, build a sum map PER LEVEL keyed by the
  // group-value prefix (level 0..L), computed across the FULL filtered set so each
  // rollup is correct even when a group spans multiple pages.
  const _GSEP = '\\u0001';
  const _gActive = _gcols.length > 0;
  let _levelSub = null;   // _levelSub[L][prefixKey] = {{count, sums}}
  if (_gActive) {{
    _levelSub = _gcols.map(() => ({{}}));
    rows.forEach(r => {{
      let prefix = '';
      for (let L = 0; L < _gcols.length; L++) {{
        prefix = (L === 0) ? String(r[_gcols[0]]) : prefix + _GSEP + String(r[_gcols[L]]);
        let g = _levelSub[L][prefix];
        if (!g) {{ g = _levelSub[L][prefix] = {{ count: 0, sums: {{}} }}; numCols.forEach(ci => g.sums[ci] = 0); }}
        g.count++;
        numCols.forEach(ci => {{ const n = parseFloat(r[ci]); if (!isNaN(n)) g.sums[ci] += n; }});
      }}
    }});
  }}

  // Body
  let tbody = '<tbody>';
  pageRows.forEach((row, ri) => {{
    tbody += '<tr>';
    row.forEach((cell, i) => {{
      if (!visibleCols[i]) return;
      const colName = cols[i];
      // Conditional formatting — color cells based on value thresholds (top 25%=green, bottom 25%=red)
      let condFmtClass = '';
      if (state.condFmt && i > 0 && data.numeric_cols && data.numeric_cols.includes(i)) {{
        const colKey = String(i);
        const colMax = data.bar_ranges && data.bar_ranges[colKey] ? data.bar_ranges[colKey] : 1;
        const n = parseFloat(cell);
        if (!isNaN(n)) {{
          const ratio = colMax > 0 ? n / colMax : 0;
          if (ratio >= 0.75)       condFmtClass = 'cond-green';
          else if (ratio <= 0.25)  condFmtClass = 'cond-red';
          else if (ratio <= 0.5)   condFmtClass = 'cond-yellow';
        }}
      }}

      // Strip .0 from whole floats in time columns (belt+suspenders for float index)
      // BUT preserve date-like strings: 2021-01, 2021-01-15 etc.
      const _cleanTimeVal = (v) => {{
        if (_isTimeCol(colName)) {{
          const sv = String(v);
          // If value contains a hyphen → date-like string, preserve exactly
          if (sv.includes('-')) return sv;
          const n = parseFloat(sv);
          return (!isNaN(n) && n === Math.floor(n) && !sv.includes('e')) ? String(Math.floor(n)) : sv;
        }}
        return v;
      }};
      const _cell = _cleanTimeVal(cell);
      let inner = i === 0 ? _cell : (
        currentColFmt === 'excel'
          ? (isPctCol(colName)
              ? fmtCol(_cell, colName)
              : _isTimeCol(colName)
                ? _cell
                : (currentNumFmt === 'actual'
                    ? (isNaN(parseFloat(_cell)) ? _cell : Math.round(parseFloat(_cell)).toLocaleString('en-IN'))
                    : fmtNum(_cell)))
          : (_isTimeCol(colName) ? _cell : fmtNum(_cell))
      );
      let extra = '';

      // Data bar
      if (state.databarOn && i > 0 && data.bar_ranges[String(i)]) {{
        const max = data.bar_ranges[String(i)];
        const val = parseFloat(cell) || 0;
        const pct = Math.abs(val/max)*100;
        const cls = val >= 0 ? 'data-bar-pos' : 'data-bar-neg';
        extra = `<div class="data-bar-bg ${{cls}}" style="width:${{pct}}%"></div>`;
      }}

      // Conditional formatting
      if (data.cf_cols && data.cf_cols.includes(colName) && data.cf_ranges[colName]) {{
        const r   = data.cf_ranges[colName];
        const val = parseFloat(cell);
        const pct = (val - r.min) / (r.max - r.min || 1);
        if      (pct >= 0.7) extra += '<span style="display:none"></span>', inner = `<span class="cf-high">${{inner}}</span>`;
        else if (pct <= 0.3) inner = `<span class="cf-low">${{inner}}</span>`;
        else                 inner = `<span class="cf-mid">${{inner}}</span>`;
      }}

      const isNumeric = i > 0 && !isNaN(parseFloat(cell));
      const tdCls     = i === 0 ? '' : (condFmtClass ? `class="${{condFmtClass}}"` : '');
      const dataRaw   = isNumeric ? `data-raw="${{cell}}" data-col="${{colName}}"` : '';
      tbody += `<td ${{tdCls}} ${{dataRaw}}>${{extra}}${{inner}}</td>`;
    }});
    // Extra computed columns per row — order MUST match header order (%, Running, then Computed)
    if (state.showPct || state.showRunning) {{
      const firstNumIdx = numCols[0];
      if (firstNumIdx !== undefined) {{
        const rowVal = parseFloat(pageRows[ri]?.[firstNumIdx]) || 0;
        const colTotal = (data.col_totals && data.col_totals[String(firstNumIdx)]) || 1;
        if (state.showPct) {{
          const pct = colTotal !== 0 ? ((rowVal / colTotal) * 100).toFixed(1) + '%' : '—';
          tbody += `<td class="extra-col">${{pct}}</td>`;
        }}
        if (state.showRunning) {{
          // Running total across the FULL filtered+sorted dataset (prefix sum)
          const absIdx = (state.page - 1) * rpp + ri;
          const running = _runPrefix ? (_runPrefix[absIdx] || 0) : 0;
          tbody += `<td class="extra-col" title="Cumulative through this row">${{fmtNum(String(running))}}</td>`;
        }}
      }}
    }}

    // Computed column cells (rendered last, matching header placement)
    compCols.forEach(cc => {{
      const val = _evalComputedRow(cc.expr, pageRows[ri], cols);
      tbody += `<td class="computed-col-cell">${{typeof val === 'number' ? val.toLocaleString() : val}}</td>`;
    }});
    // Sparkline cell — mini bar chart of the row's numeric values
    if (state.sparkOn && numCols.length > 1) {{
      const sv = numCols.map(ci => parseFloat(row[ci])).filter(v => !isNaN(v));
      tbody += `<td class="extra-col">${{_sparkSVG(sv)}}</td>`;
    }}
    tbody += '</tr>';

    // Group subtotal rows — at a group boundary, emit a rollup for every level that ends
    // here, innermost → outermost (Excel-style outline). Subtotals reflect the whole
    // group across pages, and appear on the page where that group actually ends.
    if (_gActive && _levelSub) {{
      const absIdx  = (state.page - 1) * rpp + ri;
      const nextRow = rows[absIdx + 1];
      const _prefixAt = (rr, L) => {{ let p=''; for (let x=0; x<=L; x++) p = (x===0)?String(rr[_gcols[0]]):p+_GSEP+String(rr[_gcols[x]]); return p; }};
      for (let L = _gcols.length - 1; L >= 0; L--) {{
        const curP  = _prefixAt(row, L);
        const nextP = nextRow ? _prefixAt(nextRow, L) : null;
        if (nextRow && nextP === curP) break;   // this level (and all shallower) continue
        const g = _levelSub[L][curP] || {{ sums: {{}}, count: 0 }};
        const lblParts = []; for (let x=0; x<=L; x++) lblParts.push(String(row[_gcols[x]]));
        const indent = '&nbsp;'.repeat(L * 3);
        tbody += '<tr class="subtotal-row subtotal-lvl-' + L + '">';
        cols.forEach((cn, i) => {{
          if (!visibleCols[i]) return;
          if (i === 0) {{ tbody += '<td class="subtotal-cell subtotal-idx">' + indent + '↳ ' + _escHtml(lblParts.join(' › ')) + ' · ' + g.count + ' rows</td>'; return; }}
          if (numCols.includes(i)) {{
            const v = g.sums[i];
            let disp;
            if (isPctCol(cn))                                          disp = fmtCol(String(v), cn);
            else if (currentNumFmt === 'actual' && currentColFmt === 'excel') disp = Math.round(v).toLocaleString('en-IN');
            else                                                       disp = fmtNum(String(v));
            tbody += '<td class="subtotal-cell" data-raw="' + v + '" data-col="' + cn + '">' + disp + '</td>';
          }} else {{
            tbody += '<td class="subtotal-cell">—</td>';
          }}
        }});
        if (state.showPct && numCols.length > 0)     tbody += '<td class="subtotal-cell">—</td>';
        if (state.showRunning && numCols.length > 0) tbody += '<td class="subtotal-cell">—</td>';
        compCols.forEach(() => {{ tbody += '<td class="subtotal-cell">—</td>'; }});
        if (state.sparkOn && numCols.length > 1)      tbody += '<td class="subtotal-cell">—</td>';
        tbody += '</tr>';
      }}
    }}
  }});

  // Summary
  if (summary && state.showSummary) {{
    tbody += '<tr class="summary-row"><td class="summary-cell summary-idx">∑ Summary</td>';
    summary.forEach((val, i) => {{
      if (!visibleCols[i+1]) return;
      const colName = cols[i+1];
      let display = '—', dataAttrs = '';
      if (val !== null) {{
        if (isPctCol(colName)) {{
          display = fmtCol(String(val), colName);
        }} else if (currentNumFmt === 'actual' && currentColFmt === 'excel') {{
          display = Math.round(parseFloat(val)).toLocaleString('en-IN');
        }} else {{
          display = fmtNum(String(val));
        }}
        // data-col lets setNumFmt()/setColFmt() reformat summary totals in-place
        dataAttrs = `data-raw="${{val}}" data-col="${{colName}}"`;
      }}
      tbody += `<td class="summary-cell" ${{dataAttrs}}>${{display}}</td>`;
    }});
    // Trailing cells for the extra columns — MUST match header order (%, Running, Computed, Spark)
    if (state.showPct && numCols.length > 0)     tbody += '<td class="summary-cell">—</td>';
    if (state.showRunning && numCols.length > 0) tbody += '<td class="summary-cell">—</td>';
    compCols.forEach(() => {{ tbody += '<td class="summary-cell">—</td>'; }});
    if (state.sparkOn && numCols.length > 1)      tbody += '<td class="summary-cell">—</td>';
    tbody += '</tr>';
  }}
  tbody += '</tbody>';

  // Column toggle menu
  const colMenu = `<div class="col-toggle-menu" id="colmenu-${{tableId}}" onclick="event.stopPropagation()">
    ${{cols.map((col,i) => `<label class="col-toggle-item">
      <input type="checkbox" ${{!state.hiddenCols.has(i)?'checked':''}} onchange="toggleCol('${{tableId}}',${{i}},this.checked)"> ${{_escHtml(col)}}
    </label>`).join('')}}
  </div>`;

  // Group-by menu — pick one or more columns to group rows by (checkboxes). Grouping
  // nests left-to-right by column order, with a subtotal rollup at each level.
  const _gset = new Set(state.groupByCols || []);
  const groupMenu = `<div class="col-toggle-menu" id="groupmenu-${{tableId}}" onclick="event.stopPropagation()">
    <div style="font-size:10px;color:var(--text3);padding:2px 6px 4px;">Group by (nests by column order):</div>
    ${{cols.map((col,i) => `<label class="col-toggle-item">
      <input type="checkbox" ${{_gset.has(i)?'checked':''}} onchange="setGroupBy('${{tableId}}',${{i}},this.checked)"> ${{_escHtml(col)}}
    </label>`).join('')}}
    <label class="col-toggle-item" style="border-top:0.5px solid var(--tbl-border);margin-top:4px;color:var(--text3);"><span onclick="clearGroupBy('${{tableId}}')" style="cursor:pointer;">✕ Clear grouping</span></label>
  </div>`;

  // Pagination controls
  const paginationHtml = `<div class="pagination">
    <span class="page-info">Showing ${{Math.min((state.page-1)*rpp+1,total)}}–${{Math.min(state.page*rpp,total)}} of ${{total}} rows</span>
    <div class="page-btns">
      <button class="page-btn" onclick="changePage('${{tableId}}',-1)" ${{state.page<=1?'disabled':''}}>‹ Prev</button>
      ${{Array.from({{length:pages}},(_,i)=>i+1).filter(p=>Math.abs(p-state.page)<=2||p===1||p===pages).map((p,idx,arr)=>{{
        const prev = arr[idx-1];
        const gap  = prev && p - prev > 1 ? '<span style="padding:0 4px;color:var(--text3)">…</span>' : '';
        return gap + `<button class="page-btn ${{p===state.page?'active':''}}" onclick="goPage('${{tableId}}',${{p}})">${{p}}</button>`;
      }}).join('')}}
      <button class="page-btn" onclick="changePage('${{tableId}}',1)" ${{state.page>=pages?'disabled':''}}>Next ›</button>
    </div>
  </div>`;

  return `
  ${{data.description ? `<div class="tbl-description">${{data.description}}</div>` : ''}}
  ${{_kpiCards(data, tableId)}}
  <div class="tbl-controls">
    <div class="tbl-search-wrap">
      <span class="tbl-search-icon">🔍</span>
      <input class="tbl-search" placeholder="Filter rows…" value="${{state.search}}"
        oninput="filterTable('${{tableId}}',this.value)"/>
    </div>
    <div class="tbl-right-controls">
      <select class="rpp-select" onchange="changeRpp('${{tableId}}',this.value)">
        ${{[25,50,100].map(n=>`<option value="${{n}}" ${{state.rpp===n?'selected':''}}>${{n}} rows</option>`).join('')}}
      </select>
      <button class="action-btn ${{state.databarOn?'active':''}}" onclick="toggleDataBar('${{tableId}}')">≡ Bars</button>
      <button class="action-btn ${{state.showSummary?'active':''}}" onclick="toggleSummary('${{tableId}}')" title="Toggle Summary Row">∑</button>
      <button class="action-btn ${{state.showPct?'active':''}}" onclick="togglePct('${{tableId}}')" title="% of Total column">%</button>
      <button class="action-btn ${{state.showRunning?'active':''}}" onclick="toggleRunning('${{tableId}}')" title="Running Total column">Σ↓</button>
      <button class="action-btn ${{state.condFmt?'active':''}}" onclick="toggleCondFmt('${{tableId}}')" title="Conditional cell formatting">🎨</button>
      <button class="action-btn ${{state.sparkOn?'active':''}}" onclick="toggleSpark('${{tableId}}')" title="Row sparklines (mini chart per row)">✨ Spark</button>
      <div class="col-toggle-wrap">
        <button class="col-toggle-btn" onclick="toggleColMenu('${{tableId}}')">Columns ▾</button>
        ${{colMenu}}
      </div>
      <div class="col-toggle-wrap">
        <button class="col-toggle-btn ${{(state.groupByCols&&state.groupByCols.length)?'active':''}}" onclick="toggleGroupMenu('${{tableId}}')" title="Group rows by one or more columns, with subtotals">⊕ Group ▾</button>
        ${{groupMenu}}
      </div>
      <button class="action-btn" onclick="toggleStatPanel('${{tableId}}')" title="Statistical Summary">Σ∑</button>
      <button class="action-btn" onclick="saveFilterPreset('${{tableId}}')" title="Save current filters as preset">💾 Preset</button>
      <button class="action-btn" onclick="addComputedCol('${{tableId}}')" title="Add computed column">ƒ(x)</button>
      <button class="action-btn" onclick="copyTable('${{tableId}}')">⎘ Copy</button>
      <button class="action-btn" onclick="downloadCSV('${{tableId}}')">↓ CSV</button>
      <button class="action-btn" onclick="downloadFilteredCSV('${{tableId}}')" title="Export only the currently filtered + sorted rows">↓ Filtered</button>
      <button class="action-btn" onclick="printTable('${{tableId}}')" title="Print / Save as PDF">🖨 PDF</button>
      ${{!inFullscreen ? `<button class="action-btn" onclick="openFullscreen('${{tableId}}')">⛶ Full</button>` : ''}}
      ${{!inFullscreen ? `<button class="action-btn" onclick="openPivotBuilder('${{tableId}}')" title="Build a pivot table">⊞ Pivot</button>` : ''}}
      ${{!inFullscreen ? `<button class="action-btn" style="color:var(--accent);border-color:var(--accent);" onclick="openChartBuilder('${{tableId}}')">📊 Chart</button>` : ''}}
    </div>
  </div>
  <div class="stat-panel" id="stat-panel-${{tableId}}"></div>
  <div class="preset-wrap" id="preset-wrap-${{tableId}}" style="margin-bottom:6px;"></div>
  <div id="computed-cols-display-${{tableId}}"></div>
  <div class="tbl-outer">
    <div class="tbl-wrap" id="tblwrap-${{tableId}}">
      <table id="tbl-${{tableId}}">${{thead}}${{tbody}}</table>
    </div>
    ${{paginationHtml}}
  </div>`;
}}

// ── RENDER SECTION ──
function renderSection(sec, inFullscreen, parentTid, dictIdx) {{
  dictIdx = dictIdx !== undefined ? dictIdx : currentDict;
  const _prefix = parentTid ? parentTid + '-' : 'd' + dictIdx + '-tbl-';
  const tid = _prefix + sec.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');

  if (sec.type === 1) {{
    return `<div class="section" id="sec-d${{dictIdx}}-${{sec.title.replace(/[^a-zA-Z0-9]+/g,'-')}}">
      <div class="section-header">
        <div class="section-title">${{sec.title}} <span class="type-badge t1-badge">Table</span></div>
      </div>
      <div id="tblcontainer-${{tid}}">${{buildTable(sec.data, tid, inFullscreen)}}</div>
    </div>`;
  }} else {{
    // Recursively render children — each child can be type 1 (table) or type 2 (nested group)
    const children = sec.children.map(ch => {{
      if (ch.type === 1) {{
        const ctid = tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
        return `<div class="sub-block">
          <div class="sub-label">${{ch.title}}</div>
          <div id="tblcontainer-${{ctid}}">${{buildTable(ch.data, ctid, inFullscreen)}}</div>
        </div>`;
      }} else {{
        // Nested group — recurse with indentation
        return `<div class="sub-block">
          <div class="sub-label" style="font-size:12px;color:var(--text1);font-weight:700;border-left:3px solid var(--accent);padding-left:8px;">${{ch.title}}</div>
          <div style="padding-left:12px;">
            ${{ch.children.map(gch => renderSection(gch, inFullscreen, tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,''), dictIdx)).join('')}}
          </div>
        </div>`;
      }}
    }}).join('');
    return `<div class="section" id="sec-d${{dictIdx}}-${{sec.title.replace(/[^a-zA-Z0-9]+/g,'-')}}">
      <div class="section-header">
        <div class="section-title">${{sec.title}} <span class="type-badge t2-badge">Grouped</span></div>
      </div>
      ${{children}}
    </div>`;
  }}
}}

function renderCollapsedSection(sec, parentTid, dictIdx) {{
  dictIdx = dictIdx !== undefined ? dictIdx : currentDict;
  const _prefix = parentTid ? parentTid + '-' : 'd' + dictIdx + '-tbl-';
  const tid = _prefix + sec.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');

  if (sec.type === 1) {{
    return `<div class="coll-item" id="sec-d${{dictIdx}}-${{sec.title.replace(/[^a-zA-Z0-9]+/g,'-')}}">
      <div class="coll-head" onclick="toggleColl(this)">
        <div class="coll-title">${{sec.title}} <span class="type-badge t1-badge">Table</span></div>
        <span class="coll-arrow">▶</span>
      </div>
      <div class="coll-body">
        <div class="coll-body-inner"><div id="tblcontainer-${{tid}}">${{buildTable(sec.data, tid, false)}}</div></div>
      </div>
    </div>`;
  }} else {{
    // Recursively render children in accordion
    const accs = sec.children.map(ch => {{
      if (ch.type === 1) {{
        const ctid = tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
        return `<div class="acc-item">
          <div class="acc-head" onclick="toggleAcc(this)">${{ch.title}} <span class="acc-arrow">▶</span></div>
          <div class="acc-body"><div class="acc-body-inner">
            <div id="tblcontainer-${{ctid}}">${{buildTable(ch.data, ctid, false)}}</div>
          </div></div>
        </div>`;
      }} else {{
        // Nested group inside accordion — recurse
        const nestedTid = tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
        const nestedAccs = ch.children.map(gch => renderCollapsedSection(gch, nestedTid, dictIdx)).join('');
        return `<div class="acc-item">
          <div class="acc-head" onclick="toggleAcc(this)">${{ch.title}} <span class="type-badge t2-badge">Grouped</span> <span class="acc-arrow">▶</span></div>
          <div class="acc-body"><div class="acc-body-inner" style="padding-left:12px;">
            ${{nestedAccs}}
          </div></div>
        </div>`;
      }}
    }}).join('');
    return `<div class="coll-item" id="sec-d${{dictIdx}}-${{sec.title.replace(/[^a-zA-Z0-9]+/g,'-')}}">
      <div class="coll-head" onclick="toggleColl(this)">
        <div class="coll-title">${{sec.title}} <span class="type-badge t2-badge">Grouped</span></div>
        <span class="coll-arrow">▶</span>
      </div>
      <div class="coll-body">
        <div class="coll-body-inner">
          ${{accs}}
        </div>
      </div>
    </div>`;
  }}
}}

// ── RENDERED DICT CACHE (keyed by dictIdx + view) ──
const _renderedDicts = {{}};

function renderBody() {{
  const area = document.getElementById('scroll-area');
  const key  = currentDict + '_' + currentView;

  // Hide all cached dict containers
  area.querySelectorAll(':scope > .dict-body').forEach(el => el.style.display = 'none');

  // If already rendered, just show it + resize any Plotly instances inside
  if (_renderedDicts[key]) {{
    _renderedDicts[key].style.display = '';
    if (typeof Plotly !== 'undefined') {{
      _renderedDicts[key].querySelectorAll('[id^="plotly-"]').forEach(el => {{
        try {{ Plotly.Plots.resize(el); }} catch(e) {{}}
      }});
    }}
    return;
  }}

  // First render — build DOM and cache it
  const sections = ALL_SECTIONS[currentDict];
  const html     = currentView === 'expand'
    ? sections.map(s => renderSection(s, false, null, currentDict)).join('')
    : sections.map(s => renderCollapsedSection(s, null, currentDict)).join('');

  const wrapper       = document.createElement('div');
  wrapper.className   = 'dict-body';
  wrapper.innerHTML   = html;
  area.appendChild(wrapper);
  _renderedDicts[key] = wrapper;

  // Restore any saved charts for this tab (once per dict+view)
  if (!_chartsRestored[key]) {{
    _chartsRestored[key] = true;
    setTimeout(() => _restoreChartsFor(currentDict), 60);
  }}
}}

function _invalidateBodyCache(dictIdx, view) {{
  const key = (dictIdx !== undefined ? dictIdx : currentDict) + '_' + (view || currentView);
  if (_renderedDicts[key]) {{
    _renderedDicts[key].remove();
    delete _renderedDicts[key];
  }}
}}

// ── TABLE INTERACTIONS ──
function refreshTable(tableId, preserveFocus) {{
  const container = document.getElementById('tblcontainer-' + tableId);
  if (!container) return;

  const activeEl  = document.activeElement;
  const wasSearch = activeEl && activeEl.classList.contains('tbl-search');

  // Use recursive findTableData — works for any nesting depth
  const data = findTableData(tableId);
  if (data) {{
    container.innerHTML = buildTable(data, tableId, false);
    if (wasSearch) refocusSearch(tableId);
  }}
}}

function refocusSearch(tableId) {{
  const container = document.getElementById('tblcontainer-' + tableId);
  if (!container) return;
  const input = container.querySelector('.tbl-search');
  if (input) {{
    input.focus();
    const len = input.value.length;
    input.setSelectionRange(len, len);
  }}
}}

function filterTable(tableId, val) {{
  getState(tableId).search = val;
  getState(tableId).page   = 1;
  refreshTable(tableId);
}}

// ── % OF TOTAL / RUNNING TOTAL / COND FMT TOGGLES ──
function togglePct(tableId)     {{ const s=getState(tableId); s.showPct=!s.showPct;     refreshTable(tableId); }}
function toggleRunning(tableId) {{ const s=getState(tableId); s.showRunning=!s.showRunning; refreshTable(tableId); }}
function toggleCondFmt(tableId) {{ const s=getState(tableId); s.condFmt=!s.condFmt;    refreshTable(tableId); }}
function toggleSpark(tableId)    {{ const s=getState(tableId); s.sparkOn=!s.sparkOn;    refreshTable(tableId); _persistSoon(); }}

// ── COMPACT MODE ──
let _compactMode = false;
function toggleCompact() {{
  _compactMode = !_compactMode;
  document.body.classList.toggle('compact-mode', _compactMode);
  const btn = document.getElementById('compact-btn');
  if (btn) btn.classList.toggle('active', _compactMode);
}}

// ── DARK/LIGHT QUICK TOGGLE ──
let _prevDarkTheme = 'Dark Blue';
function toggleDarkLight() {{
  const isLight = currentTheme === 'Slate Light';
  if (isLight) {{
    setTheme(_prevDarkTheme);
  }} else {{
    _prevDarkTheme = currentTheme;
    setTheme('Slate Light');
  }}
  const btn = document.getElementById('darklight-btn');
  if (btn) btn.textContent = isLight ? '☀' : '🌙';
}}

function sortTable(tableId, colIdx, evt) {{
  const s   = getState(tableId);
  const multi = evt && (evt.shiftKey || evt.ctrlKey || evt.metaKey);
  if (multi) {{
    // Multi-column sort — Shift/Ctrl+Click adds to sort keys
    const existing = s.sortKeys.find(k => k.col === colIdx);
    if (existing) {{
      existing.dir = existing.dir === 1 ? -1 : existing.dir === -1 ? 0 : 1;
      if (existing.dir === 0) s.sortKeys = s.sortKeys.filter(k => k.col !== colIdx);
    }} else {{
      s.sortKeys.push({{ col: colIdx, dir: 1 }});
    }}
    // Keep legacy sortCol/sortDir in sync with first key
    if (s.sortKeys.length > 0) {{ s.sortCol = s.sortKeys[0].col; s.sortDir = s.sortKeys[0].dir; }}
    else {{ s.sortCol = -1; s.sortDir = 0; }}
  }} else {{
    // Single sort — clear multi keys
    s.sortKeys = [];
    if (s.sortCol === colIdx) {{
      s.sortDir = s.sortDir === 1 ? -1 : s.sortDir === -1 ? 0 : 1;
      if (s.sortDir === 0) s.sortCol = -1;
    }} else {{
      s.sortCol = colIdx; s.sortDir = 1;
    }}
  }}
  refreshTable(tableId);
}}

function changePage(tableId, delta) {{
  getState(tableId).page += delta;
  refreshTable(tableId);
}}

function goPage(tableId, p) {{
  getState(tableId).page = p;
  refreshTable(tableId);
}}

function changeRpp(tableId, val) {{
  getState(tableId).rpp  = parseInt(val);
  getState(tableId).page = 1;
  refreshTable(tableId);
}}

function toggleDataBar(tableId) {{
  const s = getState(tableId);
  s.databarOn = !s.databarOn;
  refreshTable(tableId);
}}

function toggleSummary(tableId) {{
  const s = getState(tableId);
  s.showSummary = !s.showSummary;
  refreshTable(tableId);
}}

function toggleCol(tableId, colIdx, checked) {{
  const s = getState(tableId);
  checked ? s.hiddenCols.delete(colIdx) : s.hiddenCols.add(colIdx);
  refreshTable(tableId);
  // Reopen dropdown after re-render since refreshTable rebuilds the DOM
  const menu = document.getElementById('colmenu-' + tableId);
  if (menu) menu.classList.add('show');
}}

function toggleColMenu(tableId) {{
  const menu = document.getElementById('colmenu-' + tableId);
  if (menu) menu.classList.toggle('show');
}}

function toggleGroupMenu(tableId) {{
  const menu = document.getElementById('groupmenu-' + tableId);
  if (menu) menu.classList.toggle('show');
}}

function setGroupBy(tableId, colIdx, checked) {{
  const s = getState(tableId);
  if (!Array.isArray(s.groupByCols)) s.groupByCols = [];
  const set = new Set(s.groupByCols);
  if (checked) set.add(colIdx); else set.delete(colIdx);
  // Keep grouping order = column order (left-to-right nesting)
  s.groupByCols = [...set].sort((a, b) => a - b);
  s.page = 1;
  refreshTable(tableId);
  // Keep the dropdown open so multiple columns can be ticked in one go
  const menu = document.getElementById('groupmenu-' + tableId);
  if (menu) menu.classList.add('show');
  if (typeof _persistSoon === 'function') _persistSoon();
}}

function clearGroupBy(tableId) {{
  const s = getState(tableId);
  s.groupByCols = [];
  s.page = 1;
  refreshTable(tableId);
  if (typeof _persistSoon === 'function') _persistSoon();
}}



// ── ACCORDION / COLL ──
function toggleColl(head) {{
  const body = head.nextElementSibling; const arrow = head.querySelector('.coll-arrow');
  const open = body.classList.contains('open');
  body.classList.toggle('open',!open); arrow.classList.toggle('open',!open);
}}
function toggleAcc(head) {{
  const body = head.nextElementSibling; const arrow = head.querySelector('.acc-arrow');
  const open = body.classList.contains('open');
  body.classList.toggle('open',!open); arrow.classList.toggle('open',!open);
}}

// ── FULLSCREEN ──
function openFullscreen(tableId) {{
  for (let di = 0; di < ALL_SECTIONS.length; di++) {{
    for (const sec of ALL_SECTIONS[di]) {{
      const tid = 'd' + di + '-tbl-' + sec.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
      if (sec.type === 1 && tid === tableId) {{
        document.getElementById('fullscreen-title').textContent = sec.title;
        document.getElementById('fullscreen-body').innerHTML = buildTable(sec.data, tableId, true);
        document.getElementById('fullscreen-overlay').classList.add('show');
        return;
      }}
      if (sec.type === 2) {{
        for (const ch of sec.children) {{
          const ctid = tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g,'-').replace(/[^a-z0-9-]/gi,'');
          if (ctid === tableId) {{
            document.getElementById('fullscreen-title').textContent = sec.title + ' › ' + ch.title;
            document.getElementById('fullscreen-body').innerHTML = buildTable(ch.data, tableId, true);
            document.getElementById('fullscreen-overlay').classList.add('show');
            return;
          }}
        }}
      }}
    }}
  }}
}}
function closeFullscreen() {{
  document.getElementById('fullscreen-overlay').classList.remove('show');
}}
// ── KEYBOARD SHORTCUTS ──
document.addEventListener('keydown', e => {{
  // Ignore if typing in input/textarea
  if (['INPUT','TEXTAREA','SELECT'].includes(e.target.tagName)) return;
  if (e.key === 'Escape') closeFullscreen();
  if (e.key === 'e' || e.key === 'E') {{ setColFmt('excel'); showToast('Format: Excel'); }}
  if (e.key === 'r' || e.key === 'R') {{ setColFmt('raw');   showToast('Format: Raw');   }}
}});

// ── COPY / CSV ──
function copyTable(tableId) {{
  const data = findTableData(tableId);
  if (!data) return;
  const TAB  = String.fromCharCode(9);
  const NL   = String.fromCharCode(10);
  const header = data.cols.join(TAB);
  const rows   = data.rows.map(row => row.join(TAB));
  const text   = [header, ...rows].join(NL);
  navigator.clipboard.writeText(text).then(() =>
    showToast('Copied ' + data.rows.length + ' rows to clipboard!')
  );
}}

function downloadCSV(tableId) {{
  // Get ALL data from source — not just current page from DOM
  const data = findTableData(tableId);
  if (!data) return;

  // Build CSV from full dataset
  const headers = data.cols.join(',');
  const rows    = data.rows.map(row =>
    row.map(cell => '"' + String(cell).replace(/"/g, '""') + '"').join(',')
  );
  const csv = [headers, ...rows].join('\\n');

  const encoded = 'data:text/csv;charset=utf-8,' + encodeURIComponent(csv);
  const a       = document.createElement('a');
  a.href        = encoded;
  a.download    = tableId + '.csv';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  showToast('CSV downloaded (' + data.rows.length + ' rows)!');
}}

// ── GLOBAL SEARCH ──
function globalSearch(q) {{
  const box = document.getElementById('global-results');
  if (!q || q.length < 2) {{ box.classList.remove('show'); return; }}
  const ql = q.toLowerCase();
  const hits = [];
  ALL_SECTIONS.forEach((sections, di) => {{
    function searchSec(sec, path) {{
      if (sec.type === 1) {{
        sec.data.rows.forEach(row => {{
          if (row.some(c => String(c).toLowerCase().includes(ql))) {{
            const match = row.find(c => String(c).toLowerCase().includes(ql)) || '';
            const hi    = String(match).replace(new RegExp(q,'gi'), m => `<span class="global-result-match">${{m}}</span>`);
            hits.push({{ dict: di, section: sec.title, path, row, match: hi }});
          }}
        }});
      }} else {{
        sec.children.forEach(ch => searchSec(ch, path + ' › ' + ch.title));
      }}
    }}
    sections.forEach(sec => searchSec(sec, NAV_NAMES[di] + ' › ' + sec.title));
  }});
  if (hits.length === 0) {{
    box.innerHTML = '<div class="no-results">No results found</div>';
  }} else {{
    box.innerHTML = hits.slice(0,20).map(h =>
      `<div class="global-result-item" onclick="jumpToSection(${{h.dict}},'${{h.section}}');hideGlobalResults()">
        <div>${{h.match}}</div>
        <div class="global-result-path">${{h.path}}</div>
      </div>`
    ).join('');
  }}
  box.classList.add('show');
}}

function showGlobalResults() {{
  const val = document.getElementById('global-search').value;
  if (val && val.length >= 2) document.getElementById('global-results').classList.add('show');
}}
function hideGlobalResults() {{
  document.getElementById('global-results').classList.remove('show');
  document.getElementById('global-search').value = '';
}}
document.addEventListener('click', e => {{
  if (!e.target.closest('.global-search-wrap')) hideGlobalResults();
  if (!e.target.closest('.col-toggle-wrap')) {{
    document.querySelectorAll('.col-toggle-menu').forEach(m => m.classList.remove('show'));
  }}
}});

// ── TOAST ──
function showToast(msg) {{
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 2500);
}}

// ── EXPORT CHARTS ──
function exportCharts() {{
  showToast('⏳ Fetching Plotly from CDN...');
  fetch('https://cdn.plot.ly/plotly-2.27.0.min.js')
    .then(r => r.text())
    .then(plotlyCode => {{
      showToast('⏳ Building charts report...');
      setTimeout(() => _doExportCharts(plotlyCode), 100);
    }})
    .catch(() => {{
      showToast('⚠ No internet connection — cannot export charts');
      setTimeout(() => _doExportCharts(null), 100);
    }});
}}

function _doExportCharts(plotlyCode) {{
  const bg     = getComputedStyle(document.body).getPropertyValue('--bg').trim();
  const nav    = getComputedStyle(document.body).getPropertyValue('--nav').trim();
  const text1  = getComputedStyle(document.body).getPropertyValue('--text1').trim();
  const text2  = getComputedStyle(document.body).getPropertyValue('--text2').trim();
  const text3  = getComputedStyle(document.body).getPropertyValue('--text3').trim();
  const border = getComputedStyle(document.body).getPropertyValue('--tbl-border').trim();
  const accent = getComputedStyle(document.body).getPropertyValue('--accent').trim();

  let chartBlocks = [];

  Object.keys(chartTypeState).forEach(iid => {{
    const plotDiv = document.getElementById('plotly-' + iid);
    if (!plotDiv) {{ console.log('No plotDiv for', iid); return; }}

    // Plotly stores traces on ._fullData or .data after newPlot
    const traces = plotDiv.data || plotDiv._fullData;
    if (!traces || !traces.length) {{
      console.log('No chart data for', iid, '— was chart generated?');
      return;
    }}

    const titleEl = document.getElementById('chart-title-' + iid);
    const descEl  = document.getElementById('chart-desc-'  + iid);
    const title   = titleEl ? titleEl.value.trim() : '';
    const desc    = descEl  ? descEl.value.trim()  : '';
    const tableId = iid.split('-inst-')[0];

    let sectionPath = '';
    ALL_SECTIONS.forEach((sections, di) => {{
      sections.forEach(sec => {{
        const tid = 'd' + di + '-tbl-' + sec.title.replace(/[^a-zA-Z0-9]+/g, '-').replace(/[^a-z0-9-]/gi, '');
        if (sec.type === 1 && tid === tableId) {{
          sectionPath = NAV_NAMES[di] + ' \u203a ' + sec.title;
        }}
        if (sec.type === 2) {{
          sec.children.forEach(ch => {{
            const ctid = tid + '-' + ch.title.replace(/[^a-zA-Z0-9]+/g, '-').replace(/[^a-z0-9-]/gi, '');
            if (ctid === tableId) sectionPath = NAV_NAMES[di] + ' \u203a ' + sec.title + ' \u203a ' + ch.title;
          }});
        }}
      }});
    }});

    // Get traces and layout from Plotly graph object
    const chartTraces = JSON.parse(JSON.stringify(plotDiv.data   || traces || []));
    const chartLayout = JSON.parse(JSON.stringify(plotDiv.layout || plotDiv._fullLayout || {{}}));

    chartBlocks.push({{ title, desc, sectionPath, traces: chartTraces, layout: chartLayout }});
  }});

  // Embed Plotly inline — fetched from CDN or fallback to src
  const plotlyScript = plotlyCode
    ? '<scr' + 'ipt>' + plotlyCode + '</scr' + 'ipt>'
    : '<scr' + 'ipt src="plotly.min.js"></' + 'script>';

  const timestamp = new Date().toLocaleString();
  const count     = chartBlocks.length;

  // Build chart divs — each will be rendered by Plotly in the exported HTML
  let body = '';
  chartBlocks.forEach((block, i) => {{
    const tracesJson = JSON.stringify(block.traces);
    const layoutJson = JSON.stringify(block.layout);
    body += '<div class="chart-block">';
    body += '<div class="chart-meta">' + block.sectionPath + '</div>';
    if (block.title) body += '<div class="chart-title">' + block.title + '</div>';
    if (block.desc)  body += '<div class="chart-desc">'  + block.desc  + '</div>';
    body += '<div id="chart-' + i + '" style="width:100%;height:480px;"></div>';
    body += '<scr' + 'ipt>';
    body += '(function(){{';
    body += 'var traces=' + tracesJson + ';';
    body += 'var layout=' + layoutJson + ';';
    body += 'layout.paper_bgcolor="' + bg + '";';
    body += 'layout.plot_bgcolor="'  + bg + '";';
    body += 'if(layout.font) layout.font.color="' + text1 + '";';
    body += 'if(layout.xaxis){{layout.xaxis.tickfont={{color:"' + text1 + '"}};layout.xaxis.gridcolor="' + border + '";}}';
    body += 'if(layout.yaxis){{layout.yaxis.tickfont={{color:"' + text1 + '"}};layout.yaxis.gridcolor="' + border + '";}}';
    body += 'if(layout.legend) layout.legend.font={{color:"' + text1 + '"}};';
    body += 'Plotly.newPlot("chart-' + i + '",traces,layout,{{responsive:true,displaylogo:false,scrollZoom:true}});';
    body += '}})();';
    body += '</scr' + 'ipt>';
    body += '</div>';
  }});

  const html = '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"/>'
    + '<meta name="viewport" content="width=device-width,initial-scale=1.0"/>'
    + '<title>{_title} — Charts</title>'
    + '<style>'
    + '*{{box-sizing:border-box;margin:0;padding:0;}}'
    + 'body{{font-family:Arial,sans-serif;background:' + bg + ';color:' + text1 + ';padding:32px;}}'
    + '.report-header{{text-align:center;margin-bottom:40px;padding-bottom:20px;border-bottom:1px solid ' + border + ';}}'
    + '.report-header h1{{font-size:24px;font-weight:700;color:' + text1 + ';margin-bottom:6px;}}'
    + '.report-header .subtitle{{font-size:13px;color:' + text3 + ';}}'
    + '.accent{{color:' + accent + ';}}'
    + '.chart-block{{margin-bottom:40px;background:' + nav + ';border-radius:12px;padding:24px;border:0.5px solid ' + border + ';}}'
    + '.chart-meta{{font-size:11px;color:' + text3 + ';text-transform:uppercase;letter-spacing:0.5px;margin-bottom:8px;font-weight:600;}}'
    + '.chart-title{{font-size:18px;font-weight:700;color:' + text1 + ';margin-bottom:6px;}}'
    + '.chart-desc{{font-size:13px;color:' + text2 + ';margin-bottom:16px;line-height:1.6;}}'
    + '</style>'
    + plotlyScript
    + '</head><body>'
    + '<div class="report-header">'
    + '<h1>{_title}</h1>'
    + '<div class="subtitle">Generated: ' + timestamp + ' &nbsp;|&nbsp; ' + count + ' chart' + (count !== 1 ? 's' : '') + '</div>'
    + '</div>'
    + body
    + '</body></html>';

  const encoded = 'data:text/html;charset=utf-8,' + encodeURIComponent(html);
  const a       = document.createElement('a');
  a.href        = encoded;
  a.download    = 'charts_report.html';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  setTimeout(() => showToast('charts_report.html downloaded!'), 200);
}}

// ══════════════════════════════════════════════
//  v11 ADDITIONS — persistence, sparklines, KPIs,
//  trendlines, filtered export, print, help panel
// ══════════════════════════════════════════════

function _escHtml(s) {{
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

// ── ROW SPARKLINE (inline SVG mini bar chart) ──
function _sparkSVG(values) {{
  if (!values || values.length < 2) return '';
  const w=70, h=18, n=values.length;
  const mn=Math.min(...values), mx=Math.max(...values), rng=(mx-mn)||1, bw=w/n;
  let bars='';
  for (let i=0;i<n;i++) {{
    const bh = Math.max(1, ((values[i]-mn)/rng)*(h-2));
    bars += '<rect x="'+(i*bw+0.5).toFixed(1)+'" y="'+(h-bh).toFixed(1)+'" width="'+Math.max(0.5,bw-1).toFixed(1)+'" height="'+bh.toFixed(1)+'"></rect>';
  }}
  return '<svg class="row-spark" width="'+w+'" height="'+h+'" viewBox="0 0 '+w+' '+h+'">'+bars+'</svg>';
}}

// ── KPI SCORECARDS (auto from summary row) ──
function _kpiCardsInner(data) {{
  if (!data || !data.summary) return '';
  const cards = [];
  for (let i=0; i<data.summary.length && cards.length<5; i++) {{
    if (data.summary[i] === null) continue;
    const name = String(data.cols[i+1]);
    const isPct = isPctCol(name);
    let val;
    if (isPct) {{
      const vals = data.rows.map(r=>parseFloat(r[i+1])).filter(v=>!isNaN(v));
      const avg = vals.length ? vals.reduce((a,b)=>a+b,0)/vals.length : 0;
      val = fmtCol(String(avg), name);
    }} else {{
      // Mirror the table-cell formatting: K/M/B when active, else comma-grouped (Excel) or raw
      const tot = data.summary[i];
      if (currentNumFmt !== 'actual')      val = fmtNum(String(tot));
      else if (currentColFmt === 'excel')  val = Math.round(parseFloat(tot)).toLocaleString('en-IN');
      else                                 val = String(tot);
    }}
    cards.push('<div class="kpi-card"><div class="kpi-card-label">'+_escHtml(name)+
      '</div><div class="kpi-card-value">'+val+'</div><div class="kpi-card-sub">'+(isPct?'average':'total')+'</div></div>');
  }}
  return cards.join('');
}}
function _kpiCards(data, tableId) {{
  const inner = _kpiCardsInner(data);
  if (!inner) return '';
  return '<div class="kpi-row" data-tid="'+(tableId||'')+'">'+inner+'</div>';
}}
// Re-render KPI cards in place when the number/column format changes
function _refreshKpis() {{
  document.querySelectorAll('.kpi-row[data-tid]').forEach(row => {{
    const tid = row.dataset.tid;
    if (!tid) return;
    const data = findTableData(tid);
    if (data) row.innerHTML = _kpiCardsInner(data);
  }});
}}

// ── TRENDLINE / MOVING AVERAGE ──
function _computeTrend(mode, ys) {{
  const v = ys.map(y=>isNaN(y)?null:y);
  const idx = v.map((y,i)=>y===null?-1:i).filter(i=>i>=0);
  if (idx.length < 2) return null;
  if (mode==='mean') {{ const m = idx.reduce((a,i)=>a+v[i],0)/idx.length; return v.map(()=>m); }}
  if (mode==='linear') {{
    const n=idx.length; let sx=0,sy=0,sxy=0,sxx=0;
    idx.forEach(i=>{{ sx+=i; sy+=v[i]; sxy+=i*v[i]; sxx+=i*i; }});
    const denom=(n*sxx - sx*sx)||1; const b=(n*sxy - sx*sy)/denom; const a=(sy - b*sx)/n;
    return v.map((_,i)=>a+b*i);
  }}
  const win = mode==='ma3'?3:mode==='ma5'?5:7;
  return v.map((_,i)=>{{
    let s=0,c=0; for (let j=Math.max(0,i-win+1); j<=i; j++) {{ if(v[j]!==null){{ s+=v[j]; c++; }} }}
    return c?s/c:null;
  }});
}}
function _trendLabel(mode) {{ return ({{mean:'Mean',linear:'Linear fit',ma3:'MA(3)',ma5:'MA(5)',ma7:'MA(7)'}})[mode]||'Trend'; }}

// ── EXPORT CURRENTLY FILTERED + SORTED ROWS ──
function _filteredSortedRows(tableId) {{
  const data = findTableData(tableId);
  if (!data) return {{cols:[],rows:[]}};
  const st = getState(tableId);
  let rows = applyFilters(tableId, data.rows);
  if (st.search) {{ const q=st.search.toLowerCase(); rows=rows.filter(r=>r.some(c=>String(c).toLowerCase().includes(q))); }}
  const _sks = (st.sortKeys&&st.sortKeys.length)?st.sortKeys:(st.sortCol>=0&&st.sortDir!==0?[{{col:st.sortCol,dir:st.sortDir}}]:[]);
  if (_sks.length) rows=[...rows].sort((a,b)=>{{ for(const sk of _sks){{ const c=_chronoCmp(a[sk.col],b[sk.col]); if(c!==0) return c*sk.dir; }} return 0; }});
  return {{cols:data.cols, rows}};
}}
function downloadFilteredCSV(tableId) {{
  const {{cols, rows}} = _filteredSortedRows(tableId);
  if (!cols.length) return;
  const esc = c => '"'+String(c).replace(/"/g,'""')+'"';
  const csv = [cols.map(esc).join(','), ...rows.map(r=>r.map(esc).join(','))].join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download=tableId+'_filtered.csv';
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  showToast('Filtered CSV downloaded ('+rows.length+' rows)');
}}

// ── PRINT / PDF (current filtered + sorted view) ──
function printTable(tableId) {{
  const {{cols, rows}} = _filteredSortedRows(tableId);
  if (!cols.length) return;
  let html='<html><head><title>'+_escHtml(tableId)+'</title><style>'
    +'body{{font-family:Arial,sans-serif;padding:20px;}}table{{border-collapse:collapse;width:100%;font-size:12px;}}'
    +'th,td{{border:1px solid #ccc;padding:6px 10px;text-align:right;}}th{{background:#1e2236;color:#fff;}}'
    +'td:first-child,th:first-child{{text-align:left;}}</style></head><body>';
  html+='<table><thead><tr>'+cols.map(c=>'<th>'+_escHtml(c)+'</th>').join('')+'</tr></thead><tbody>';
  rows.forEach(r=>{{ html+='<tr>'+r.map((c,i)=>'<td>'+_escHtml(i>0?fmtNum(String(c)):c)+'</td>').join('')+'</tr>'; }});
  html+='</tbody></table></body></html>';
  const wn=window.open('','_blank');
  if(!wn){{ showToast('Allow pop-ups to print/PDF'); return; }}
  wn.document.write(html); wn.document.close(); wn.focus();
  setTimeout(()=>{{ try{{ wn.print(); }}catch(e){{}} }}, 350);
}}

// ── CHART CONFIG PERSISTENCE ──
const _savedCharts    = {{}};   // _savedCharts[tableId][iid] = cfg
const _chartsRestored = {{}};

function _snapshotChart(iid, tableId) {{
  const getV = id => {{ const el=document.getElementById(id); return el?el.value:undefined; }};
  const ycEl = document.getElementById('ycols-'+iid);
  const ycols = ycEl ? Array.from(ycEl.querySelectorAll('input:checked')).map(i=>i.value) : [];
  const clEl = document.getElementById('combo-left-cols-'+iid);
  const crEl = document.getElementById('combo-right-cols-'+iid);
  const cleft  = clEl ? Array.from(clEl.querySelectorAll('input:checked')).map(i=>i.value) : [];
  const cright = crEl ? Array.from(crEl.querySelectorAll('input:checked')).map(i=>i.value) : [];
  const cft = (chartFilterState[iid]||[]).map(r=>({{ colIdx:r.colIdx, _colName:r._colName, allowed: r.allowed?[...r.allowed]:null }}));
  const cfg = {{
    ctype: chartTypeState[iid]||'bar', opt: Object.assign({{}}, _getOpt(iid)),
    palette: paletteState[iid]||'PayU', combo: Object.assign({{}}, _getComboTypes(iid)),
    label: !!labelState[iid],
    xcol:getV('xcol-'+iid), colorby:getV('colorbycol-'+iid),
    sort:getV('sort-'+iid), topn:getV('topn-'+iid), xtype:getV('xaxis-type-'+iid),
    yscale:getV('yscale-'+iid), hover:getV('hovermode-'+iid), colormode:getV('colormode-'+iid),
    agg:getV('aggmode-'+iid), bg:getV('chartbg-'+iid), grid:getV('chartgrid-'+iid),
    font:getV('chartfont-'+iid), border:getV('chartborder-'+iid), trend:getV('trend-'+iid),
    refline:getV('refline-'+iid),
    title:getV('chart-title-'+iid), desc:getV('chart-desc-'+iid),
    xlabel:getV('xlabel-custom-'+iid), ylabel:getV('ylabel-custom-'+iid), ylabel2:getV('ylabel2-custom-'+iid),
    ycols, cleft, cright, cft
  }};
  if (!_savedCharts[tableId]) _savedCharts[tableId] = {{}};
  _savedCharts[tableId][iid] = cfg;
  _persistSoon();
}}

function _applyChartCfg(iid, cfg) {{
  chartTypeState[iid] = cfg.ctype||'bar';
  chartOptState[iid]  = Object.assign({{barmode:'group',orient:'v',corners:false,pattern:false,opacity:false,rangeslider:false,annotate:false,condcolor:false,pinned:false}}, cfg.opt||{{}});
  paletteState[iid]   = cfg.palette||'PayU';
  comboTypeState[iid] = Object.assign({{left:'bar',right:'line'}}, cfg.combo||{{}});
  labelState[iid]     = !!cfg.label;
  chartFilterState[iid] = (cfg.cft||[]).map((r,i)=>({{ rid: iid+'-r'+i, colIdx:r.colIdx, _colName:r._colName, allowed: r.allowed?new Set(r.allowed):null }}));
  _cftRowCounter[iid] = (cfg.cft||[]).length;
  const setV = (id,v)=>{{ const el=document.getElementById(id); if(el && v!==undefined && v!==null) el.value=v; }};
  setV('xcol-'+iid,cfg.xcol); setV('colorbycol-'+iid,cfg.colorby);
  setV('sort-'+iid,cfg.sort); setV('topn-'+iid,cfg.topn); setV('xaxis-type-'+iid,cfg.xtype);
  setV('yscale-'+iid,cfg.yscale); setV('hovermode-'+iid,cfg.hover); setV('colormode-'+iid,cfg.colormode);
  setV('aggmode-'+iid,cfg.agg); setV('chartbg-'+iid,cfg.bg); setV('chartgrid-'+iid,cfg.grid);
  setV('chartfont-'+iid,cfg.font); setV('chartborder-'+iid,cfg.border); setV('trend-'+iid,cfg.trend);
  setV('refline-'+iid,cfg.refline);
  setV('chart-title-'+iid,cfg.title); setV('chart-desc-'+iid,cfg.desc);
  setV('xlabel-custom-'+iid,cfg.xlabel); setV('ylabel-custom-'+iid,cfg.ylabel); setV('ylabel2-custom-'+iid,cfg.ylabel2);
  // Sync UI visibility for the chart type FIRST (it resets the Y checkboxes)
  const tbtn = document.querySelector('#chartbuilder-'+iid+' .chart-type-btn[data-ctype="'+(cfg.ctype||'bar')+'"]');
  if (tbtn) selectChartType(iid, cfg.ctype||'bar', tbtn);
  // Then restore the saved column selections
  const yc = document.getElementById('ycols-'+iid);
  if (yc && cfg.ycols && cfg.ycols.length) yc.querySelectorAll('input').forEach(i=> i.checked = cfg.ycols.includes(i.value));
  const setChecks=(cid,arr)=>{{ const c=document.getElementById(cid); if(c&&arr) c.querySelectorAll('input').forEach(i=>i.checked=arr.includes(i.value)); }};
  setChecks('combo-left-cols-'+iid,cfg.cleft); setChecks('combo-right-cols-'+iid,cfg.cright);
  // Restore combo type button highlight
  ['left','right'].forEach(side=>{{
    const grp=document.getElementById('combo-'+side+'-type-'+iid);
    if(grp){{ const _ct=(comboTypeState[iid][side]||'').toLowerCase(); grp.querySelectorAll('.combo-type-btn').forEach(b=>{{ b.classList.toggle('active', !!_ct && b.textContent.toLowerCase().includes(_ct)); }}); }}
  }});
  _cftRender(iid);
  generateChart(iid, iid.split('-inst-')[0]);
}}

function _recreateChart(tableId, cfg) {{
  const before = (_chartInstancesByTable[tableId]||[]).slice();
  openChartBuilder(tableId);  // ensures container + adds one instance
  const after = (_chartInstancesByTable[tableId]||[]);
  const iid = after.find(x => !before.includes(x));
  if (iid) {{ try {{ _applyChartCfg(iid, cfg); }} catch(e) {{ console.log('chart restore failed', e); }} }}
}}

function _restoreChartsFor(di) {{
  Object.keys(_savedCharts).forEach(tableId => {{
    if (!tableId.startsWith('d'+di+'-')) return;
    if (!findTableData(tableId)) {{ if (tableId.endsWith('-pivot')) delete _savedCharts[tableId]; return; }}   // skip/cleanup ephemeral tables (pivots)
    const cfgs = Object.values(_savedCharts[tableId]||{{}});
    if (!cfgs.length) return;
    delete _savedCharts[tableId];   // repopulated by snapshots from the recreated charts
    cfgs.forEach(cfg => _recreateChart(tableId, cfg));
  }});
}}

// ── LOCALSTORAGE PERSISTENCE ──
const _DASH_KEY = 'pudbo_dashboard_state_v11';
let _persistTimer = null;
function _persistSoon() {{ clearTimeout(_persistTimer); _persistTimer = setTimeout(_persistState, 400); }}

function _persistState() {{
  try {{
    const tbl = {{}};
    Object.keys(tableState).forEach(tid => {{
      const s = tableState[tid];
      const filters = {{}};
      Object.keys(s.filters||{{}}).forEach(k => {{
        const f = s.filters[k];
        filters[k] = f.type==='cat' ? Object.assign({{}}, f, {{allowed:[...f.allowed]}}) : Object.assign({{}}, f);
      }});
      tbl[tid] = {{ search:s.search, page:s.page, rpp:s.rpp, sortCol:s.sortCol, sortDir:s.sortDir,
        sortKeys:s.sortKeys, showPct:s.showPct, showRunning:s.showRunning, condFmt:s.condFmt,
        databarOn:s.databarOn, showSummary:s.showSummary, sparkOn:s.sparkOn, groupByCols:s.groupByCols,
        hiddenCols:[...(s.hiddenCols||[])], filters }};
    }});
    const presets = {{}};
    Object.keys(_filterPresets).forEach(tid => {{
      presets[tid] = (_filterPresets[tid]||[]).map(p => {{
        const fl = {{}};
        Object.keys(p.filters||{{}}).forEach(k => {{ const f=p.filters[k]; fl[k]= f.type==='cat'?Object.assign({{}},f,{{allowed:[...f.allowed]}}):Object.assign({{}},f); }});
        return {{ name:p.name, filters:fl }};
      }});
    }});
    const payload = {{
      ui: {{ theme:currentTheme, view:currentView, numfmt:currentNumFmt, colfmt:currentColFmt, compact:_compactMode, dict:currentDict }},
      tables: tbl, presets, computed:_computedCols, charts:_savedCharts
    }};
    localStorage.setItem(_DASH_KEY, JSON.stringify(payload));
  }} catch(e) {{}}
}}

function _restoreState() {{
  let raw; try {{ raw = localStorage.getItem(_DASH_KEY); }} catch(e) {{ return; }}
  if (!raw) return;
  let p; try {{ p = JSON.parse(raw); }} catch(e) {{ return; }}
  if (p.ui) {{
    if (p.ui.theme)  currentTheme  = p.ui.theme;
    if (p.ui.view)   currentView   = p.ui.view;
    if (p.ui.numfmt) currentNumFmt = p.ui.numfmt;
    if (p.ui.colfmt) currentColFmt = p.ui.colfmt;
    if (typeof p.ui.dict === 'number') currentDict = p.ui.dict;
    _compactMode = !!p.ui.compact;
  }}
  if (p.tables) {{
    Object.keys(p.tables).forEach(tid => {{
      const s = getState(tid); const d = p.tables[tid];
      Object.assign(s, {{ search:d.search||'', page:d.page||1, rpp:d.rpp||DEFAULT_RPP,
        sortCol:(d.sortCol==null?-1:d.sortCol), sortDir:(d.sortDir==null?0:d.sortDir), sortKeys:d.sortKeys||[],
        showPct:!!d.showPct, showRunning:!!d.showRunning, condFmt:!!d.condFmt,
        databarOn:!!d.databarOn, showSummary:d.showSummary!==false, sparkOn:!!d.sparkOn,
        groupByCols:(Array.isArray(d.groupByCols)?d.groupByCols:(typeof d.groupByCol==='number'&&d.groupByCol>=0?[d.groupByCol]:[])),
        hiddenCols:new Set(d.hiddenCols||[]) }});
      const filters = {{}};
      Object.keys(d.filters||{{}}).forEach(k => {{ const f=d.filters[k]; filters[k]= f.type==='cat'?Object.assign({{}},f,{{allowed:new Set(f.allowed)}}):Object.assign({{}},f); }});
      s.filters = filters;
    }});
  }}
  if (p.presets) {{
    Object.keys(p.presets).forEach(tid => {{
      _filterPresets[tid] = (p.presets[tid]||[]).map(pr => {{
        const fl={{}}; Object.keys(pr.filters||{{}}).forEach(k=>{{const f=pr.filters[k];fl[k]=f.type==='cat'?Object.assign({{}},f,{{allowed:new Set(f.allowed)}}):Object.assign({{}},f);}});
        return {{ name:pr.name, filters:fl }};
      }});
    }});
  }}
  if (p.computed) Object.assign(_computedCols, p.computed);
  if (p.charts)   Object.assign(_savedCharts, p.charts);
}}

function _resetDashboard() {{
  try {{ localStorage.removeItem(_DASH_KEY); }} catch(e) {{}}
  showToast('Saved layout cleared — reloading…');
  setTimeout(()=>location.reload(), 600);
}}

// ── CAVEMAN HELP / STATUS PANEL ──
const _CAVEMAN_ROWS = [
  ['📊 Chart', 'Open the chart builder under any table'],
  ['✨ Spark', 'Row sparklines — mini bar chart per row'],
  ['Σ↓', 'Running total across the FULL filtered dataset'],
  ['%', '% of total column'],
  ['🎨', 'Conditional cell colouring (top/bottom quartiles)'],
  ['💾 Preset', 'Save current filters; right-click a preset to delete'],
  ['ƒ(x)', 'Add a computed column from a formula'],
  ['↓ Filtered', 'Export only the rows currently shown (filters + sort)'],
  ['🖨 PDF', 'Print / save the current view as PDF'],
  ['Trend', 'Chart trendline: mean / linear fit / moving average'],
  ['Auto-save', 'Theme, filters, charts & layout persist across refresh'],
  ['#tab links', 'The active tab is stored in the URL — shareable'],
];
let _cavemanOpen=false;
function toggleCaveman() {{
  _cavemanOpen=!_cavemanOpen;
  let m=document.getElementById('caveman-modal');
  if(_cavemanOpen){{
    if(!m){{
      m=document.createElement('div'); m.id='caveman-modal'; m.className='shortcuts-modal';
      m.onclick=e=>{{ if(e.target===m) toggleCaveman(); }};
      const rows = _CAVEMAN_ROWS.map(r=>'<div class="shortcut-row"><span class="shortcut-key">'+_escHtml(r[0])+'</span><span class="shortcut-desc">'+_escHtml(r[1])+'</span></div>').join('');
      m.innerHTML='<div class="shortcuts-box"><div class="shortcuts-title">🪨 Caveman Guide</div>'+rows+
        '<div style="margin-top:16px;display:flex;justify-content:space-between;align-items:center;">'+
        '<button class="action-btn" style="color:#f75a7a;border-color:#f75a7a;" onclick="_resetDashboard()">↺ Reset saved layout</button>'+
        '<button class="action-btn" onclick="toggleCaveman()">Close</button></div></div>';
      document.body.appendChild(m);
    }}
    m.style.display='flex';
  }} else if(m){{ m.style.display='none'; }}
}}

// ══════════════════════════════════════════════
//  v12 — SHARED INFRASTRUCTURE for derived tables
// ══════════════════════════════════════════════

// Registry for tables produced at runtime (pivots, comparisons). findTableData() checks this first.
const _derivedTables = {{}};

// Enumerate every real table across all dicts → [{{tableId, label}}]
function listAllTables() {{
  const out = [];
  function walk(sections, di, parentTid, crumb) {{
    sections.forEach(sec => {{
      const _prefix = parentTid ? parentTid + '-' : 'd' + di + '-tbl-';
      const tid = _prefix + sec.title.normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').replace(/[^a-zA-Z0-9]+/g,'-').replace(/^-+|-+$/g,'').replace(/[^a-z0-9-]/gi,'');
      const lbl = (crumb ? crumb + ' › ' : '') + sec.title;
      if (sec.type === 1) out.push({{ tableId: tid, label: lbl }});
      else if (sec.type === 2) walk(sec.children, di, tid, lbl);
    }});
  }}
  ALL_SECTIONS.forEach((secs, di) => walk(secs, di, null, NAV_NAMES[di]));
  return out;
}}

// Aggregate a list of values by mode (sum/avg/count/max/min) with smart rounding
function _aggValues(vals, mode) {{
  if (mode === 'count') return vals.length;
  const nums = vals.map(v => parseFloat(v)).filter(v => !isNaN(v));
  if (nums.length === 0) return null;
  let raw;
  if (mode === 'avg')      raw = nums.reduce((a,b)=>a+b,0) / nums.length;
  else if (mode === 'max') raw = Math.max(...nums);
  else if (mode === 'min') raw = Math.min(...nums);
  else                     raw = nums.reduce((a,b)=>a+b,0);   // sum
  return _smartRound(raw, mode);
}}

// Build a full table-data object (matching _df_to_json_rows shape) from cols + rows.
// Keys numeric_cols / bar_ranges / col_totals by FLAT column index (what buildTable reads).
function _makeTableData(cols, rows) {{
  const numeric_cols = [];
  const bar_ranges = {{}};
  const col_totals = {{}};
  const summary = [];
  for (let i = 1; i < cols.length; i++) {{
    const vals = rows.map(r => parseFloat(r[i])).filter(v => !isNaN(v));
    const isNum = rows.length > 0 && vals.length >= rows.length * 0.6;
    if (isNum) {{
      numeric_cols.push(i);
      bar_ranges[String(i)] = vals.reduce((m,v)=>Math.max(m,Math.abs(v)),0) || 1;
      const tot = vals.reduce((a,b)=>a+b,0);
      col_totals[String(i)] = tot !== 0 ? tot : 1;
      summary.push(Math.round(tot*100)/100);
    }} else {{
      summary.push(null);
    }}
  }}
  return {{ cols, rows, summary, numeric_cols, bar_ranges, col_totals,
           cf_cols:[], cf_ranges:{{}}, sp_col:null, description:'' }};
}}

// ══════════════════════════════════════════════
//  v12 — PIVOT TABLE BUILDER
// ══════════════════════════════════════════════

function openPivotBuilder(tableId) {{
  let panel = document.getElementById('pivot-panel-' + tableId);
  if (panel) {{ panel.style.display = (panel.style.display === 'none') ? '' : 'none'; return; }}
  const data = findTableData(tableId);
  if (!data) return;
  const cont = document.getElementById('tblcontainer-' + tableId);
  if (!cont) return;
  panel = document.createElement('div');
  panel.id = 'pivot-panel-' + tableId;
  panel.className = 'pivot-panel';
  panel.innerHTML = _buildPivotBuilderHTML(tableId, data);
  cont.parentElement.appendChild(panel);
  panel.scrollIntoView({{behavior:'smooth', block:'nearest'}});
}}

function _buildPivotBuilderHTML(tableId, data) {{
  const cols = data.cols;
  const allOpts = cols.map((c,i) => '<option value="'+i+'">'+_escHtml(String(c))+'</option>').join('');
  const numOpts = cols.map((c,i) => (data.numeric_cols && data.numeric_cols.includes(i))
      ? '<option value="'+i+'">'+_escHtml(String(c))+'</option>' : '').join('');
  // First row dimension pre-selected for a sensible default
  const rowOpts = cols.map((c,i) => '<option value="'+i+'"'+(i===0?' selected':'')+'>'+_escHtml(String(c))+'</option>').join('');
  return '<div class="pivot-builder">'
    + '<div class="pivot-builder-title">⊞ Pivot Table <span style="font-size:10px;color:var(--text3);font-weight:400;">— respects current filters · Ctrl/Cmd-click to pick multiple</span></div>'
    + '<div class="pivot-controls">'
    +   '<div class="pivot-field"><label>Rows (dimension)</label><select class="chart-opt-select pivot-multi" multiple size="5" id="pivot-row-'+tableId+'">'+rowOpts+'</select></div>'
    +   '<div class="pivot-field"><label>Columns (optional)</label><select class="chart-opt-select pivot-multi" multiple size="5" id="pivot-col-'+tableId+'">'+allOpts+'</select></div>'
    +   '<div class="pivot-field"><label>Value (metric)</label><select class="chart-opt-select" id="pivot-val-'+tableId+'">'+numOpts+'</select></div>'
    +   '<div class="pivot-field"><label>Aggregate</label><select class="chart-opt-select" id="pivot-agg-'+tableId+'"><option value="sum">Sum</option><option value="avg">Avg</option><option value="count">Count</option><option value="max">Max</option><option value="min">Min</option></select></div>'
    +   '<button class="chart-generate-btn" onclick="generatePivot(\\''+tableId+'\\')">▶ Build Pivot</button>'
    +   '<button class="chart-opt-toggle" onclick="openPivotBuilder(\\''+tableId+'\\')">Hide</button>'
    + '</div>'
    + '<div id="pivot-out-'+tableId+'" class="pivot-out"></div>'
    + '</div>';
}}

// Multi-field pivot. rowIdxs / colIdxs are arrays of column indices.
// Each row field becomes its own leading output column; column fields are
// combined into composite headers ("A | B").
function _computePivot(srcRows, rowIdxs, colIdxs, valIdx, aggMode, cols) {{
  const SEP = '\\u0001';
  const rowNames = rowIdxs.map(i => String(cols[i]));
  const valName  = String(cols[valIdx]);
  const rowKeyOf  = r => rowIdxs.map(i => String(r[i])).join(SEP);
  const rowValsOf = r => rowIdxs.map(i => String(r[i]));
  if (!colIdxs || colIdxs.length === 0) {{
    const groups = {{}}, order = [], rowVals = {{}};
    srcRows.forEach(r => {{ const k = rowKeyOf(r); if(!(k in groups)){{ groups[k]=[]; order.push(k); rowVals[k]=rowValsOf(r); }} groups[k].push(r[valIdx]); }});
    const outCols = rowNames.concat([aggMode + '(' + valName + ')']);
    const outRows = order.map(k => rowVals[k].concat([_aggValues(groups[k], aggMode)]));
    return _makeTableData(outCols, outRows);
  }} else {{
    const colKeyOf = r => colIdxs.map(i => String(r[i])).join(' | ');
    const colKeys = [], seenC = {{}};
    srcRows.forEach(r => {{ const c = colKeyOf(r); if(!(c in seenC)){{ seenC[c]=1; colKeys.push(c); }} }});
    colKeys.sort(_chronoCmp);
    const groups = {{}}, order = [], rowVals = {{}};
    srcRows.forEach(r => {{
      const k = rowKeyOf(r); if(!(k in groups)){{ groups[k]={{}}; order.push(k); rowVals[k]=rowValsOf(r); }}
      const c = colKeyOf(r); (groups[k][c] = groups[k][c] || []).push(r[valIdx]);
    }});
    const outCols = rowNames.concat(colKeys);
    const outRows = order.map(k => rowVals[k].concat(colKeys.map(c => groups[k][c] ? _aggValues(groups[k][c], aggMode) : null)));
    return _makeTableData(outCols, outRows);
  }}
}}

function generatePivot(tableId) {{
  const data = findTableData(tableId);
  if (!data) return;
  const _selIdxs = sel => sel ? Array.from(sel.selectedOptions).map(o => parseInt(o.value)).filter(v => !isNaN(v) && v >= 0) : [];
  const rowIdxs = _selIdxs(document.getElementById('pivot-row-'+tableId));
  const colIdxs = _selIdxs(document.getElementById('pivot-col-'+tableId));
  const valIdx  = parseInt((document.getElementById('pivot-val-'+tableId)||{{}}).value);
  const agg     = (document.getElementById('pivot-agg-'+tableId)||{{}}).value || 'sum';
  if (rowIdxs.length === 0 || isNaN(valIdx)) {{ showToast('Pick at least one Row dimension and a Value column'); return; }}
  if (rowIdxs.includes(valIdx) || colIdxs.includes(valIdx)) {{ showToast('Value column must differ from Row/Column fields'); return; }}
  // Respect the source table's current filters + search
  let rows = applyFilters(tableId, data.rows);
  const st = getState(tableId);
  if (st.search) {{ const q = st.search.toLowerCase(); rows = rows.filter(r => r.some(c => String(c).toLowerCase().includes(q))); }}
  rows = rows.filter(r => !isGrandTotal(r[rowIdxs[0]]));
  if (rows.length === 0) {{ showToast('No rows to pivot (check filters)'); return; }}
  const pdata = _computePivot(rows, rowIdxs, colIdxs, valIdx, agg, data.cols);
  const ptid  = tableId + '-pivot';
  _derivedTables[ptid] = pdata;
  delete tableState[ptid];   // fresh state for the regenerated pivot
  const out = document.getElementById('pivot-out-' + tableId);
  if (out) out.innerHTML = '<div class="pivot-result-label">⊞ Pivot result ('+pdata.rows.length+' rows)</div>' + buildTable(pdata, ptid, false);
}}

// ══════════════════════════════════════════════
//  v12 — CROSS-TABLE / PERIOD COMPARISON
// ══════════════════════════════════════════════
// Compares a chosen metric across up to 4 tables ("periods"), joined on a key
// dimension. Reuses listAllTables / _aggValues / _makeTableData / buildTable.

const CMP_SLOTS = 4;

function openCompareBuilder() {{
  const ov = document.getElementById('cmp-overlay');
  if (!ov) return;
  const tables = listAllTables();
  if (tables.length === 0) {{ showToast('No tables to compare'); return; }}
  const tblOpts = '<option value="">— pick a table —</option>' +
    tables.map(t => '<option value="'+_escHtml(t.tableId)+'">'+_escHtml(t.label)+'</option>').join('');
  let slots = '';
  for (let s = 0; s < CMP_SLOTS; s++) {{
    slots +=
      '<div class="cmp-slot">'
      + '<div class="pivot-field"><label>Table '+(s+1)+(s<2?' *':'')+'</label>'
      +   '<select class="chart-opt-select" id="cmp-tbl-'+s+'" onchange="cmpSlotTableChanged('+s+')">'+tblOpts+'</select></div>'
      + '<div class="pivot-field"><label>Key (dimension)</label>'
      +   '<select class="chart-opt-select" id="cmp-key-'+s+'"><option value="-1">—</option></select></div>'
      + '<div class="pivot-field"><label>Metric</label>'
      +   '<select class="chart-opt-select" id="cmp-val-'+s+'"><option value="-1">—</option></select></div>'
      + '<div class="pivot-field"><label>Series label</label>'
      +   '<input class="chart-opt-select" id="cmp-lbl-'+s+'" placeholder="(auto)" style="min-width:120px;"></div>'
      + '</div>';
  }}
  ov.innerHTML =
    '<div class="cmp-modal">'
    + '<div class="cmp-modal-head"><span class="cmp-modal-title">⇄ Compare across tables / periods</span>'
    +   '<button class="action-btn" onclick="closeCompareBuilder()">✕ Close</button></div>'
    + '<div style="font-size:11px;color:var(--text3);margin-bottom:10px;">Pick 2+ tables, a key column to align on, and a metric. Values are aggregated per key and shown side by side.</div>'
    + slots
    + '<div class="cmp-actions">'
    +   '<div class="pivot-field"><label>Aggregate</label><select class="chart-opt-select" id="cmp-agg"><option value="sum">Sum</option><option value="avg">Avg</option><option value="count">Count</option><option value="max">Max</option><option value="min">Min</option></select></div>'
    +   '<button class="chart-generate-btn" onclick="generateComparison()">▶ Build Comparison</button>'
    +   '<button class="action-btn" onclick="buildComparisonChart()" title="Chart the comparison result">📊 Chart it</button>'
    + '</div>'
    + '<div id="cmp-out" class="cmp-out"></div>'
    + '</div>';
  ov.classList.add('open');
}}

function closeCompareBuilder() {{
  const ov = document.getElementById('cmp-overlay');
  if (ov) ov.classList.remove('open');
}}

// When a slot's table changes, repopulate its Key + Metric selects from that table's columns.
function cmpSlotTableChanged(s) {{
  const tid = (document.getElementById('cmp-tbl-'+s)||{{}}).value;
  const keySel = document.getElementById('cmp-key-'+s);
  const valSel = document.getElementById('cmp-val-'+s);
  if (!keySel || !valSel) return;
  if (!tid) {{ keySel.innerHTML = '<option value="-1">—</option>'; valSel.innerHTML = '<option value="-1">—</option>'; return; }}
  const data = findTableData(tid);
  if (!data) return;
  let keyOpts = '', valOpts = '';
  data.cols.forEach((c, i) => {{
    keyOpts += '<option value="'+i+'">'+_escHtml(String(c))+'</option>';
    if (data.numeric_cols && data.numeric_cols.includes(i)) valOpts += '<option value="'+i+'">'+_escHtml(String(c))+'</option>';
  }});
  keySel.innerHTML = keyOpts;
  valSel.innerHTML = valOpts || '<option value="-1">—</option>';
}}

// Read the configured slots into [{{tid,label,keyIdx,valIdx,data}}]
function _cmpReadSlots() {{
  const slots = [];
  for (let s = 0; s < CMP_SLOTS; s++) {{
    const tid = (document.getElementById('cmp-tbl-'+s)||{{}}).value;
    if (!tid) continue;
    const keyIdx = parseInt((document.getElementById('cmp-key-'+s)||{{}}).value);
    const valIdx = parseInt((document.getElementById('cmp-val-'+s)||{{}}).value);
    if (isNaN(keyIdx) || isNaN(valIdx) || keyIdx < 0 || valIdx < 0) continue;
    const data = findTableData(tid);
    if (!data) continue;
    const custom = (document.getElementById('cmp-lbl-'+s)||{{}}).value;
    const auto = String(data.cols[valIdx]);
    slots.push({{ tid, keyIdx, valIdx, data, label: (custom && custom.trim()) ? custom.trim() : auto }});
  }}
  return slots;
}}

// Build a merged comparison table object (cached as _derivedTables['cmp-result']).
function _cmpBuildData() {{
  const slots = _cmpReadSlots();
  if (slots.length < 2) {{ showToast('Pick at least 2 tables with a key + metric'); return null; }}
  const agg = (document.getElementById('cmp-agg')||{{}}).value || 'sum';
  // Per-slot: group metric values by key string → aggregate
  const keyOrder = [], seenKey = {{}};
  const slotMaps = slots.map(sl => {{
    const groups = {{}};
    sl.data.rows.forEach(r => {{
      if (isGrandTotal(r[sl.keyIdx])) return;
      const k = String(r[sl.keyIdx]);
      (groups[k] = groups[k] || []).push(r[sl.valIdx]);
      if (!(k in seenKey)) {{ seenKey[k] = 1; keyOrder.push(k); }}
    }});
    const out = {{}};
    Object.keys(groups).forEach(k => {{ out[k] = _aggValues(groups[k], agg); }});
    return out;
  }});
  keyOrder.sort(_chronoCmp);
  // De-duplicate identical series labels (append #2, #3…)
  const seenLbl = {{}};
  const labels = slots.map(sl => {{
    let base = sl.label; let lbl = base; let n = 2;
    while (lbl in seenLbl) {{ lbl = base + ' #' + n; n++; }}
    seenLbl[lbl] = 1; return lbl;
  }});
  const cols = ['Key'].concat(labels);
  const rows = keyOrder.map(k => [k].concat(slotMaps.map(m => (k in m) ? m[k] : null)));
  const pdata = _makeTableData(cols, rows);
  pdata.description = 'Comparison of ' + agg + '() across ' + slots.length + ' tables, aligned on key column.';
  return pdata;
}}

function generateComparison() {{
  const pdata = _cmpBuildData();
  if (!pdata) return;
  _derivedTables['cmp-result'] = pdata;
  delete tableState['cmp-result'];
  const out = document.getElementById('cmp-out');
  if (out) out.innerHTML = '<div class="pivot-result-label">⇄ Comparison ('+pdata.rows.length+' keys)</div>'
    + '<div id="tblcontainer-cmp-result">' + buildTable(pdata, 'cmp-result', false) + '</div>';
}}

// Build the comparison then open the dual/standard chart builder on the result table.
function buildComparisonChart() {{
  const pdata = _cmpBuildData();
  if (!pdata) return;
  _derivedTables['cmp-result'] = pdata;
  delete tableState['cmp-result'];
  const out = document.getElementById('cmp-out');
  if (out) {{
    out.innerHTML = '<div class="pivot-result-label">⇄ Comparison ('+pdata.rows.length+' keys)</div>'
      + '<div id="tblcontainer-cmp-result">' + buildTable(pdata, 'cmp-result', false) + '</div>';
  }}
  // Reuse the standard chart builder on the derived table
  if (typeof openChartBuilder === 'function') openChartBuilder('cmp-result');
}}

// ── INIT ──
_restoreState();
// Deep-link: #tabN in URL overrides the saved tab
try {{
  const _hm = (location.hash||'').match(/^#tab(\\d+)$/);
  if (_hm) {{ const _ti=parseInt(_hm[1]); if(_ti>=0 && _ti<ALL_SECTIONS.length) currentDict=_ti; }}
}} catch(e) {{}}
if (currentDict < 0 || currentDict >= ALL_SECTIONS.length) currentDict = 0;

setTheme(currentTheme);
buildNav();
document.querySelectorAll('.nav-pill').forEach((p,idx)=>p.classList.toggle('active', idx===currentDict));
buildSidebar();
renderBody();

// Reflect restored UI state in the toolbar + cells
document.getElementById('btn-expand').classList.toggle('active',   currentView==='expand');
document.getElementById('btn-collapse').classList.toggle('active', currentView==='collapse');
(function(){{ const b=document.getElementById('compact-btn'); if(b) b.classList.toggle('active', _compactMode); }})();
(function(){{ const b=document.getElementById('darklight-btn'); if(b) b.textContent = (currentTheme==='Slate Light') ? '🌙' : '☀'; }})();
setNumFmt(currentNumFmt);
setColFmt(currentColFmt);

// Add the floating Caveman help button
(function(){{
  const fab=document.createElement('button');
  fab.className='caveman-fab'; fab.title='Caveman guide'; fab.textContent='🪨';
  fab.onclick=toggleCaveman; document.body.appendChild(fab);
}})();

// Persist on tab hide + before unload
window.addEventListener('beforeunload', _persistState);
document.addEventListener('visibilitychange', ()=>{{ if(document.hidden) _persistState(); }});
window.addEventListener('hashchange', ()=>{{
  const m=(location.hash||'').match(/^#tab(\\d+)$/);
  if(m){{ const ti=parseInt(m[1]); if(ti!==currentDict && ti>=0 && ti<ALL_SECTIONS.length) switchDict(ti); }}
}});
</script>
</body>
</html>"""

    with open(output, "w", encoding="utf-8") as f:
        f.write(html)

    abs_path = os.path.abspath(output)
    print(f"✅ Dashboard saved → {abs_path}")
    webbrowser.open(f"file://{abs_path}")


# ─────────────────────────────────────────────
#  PLOTLY SETUP — download once for offline use
# ─────────────────────────────────────────────

def setup_plotly(output_dir="."):
    """
    Download Plotly JS once and save it locally for offline use.
    Place plotly.min.js in the same folder as your dashboard.html.

    Run this ONCE:
        setup_plotly()   # saves plotly.min.js in current directory
    """
    import urllib.request
    path = os.path.join(output_dir, "plotly.min.js")
    if os.path.exists(path):
        print(f"✅ Plotly already available at: {path}")
        return path
    print("⏳ Downloading Plotly JS (~3.5MB) — one time only...")
    urllib.request.urlretrieve(
        "https://cdn.plot.ly/plotly-2.27.0.min.js",
        path
    )
    print(f"✅ Plotly saved → {path}")
    print("   Keep plotly.min.js in the same folder as dashboard.html for offline use!")
    return path


# ─────────────────────────────────────────────
#  PICKLE HELPER
# ─────────────────────────────────────────────

def load_and_generate(
    pkl_path: str,
    output: str = "dashboard.html",
    default_theme: str = "Dark Blue",
    rows_per_page: int = 25,
    sparkline_col: str = None,
    cf_cols: list = None,
):
    """
    Load a pkl file and generate dashboard in one line.

    pkl must contain:
        {{
            "Payment Summary": dict1,
            "Bank Wise":       dict2,
        }}
    Keys become navbar names automatically.
    """
    import pickle
    with open(pkl_path, "rb") as f:
        data = pickle.load(f)
    assert isinstance(data, dict), "pkl must contain a dict at top level!"
    generate_dashboard(
        dicts          = list(data.values()),
        names          = list(data.keys()),
        output         = output,
        default_theme  = default_theme,
        rows_per_page  = rows_per_page,
        sparkline_col  = sparkline_col,
        cf_cols        = cf_cols,
    )


# ─────────────────────────────────────────────
#  USAGE EXAMPLES
# ─────────────────────────────────────────────

# ── Option 1: Direct ──
# generate_dashboard(
#     dicts         = [dict1, dict2, dict3, dict4],
#     names         = ["Payment Summary", "Bank Wise", "PG Wise", "Merchant Wise"],
#     output        = "dashboard.html",
#     default_theme = "Dark Blue",
#     rows_per_page = 25,
#     cf_cols       = ["Success %", "Failed"],   # highlight high/low values
# )

# ── Option 2: From pkl ──
# load_and_generate(
#     pkl_path      = "dashboard_data.pkl",
#     output        = "dashboard.html",
#     default_theme = "Dark Blue",
#     rows_per_page = 25,
# )
